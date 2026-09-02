"""口碑巡检独立领域、合成验收运行与交付物测试。"""

from __future__ import annotations

import asyncio
import csv
import hashlib
import json
import tempfile
import unittest
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from patchright.async_api import Error as PlaywrightError
from PIL import Image
from sqlalchemy import func, select

from threadsnap.app import create_app
from threadsnap.config import Settings
from threadsnap.models import (
    ExtractionRun,
    PlatformConfig,
    ReputationResult,
    ReputationRun,
    ReputationScheduleEvent,
    ReputationScopeDraft,
    ReputationScopeVersion,
    ReputationTombstone,
)
from threadsnap.reputation_autohome import normalize_series_url as normalize_autohome_url
from threadsnap.reputation_dongchedi import (
    DongchediReputationAdapter,
    ReputationAdapterError,
    ReputationMappingTarget,
    ReputationPageResult,
    _metric_rect,
    normalize_series_url,
)
from threadsnap.reputation_scheduler import ReputationCoordinator
from threadsnap.reputation_yiche import normalize_series_url as normalize_yiche_url


class OfficialFakeAdapter:
    """支持证据策略和可控失败的正式巡检确定性适配器。"""

    score_overrides: dict[str, str] = {}
    negative_rate_overrides: dict[str, str] = {}
    failures: set[str] = set()
    retry_once: set[str] = set()
    progress_probe = None
    last_prefer_http_first: bool | None = None
    last_include_negative_rate: bool | None = None
    last_concurrency: int | None = None

    def __init__(
        self,
        *_args,
        evidence_policy=None,
        prefer_http_first: bool = False,
        include_negative_rate: bool = False,
        **_kwargs,
    ):
        self.evidence_policy = evidence_policy
        self.prefer_http_first = prefer_http_first
        self.validation_calls = 0
        type(self).last_prefer_http_first = prefer_http_first
        type(self).last_include_negative_rate = include_negative_rate
        type(self).last_concurrency = int(_kwargs["concurrency"])

    def validate_sync(self, targets, output_dir, on_result=None):
        self.validation_calls += 1
        output_dir.mkdir(parents=True)
        values = []
        for index, target in enumerate(targets):
            if target.vehicle_id in self.retry_once and self.validation_calls == 1:
                result = ReputationAdapterError(
                    "REPUTATION_PAGE_FIXTURE_RETRY",
                    "确定性正式巡检暂时失败。",
                    retryable=True,
                )
            elif target.vehicle_id in self.failures:
                result = ReputationAdapterError(
                    "REPUTATION_PAGE_FIXTURE_FAILURE", "确定性正式巡检失败。"
                )
            else:
                score = self.score_overrides.get(target.vehicle_id, "3.80")
                negative_rate = self.negative_rate_overrides.get(target.vehicle_id, "37%")
                measurement = {
                    "score_raw": score,
                    "rank_raw": "4",
                    "volume_raw": str(500 + index),
                    "rank_scope": "同级车评分",
                }
                capture = self.evidence_policy is None or self.evidence_policy(
                    target, measurement
                )
                metric = None
                digest = None
                if capture:
                    target_dir = output_dir / target.vehicle_id
                    target_dir.mkdir()
                    metric = target_dir / "region.png"
                    Image.new("RGB", (600, 240), "white").save(metric)
                    digest = hashlib.sha256(metric.read_bytes()).hexdigest()
                result = ReputationPageResult(
                    vehicle_id=target.vehicle_id,
                    platform_vehicle_id=target.platform_vehicle_id,
                    mapping_hash=target.mapping_hash,
                    final_url=target.platform_url,
                    actual_name=target.platform_display_name,
                    score_raw=score,
                    rank_raw="4",
                    volume_raw=str(500 + index),
                    review_article_count_raw=str(5000 + index),
                    review_article_count_url=target.platform_url,
                    rank_scope="同级车评分",
                    measurements=[measurement] * 3,
                    full_page_path=metric,
                    metric_region_path=metric,
                    full_page_sha256=digest,
                    metric_region_sha256=digest,
                    width=600,
                    height=240,
                    metric_rect={"x": 0, "y": 0, "width": 600, "height": 240},
                    duration_ms=25,
                    negative_rate_raw=negative_rate,
                    negative_rate_url=(
                        "https://api.dcarapi.com/motor/car_score/api/v1/landing_page/"
                        f"get_detail/?series_id={target.platform_vehicle_id}"
                    ),
                    negative_rate_positive_count=214,
                    negative_rate_negative_count=128,
                )
            values.append(result)
            if on_result:
                on_result(index, target, result)
            if type(self).progress_probe:
                type(self).progress_probe()
        return values


class ReputationInspectionTest(unittest.TestCase):
    """使用三因子隔离测试配置验证口碑巡检完整纵切。"""

    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.settings = Settings(
            database_url=f"sqlite:///{(self.root / 'test.db').as_posix()}",
            data_dir=self.root / "data",
            start_background_services=False,
            runtime_mode="test",
            enable_reputation_synthetic_runs=True,
            reputation_test_database=True,
        )
        self.client_context = TestClient(create_app(self.settings))
        self.client = self.client_context.__enter__()

    def tearDown(self) -> None:
        self.client_context.__exit__(None, None, None)
        self.temporary.cleanup()

    def create_run(self, scenario_id: str) -> dict:
        """触发一个固定场景并返回运行详情。"""

        response = self.client.post(
            "/api/v1/reputation/test-runs",
            json={"scenario_id": scenario_id},
        )
        self.assertEqual(response.status_code, 202, response.text)
        return response.json()

    def test_three_platform_registry_mapping_validation_and_publish_denominator(self) -> None:
        """汽车之家和易车复用懂车帝映射验证，并把发布门扩为27乘3。"""

        scope = self.initialize_scope("three-platform-scope.csv")
        container = self.client.app.state.container
        platform_rows = {
            "autohome": [
                {
                    "vehicle_id": vehicle["id"],
                    "platform_vehicle_id": str(7000 + index),
                    "platform_url": f"https://k.autohome.com.cn/{7000 + index}/",
                    "platform_display_name": vehicle["vehicle_name"],
                }
                for index, vehicle in enumerate(scope["vehicles"])
            ],
            "yiche": [
                {
                    "vehicle_id": vehicle["id"],
                    "platform_vehicle_id": str(9000 + index),
                    "platform_url": f"https://dianping.yiche.com/fixture{index}/koubei/",
                    "platform_display_name": vehicle["vehicle_name"],
                }
                for index, vehicle in enumerate(scope["vehicles"])
            ],
        }
        for platform_code, rows in platform_rows.items():
            saved = self.client.put(
                "/api/v1/reputation/scope/mappings",
                json={
                    "revision": scope["revision"],
                    "platform_code": platform_code,
                    "rows": rows,
                },
            )
            self.assertEqual(200, saved.status_code, saved.text)
            scope = saved.json()
            container.session_store.import_state(
                platform_code,
                {"cookies": [{"name": "fixture", "value": "session", "domain": ".example.com", "path": "/"}]},
            )
            container.reputation.adapter_factories[platform_code] = OfficialFakeAdapter
            validated = self.client.post(
                "/api/v1/reputation/scope/mapping-validations",
                json={"revision": scope["revision"], "platform_code": platform_code},
            )
            self.assertEqual(200, validated.status_code, validated.text)
            payload = validated.json()
            self.assertEqual(27, payload["succeeded_count"])
            self.assertEqual(platform_code, payload["platform_code"])
            scope = payload["scope"]

        preview = self.client.get("/api/v1/reputation/scope/publish-preview")
        self.assertEqual(200, preview.status_code, preview.text)
        self.assertEqual(81, preview.json()["expected_mapping_count"])
        self.assertEqual(54, preview.json()["verified_mapping_count"])
        self.assertFalse(preview.json()["can_publish"])

        container.session_store.import_state(
            "dongchedi",
            {"cookies": [{"name": "fixture", "value": "session", "domain": ".example.com", "path": "/"}]},
        )
        container.reputation.adapter_factory = OfficialFakeAdapter
        container.reputation.adapter_factories["dongchedi"] = OfficialFakeAdapter
        validated = self.client.post(
            "/api/v1/reputation/scope/mapping-validations",
            json={"revision": scope["revision"], "platform_code": "dongchedi"},
        )
        self.assertEqual(200, validated.status_code, validated.text)
        scope = validated.json()["scope"]
        preview = self.client.get("/api/v1/reputation/scope/publish-preview").json()
        self.assertEqual(81, preview["verified_mapping_count"])
        self.assertTrue(preview["can_publish"])
        published = self.client.post(
            "/api/v1/reputation/scope/publish",
            json={
                "revision": scope["revision"],
                "initial_review_acknowledged": True,
            },
        )
        self.assertEqual(200, published.status_code, published.text)
        with container.sessions.begin() as db:
            for platform_code in ("dongchedi", "autohome", "yiche"):
                db.get(PlatformConfig, platform_code).enabled = True

        due = container.reputation.check_schedule(
            datetime(2030, 1, 2, 2, 0, tzinfo=timezone.utc)
        )
        self.assertEqual(1, len(due["queued_run_ids"]))
        finished = container.reputation.execute_run(due["queued_run_ids"][0])
        self.assertEqual(["dongchedi", "autohome", "yiche"], finished["platform_codes"])
        self.assertEqual(81, finished["planned_count"])
        self.assertEqual(81, finished["completed_count"])
        self.assertEqual(81, finished["complete_evidence_count"])
        self.assertEqual(2, finished["concurrency"])
        self.assertEqual(2, OfficialFakeAdapter.last_concurrency)
        self.assertEqual(
            81,
            len(
                {
                    (item["vehicle_id"], item["platform_code"])
                    for item in finished["results"]
                }
            ),
        )
        generated = container.reputation.generate_report(
            finished["id"], datetime(2030, 1, 2, 2, 1, tzinfo=timezone.utc)
        )
        xlsx_response = self.client.get(generated["downloads"]["xlsx"])
        self.assertEqual(200, xlsx_response.status_code)
        xlsx_path = self.root / "three-platform.xlsx"
        xlsx_path.write_bytes(xlsx_response.content)
        sheet = load_workbook(xlsx_path)["口碑巡检"]
        self.assertEqual((28, 20), (sheet.max_row, sheet.max_column))
        self.assertEqual("懂车帝-口碑分", sheet["E1"].value)
        self.assertEqual("汽车之家-口碑分", sheet["J1"].value)
        self.assertEqual("易车-口碑分", sheet["O1"].value)
        self.assertEqual("备注", sheet["T1"].value)
        self.assertEqual(27, len(sheet._images))
        preview_manifest = (
            self.settings.reputation_dir
            / finished["id"]
            / "xlsx-previews"
            / "manifest.json"
        )
        self.assertEqual(
            27,
            len(json.loads(preview_manifest.read_text(encoding="utf-8"))["items"]),
        )

    def test_later_platform_url_contracts(self) -> None:
        """两个后续平台都规范到已验证的车型口碑入口。"""

        self.assertEqual(
            "https://k.autohome.com.cn/6664/",
            normalize_autohome_url("https://k.autohome.com.cn/6664", "6664"),
        )
        self.assertEqual(
            "https://dianping.yiche.com/ruihu8/koubei/",
            normalize_yiche_url("https://car.yiche.com/ruihu8/", "5313"),
        )
        with self.assertRaises(ReputationAdapterError):
            normalize_autohome_url("https://k.autohome.com.cn/6664/", "5313")

    def test_browser_runtime_error_is_logged_and_marked_retryable(self) -> None:
        """Patchright通用异常必须保留服务端阶段诊断并标记为可重试。"""

        class BrokenBrowser:
            @staticmethod
            async def new_context(**_kwargs):
                raise PlaywrightError("page context closed during navigation")

        adapter = DongchediReputationAdapter(None)
        target = ReputationMappingTarget(
            vehicle_id="vehicle-browser-error",
            platform_vehicle_id="6227",
            platform_url="https://www.dongchedi.com/auto/series/score/6227-x-x-x-x-x",
            platform_display_name="瑞虎9",
            mapping_hash="fixture",
        )

        with self.assertLogs("threadsnap.reputation_dongchedi", level="WARNING") as logs:
            with self.assertRaises(ReputationAdapterError) as raised:
                asyncio.run(
                    adapter._visit(  # noqa: SLF001 - 显式验证浏览器异常分类边界。
                        BrokenBrowser(),  # type: ignore[arg-type]
                        target,
                        self.root / "browser-error",
                    )
                )

        self.assertEqual("REPUTATION_BROWSER_RUNTIME_ERROR", raised.exception.code)
        self.assertTrue(raised.exception.retryable)
        self.assertNotIn("page context closed", raised.exception.message)
        joined = "\n".join(logs.output)
        self.assertIn("stage=创建页面上下文", joined)
        self.assertIn("page context closed during navigation", joined)

    def initialize_scope(self, filename: str = "scope.csv") -> dict:
        """写入固定 27 车型初始化清单并返回当前范围。"""

        csv_path = self.root / filename
        headers = [
            "schema_version",
            "seed_key",
            "series_name",
            "vehicle_name",
            "role",
            "role_order",
            "platform_code",
            "platform_vehicle_id",
            "platform_url",
            "platform_display_name",
        ]
        with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            for index in range(27):
                focus = index < 14
                order = index + 1 if focus else index - 13
                writer.writerow(
                    {
                        "schema_version": "reputation-scope-v1",
                        "seed_key": f"vehicle-{index + 1:02d}",
                        "series_name": f"车系{index // 5 + 1}",
                        "vehicle_name": f"车型{index + 1:02d}",
                        "role": "focus" if focus else "competitor",
                        "role_order": order,
                        "platform_code": "dongchedi",
                        "platform_vehicle_id": str(10000 + index),
                        "platform_url": (
                            f"https://www.dongchedi.com/auto/series/score/{10000 + index}-x-x-x-x-x"
                        ),
                        "platform_display_name": f"页面车型{index + 1:02d}",
                    }
                )
        return self.client.app.state.container.reputation.initialize_scope_csv(csv_path)

    def test_synthetic_capability_requires_all_three_guards(self) -> None:
        capability = self.client.get("/api/v1/reputation/capabilities").json()
        self.assertTrue(capability["reputation_synthetic_runs"])
        self.assertEqual(len(capability["scenarios"]), 3)
        self.assertEqual(capability["real_adapter_status"], "available")

        disabled = Settings(
            database_url=f"sqlite:///{(self.root / 'production.db').as_posix()}",
            data_dir=self.root / "production-data",
            start_background_services=False,
            runtime_mode="production",
            enable_reputation_synthetic_runs=True,
            reputation_test_database=True,
        )
        with TestClient(create_app(disabled)) as client:
            self.assertFalse(
                client.get("/api/v1/reputation/capabilities").json()["reputation_synthetic_runs"]
            )
            response = client.post(
                "/api/v1/reputation/test-runs",
                json={"scenario_id": "daily_mixed_changes"},
            )
            self.assertEqual(response.status_code, 404)

    def test_three_scenarios_preserve_counts_colors_and_extraction_domain(self) -> None:
        baseline = self.create_run("baseline_initialization")
        daily = self.create_run("daily_mixed_changes")
        month_end = self.create_run("month_end_mixed_changes")

        self.assertEqual(len(baseline["results"]), 27)
        self.assertEqual(baseline["required_evidence_count"], 27)
        self.assertEqual(daily["required_evidence_count"], 27)
        self.assertEqual(month_end["required_evidence_count"], 27)
        self.assertTrue(all(item["evidence_required"] for item in daily["results"]))
        self.assertTrue(all(item["evidence"] for item in daily["results"]))
        self.assertEqual(daily["status"], "partial_success")
        tones = {
            metric["tone"] for result in daily["results"] for metric in result["metrics"].values()
        }
        self.assertEqual(tones, {"positive", "negative", "neutral"})
        self.assertEqual(
            daily["results"][0]["metrics"]["rank"]["tone"],
            "positive",
            "排名数值下降代表名次上升，应按正向变化着色。",
        )

        with self.client.app.state.container.sessions() as db:
            extraction_count = db.scalar(select(func.count()).select_from(ExtractionRun))
        self.assertEqual(extraction_count, 0)

    def test_txt_xlsx_and_evidence_zip_are_real_traceable_artifacts(self) -> None:
        run = self.create_run("daily_mixed_changes")
        report = self.client.get(run["downloads"]["txt"])
        self.assertEqual(report.status_code, 200)
        self.assertIn("口碑巡检指标变动如下", report.content.decode("utf-8"))
        self.assertIn("差评率", report.content.decode("utf-8"))
        self.assertNotIn("#E2F0D9", report.content.decode("utf-8"))

        xlsx = self.client.get(run["downloads"]["xlsx"])
        xlsx_path = self.root / "result.xlsx"
        xlsx_path.write_bytes(xlsx.content)
        workbook = load_workbook(xlsx_path)
        sheet = workbook["口碑巡检"]
        self.assertEqual(sheet.max_row, 28)
        fills = {
            sheet.cell(row, column).fill.fgColor.rgb
            for row in range(2, 29)
            for column in (5, 6, 7, 8, 9)
        }
        self.assertEqual(sheet["H1"].value, "口碑评价篇数")
        self.assertEqual(sheet["I1"].value, "差评率")
        self.assertTrue(any(str(value).endswith("E2F0D9") for value in fills))
        self.assertTrue(any(str(value).endswith("F4CCCC") for value in fills))
        circle_fills = {sheet.cell(row, 8).fill.fgColor.rgb for row in range(2, 29)}
        self.assertTrue(any(str(value).endswith("E2F0D9") for value in circle_fills))
        self.assertTrue(any(str(value).endswith("F4CCCC") for value in circle_fills))
        negative_rate_fills = {sheet.cell(row, 9).fill.fgColor.rgb for row in range(2, 29)}
        self.assertTrue(any(str(value).endswith("E2F0D9") for value in negative_rate_fills))
        self.assertTrue(any(str(value).endswith("F4CCCC") for value in negative_rate_fills))
        self.assertGreater(len(sheet._images), 0)

        archive = self.client.get(run["downloads"]["evidence_zip"])
        zip_path = self.root / "evidence.zip"
        zip_path.write_bytes(archive.content)
        with zipfile.ZipFile(zip_path) as bundle:
            manifest = json.loads(bundle.read("manifest.json"))
            checksums = bundle.read("SHA256SUMS").decode("utf-8").splitlines()
            self.assertEqual(len(manifest["items"]), 27)
            self.assertEqual(len(checksums), 27)
            self.assertEqual(manifest["schema_version"], "reputation-evidence-region-v1")
            digest, name = checksums[0].split("  ", 1)
            self.assertEqual(hashlib.sha256(bundle.read(name)).hexdigest(), digest)

    def test_scope_initialization_and_atomic_mapping_preview(self) -> None:
        csv_path = self.root / "scope.csv"
        headers = [
            "schema_version",
            "seed_key",
            "series_name",
            "vehicle_name",
            "role",
            "role_order",
            "platform_code",
            "platform_vehicle_id",
            "platform_url",
            "platform_display_name",
        ]
        with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            for index in range(27):
                focus = index < 14
                order = index + 1 if focus else index - 13
                writer.writerow(
                    {
                        "schema_version": "reputation-scope-v1",
                        "seed_key": f"vehicle-{index + 1:02d}",
                        "series_name": f"车系{index // 5 + 1}",
                        "vehicle_name": f"车型{index + 1:02d}",
                        "role": "focus" if focus else "competitor",
                        "role_order": order,
                        "platform_code": "dongchedi",
                        "platform_vehicle_id": str(10000 + index),
                        "platform_url": (
                            f"https://www.dongchedi.com/auto/series/score/{10000 + index}-x-x-x-x-x"
                        ),
                        "platform_display_name": f"页面车型{index + 1:02d}",
                    }
                )
        scope = self.client.app.state.container.reputation.initialize_scope_csv(csv_path)
        self.assertEqual(len(scope["vehicles"]), 27)
        preview = self.client.post(
            "/api/v1/reputation/scope/mappings/preview",
            json={
                "revision": scope["revision"],
                "platform_code": "dongchedi",
                "rows": [
                    {
                        "vehicle_id": "vehicle-01",
                        "platform_vehicle_id": "duplicate",
                        "platform_url": "bad-url",
                        "platform_display_name": "错误样例",
                    },
                    {
                        "vehicle_id": "vehicle-01",
                        "platform_vehicle_id": "duplicate",
                        "platform_url": "https://example.test/ok",
                        "platform_display_name": "重复样例",
                    },
                ],
            },
        ).json()
        self.assertFalse(preview["valid"])
        before = self.client.get("/api/v1/reputation/scope").json()
        save = self.client.put(
            "/api/v1/reputation/scope/mappings",
            json={
                "revision": scope["revision"],
                "platform_code": "dongchedi",
                "rows": [
                    {
                        "vehicle_id": "vehicle-01",
                        "platform_vehicle_id": "duplicate",
                        "platform_url": "bad-url",
                        "platform_display_name": "错误样例",
                    }
                ],
            },
        )
        self.assertEqual(save.status_code, 400)
        after = self.client.get("/api/v1/reputation/scope").json()
        self.assertEqual(before["revision"], after["revision"])

    def test_scope_vehicle_add_delete_disable_and_restore_preserve_history(self) -> None:
        scope = self.initialize_scope("vehicle-maintenance.csv")
        self.assertEqual(
            {row["project_group"] for row in scope["vehicles"]}, {"奇瑞项目组"}
        )
        create_payload = {
            "revision": scope["revision"],
            "series_name": "新增车系",
            "vehicle_name": "新增车型",
            "role": "competitor",
            "platform_code": "dongchedi",
            "platform_vehicle_id": "39999",
            "platform_url": "https://www.dongchedi.com/auto/series/score/39999-x-x-x-x-x",
            "platform_display_name": "新增车型页面",
        }
        missing_project_group = self.client.post(
            "/api/v1/reputation/scope/vehicles", json=create_payload
        )
        self.assertEqual(missing_project_group.status_code, 422)
        blank_project_group = self.client.post(
            "/api/v1/reputation/scope/vehicles",
            json={**create_payload, "project_group": "   "},
        )
        self.assertEqual(blank_project_group.status_code, 400)

        created_response = self.client.post(
            "/api/v1/reputation/scope/vehicles",
            json={**create_payload, "project_group": "新能源项目组"},
        )
        self.assertEqual(created_response.status_code, 200, created_response.text)
        created_scope = created_response.json()
        created = next(row for row in created_scope["vehicles"] if row["vehicle_name"] == "新增车型")
        self.assertTrue(created["id"].startswith("rep-"))
        self.assertEqual(created["project_group"], "新能源项目组")
        self.assertEqual(created["removal_mode"], "delete")
        self.assertEqual(created["mappings"]["dongchedi"]["validation_status"], "unverified")

        with self.client.app.state.container.sessions.begin() as db:
            draft = db.get(ReputationScopeDraft, "current")
            assert draft is not None
            data = json.loads(json.dumps(draft.data, ensure_ascii=False))
            stored = next(row for row in data["vehicles"] if row["id"] == created["id"])
            stored["mappings"]["dongchedi"].update(
                {
                    "validation_status": "verified",
                    "validated_at": "2030-01-01T00:00:00+00:00",
                    "validated_mapping_hash": "fixture",
                    "actual_name": "新增车型页面",
                    "latest_metrics": {"score": "4.00", "rank": "2", "volume": "100"},
                }
            )
            draft.data = data
        verified_scope = self.client.get("/api/v1/reputation/scope").json()
        edited_response = self.client.patch(
            f"/api/v1/reputation/scope/vehicles/{created['id']}",
            json={
                "revision": verified_scope["revision"],
                "series_name": "修改车系",
                "vehicle_name": "修改车型",
                "project_group": "其他项目组",
                "role": "focus",
                "platform_code": "dongchedi",
                "platform_vehicle_id": "39999",
                "platform_url": "https://www.dongchedi.com/auto/series/score/39999-x-x-x-x-x",
                "platform_display_name": "新增车型页面",
            },
        )
        self.assertEqual(edited_response.status_code, 200, edited_response.text)
        edited_scope = edited_response.json()
        edited = next(row for row in edited_scope["vehicles"] if row["id"] == created["id"])
        self.assertEqual(edited["series_name"], "修改车系")
        self.assertEqual(edited["vehicle_name"], "修改车型")
        self.assertEqual(edited["project_group"], "其他项目组")
        self.assertEqual(edited["role"], "focus")
        self.assertEqual(edited["role_order"], 15)
        self.assertEqual(edited["mappings"]["dongchedi"]["validation_status"], "verified")
        self.assertIn("latest_metrics", edited["mappings"]["dongchedi"])
        self.assertFalse(edited_scope["last_vehicle_mapping_changed"])

        stale_edit = self.client.patch(
            f"/api/v1/reputation/scope/vehicles/{created['id']}",
            json={
                **create_payload,
                "revision": verified_scope["revision"],
                "project_group": "过期修改",
            },
        )
        self.assertEqual(stale_edit.status_code, 409)

        mapping_edit = self.client.patch(
            f"/api/v1/reputation/scope/vehicles/{created['id']}",
            json={
                "revision": edited_scope["revision"],
                "series_name": "修改车系",
                "vehicle_name": "修改车型",
                "project_group": "其他项目组",
                "role": "focus",
                "platform_code": "dongchedi",
                "platform_vehicle_id": "39998",
                "platform_url": "https://www.dongchedi.com/auto/series/score/39998-x-x-x-x-x",
                "platform_display_name": "修改车型页面",
            },
        )
        self.assertEqual(mapping_edit.status_code, 200, mapping_edit.text)
        mapped_scope = mapping_edit.json()
        mapped = next(row for row in mapped_scope["vehicles"] if row["id"] == created["id"])
        self.assertEqual(mapped["mappings"]["dongchedi"]["validation_status"], "unverified")
        self.assertNotIn("latest_metrics", mapped["mappings"]["dongchedi"])
        self.assertTrue(mapped_scope["last_vehicle_mapping_changed"])

        duplicate = self.client.post(
            "/api/v1/reputation/scope/vehicles",
            json={
                "revision": mapped_scope["revision"],
                "series_name": "重复车系",
                "vehicle_name": "重复车型",
                "project_group": "新能源项目组",
                "role": "focus",
                "platform_vehicle_id": "39998",
                "platform_url": "https://www.dongchedi.com/auto/series/score/39998-x-x-x-x-x",
                "platform_display_name": "重复页面",
            },
        )
        self.assertEqual(duplicate.status_code, 400)

        removed = self.client.delete(
            f"/api/v1/reputation/scope/vehicles/{created['id']}",
            params={"revision": mapped_scope["revision"]},
        )
        self.assertEqual(removed.status_code, 200, removed.text)
        removed_scope = removed.json()
        self.assertEqual(removed_scope["last_vehicle_action"], "deleted")
        self.assertNotIn(created["id"], {row["id"] for row in removed_scope["vehicles"]})

        with self.client.app.state.container.sessions.begin() as db:
            draft = db.get(ReputationScopeDraft, "current")
            assert draft is not None
            version = ReputationScopeVersion(
                version=1,
                snapshot=json.loads(json.dumps(draft.data, ensure_ascii=False)),
                source_revision=draft.revision,
            )
            db.add(version)
            db.flush()
            draft.published_version_id = version.id

        published_scope = self.client.get("/api/v1/reputation/scope").json()
        original = next(row for row in published_scope["vehicles"] if row["id"] == "vehicle-01")
        self.assertEqual(original["removal_mode"], "disable")
        disabled = self.client.delete(
            "/api/v1/reputation/scope/vehicles/vehicle-01",
            params={"revision": published_scope["revision"]},
        )
        self.assertEqual(disabled.status_code, 200, disabled.text)
        disabled_scope = disabled.json()
        self.assertEqual(disabled_scope["last_vehicle_action"], "disabled")
        disabled_vehicle = next(
            row for row in disabled_scope["vehicles"] if row["id"] == "vehicle-01"
        )
        self.assertFalse(disabled_vehicle["enabled"])
        preview = self.client.get("/api/v1/reputation/scope/publish-preview").json()
        self.assertEqual(preview["vehicle_count"], 26)
        self.assertEqual(preview["disabled_count"], 1)

        restored = self.client.post(
            "/api/v1/reputation/scope/vehicles/vehicle-01/restore",
            json={"revision": disabled_scope["revision"]},
        )
        self.assertEqual(restored.status_code, 200, restored.text)
        restored_scope = restored.json()
        self.assertTrue(
            next(row for row in restored_scope["vehicles"] if row["id"] == "vehicle-01")[
                "enabled"
            ]
        )
        with self.client.app.state.container.sessions() as db:
            version = db.scalar(select(ReputationScopeVersion))
            assert version is not None
            self.assertTrue(version.snapshot["vehicles"][0]["enabled"])

    def test_real_mapping_validation_binds_live_metrics_and_evidence(self) -> None:
        csv_path = self.root / "real-scope.csv"
        headers = [
            "schema_version",
            "seed_key",
            "series_name",
            "vehicle_name",
            "role",
            "role_order",
            "platform_code",
            "platform_vehicle_id",
            "platform_url",
            "platform_display_name",
        ]
        with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            ordered_indexes: list[int] = []
            for position in range(14):
                ordered_indexes.append(position)
                if position < 13:
                    ordered_indexes.append(14 + position)
            for index in ordered_indexes:
                focus = index < 14
                platform_id = str(20000 + index)
                writer.writerow(
                    {
                        "schema_version": "reputation-scope-v1",
                        "seed_key": f"real-{index + 1:02d}",
                        "series_name": f"车系{index // 5 + 1}",
                        "vehicle_name": f"车型{index + 1:02d}",
                        "role": "focus" if focus else "competitor",
                        "role_order": index + 1 if focus else index - 13,
                        "platform_code": "dongchedi",
                        "platform_vehicle_id": platform_id,
                        "platform_url": (
                            f"https://www.dongchedi.com/auto/series/score/{platform_id}-x-x-x-x-x"
                        ),
                        "platform_display_name": f"页面车型{index + 1:02d}",
                    }
                )
        scope = self.client.app.state.container.reputation.initialize_scope_csv(csv_path)
        self.client.app.state.container.session_store.import_state(
            "dongchedi",
            {
                "cookies": [
                    {
                        "name": "fixture",
                        "value": "session",
                        "domain": ".dongchedi.com",
                        "path": "/",
                    }
                ]
            },
        )

        class FakeAdapter:
            def __init__(self, *_args, **_kwargs):
                pass

            def validate_sync(self, targets, output_dir):
                output_dir.mkdir(parents=True)
                values = []
                for index, target in enumerate(targets):
                    target_dir = output_dir / target.vehicle_id
                    target_dir.mkdir()
                    full = target_dir / "full.png"
                    metric = target_dir / "metric.png"
                    Image.new("RGB", (800, 1200), "white").save(full)
                    Image.new("RGB", (600, 240), "white").save(metric)
                    values.append(
                        ReputationPageResult(
                            vehicle_id=target.vehicle_id,
                            platform_vehicle_id=target.platform_vehicle_id,
                            mapping_hash=target.mapping_hash,
                            final_url=target.platform_url,
                            actual_name=target.platform_display_name,
                            score_raw="3.80",
                            rank_raw="4",
                            volume_raw=str(500 + index),
                            review_article_count_raw=None,
                            review_article_count_url=None,
                            rank_scope="同级车评分",
                            measurements=[{"stable": True}] * 3,
                            full_page_path=full,
                            metric_region_path=metric,
                            full_page_sha256=hashlib.sha256(full.read_bytes()).hexdigest(),
                            metric_region_sha256=hashlib.sha256(metric.read_bytes()).hexdigest(),
                            width=800,
                            height=1200,
                            metric_rect={"x": 0, "y": 0, "width": 600, "height": 240},
                            duration_ms=100,
                        )
                    )
                return values

        self.client.app.state.container.reputation.adapter_factory = FakeAdapter
        response = self.client.post(
            "/api/v1/reputation/scope/mapping-validations",
            json={"revision": scope["revision"]},
        )
        self.assertEqual(response.status_code, 200, response.text)
        validation = response.json()
        self.assertEqual(validation["succeeded_count"], 27)
        self.assertEqual(validation["failed_count"], 0)
        verified = validation["scope"]["vehicles"]
        self.assertTrue(
            all(row["mappings"]["dongchedi"]["validation_status"] == "verified" for row in verified)
        )
        self.assertEqual(verified[0]["mappings"]["dongchedi"]["latest_metrics"]["volume"], "500")
        evidence = self.client.get(validation["attempts"][0]["metric_region_url"])
        self.assertEqual(evidence.status_code, 200)
        self.assertEqual(evidence.headers["content-type"], "image/png")
        publish = self.client.post(
            "/api/v1/reputation/scope/publish",
            json={
                "revision": validation["scope"]["revision"],
                "initial_review_acknowledged": True,
            },
        )
        self.assertEqual(publish.status_code, 200, publish.text)
        self.assertEqual(publish.json()["published_version"]["version"], 1)
        acceptance = self.client.app.state.container.reputation.create_real_acceptance(
            [validation["id"]]
        )
        self.assertEqual(acceptance["source_type"], "real_acceptance")
        self.assertEqual(acceptance["status"], "success")
        self.assertEqual(len(acceptance["results"]), 27)
        self.assertEqual(
            [row["vehicle_id"] for row in acceptance["results"]],
            [row["id"] for row in verified],
        )
        self.assertEqual(acceptance["complete_evidence_count"], 27)
        compacted = self.client.app.state.container.reputation.compact_region_evidence()
        self.assertGreaterEqual(compacted["validation_attempts"], 27)
        self.assertGreaterEqual(compacted["run_evidence"], 27)
        refreshed = self.client.get(f"/api/v1/reputation/runs/{acceptance['id']}").json()
        run_evidence = refreshed["results"][0]["evidence"]
        self.assertEqual(run_evidence["full_page_sha256"], run_evidence["metric_region_sha256"])
        self.assertEqual(
            self.client.get(run_evidence["full_page_url"]).content,
            self.client.get(run_evidence["metric_region_url"]).content,
        )

    def test_dongchedi_reputation_url_requires_matching_stable_id(self) -> None:
        url = "https://www.dongchedi.com/auto/series/score/24729-x-x-x-x-x"
        self.assertEqual(normalize_series_url(url, "24729"), url)
        with self.assertRaisesRegex(Exception, "车型ID"):
            normalize_series_url(url, "10170")

    def test_dongchedi_metric_region_trims_decoration_above_heading(self) -> None:
        """证据顶部只留4px，不把标题上方的平台装饰横条带入截图。"""

        rect = _metric_rect(
            {
                "heading_box": {"x": 100, "y": 200, "width": 100, "height": 30},
                "score_box": {"x": 300, "y": 250, "width": 20, "height": 20},
                "volume_box": {"x": 330, "y": 240, "width": 80, "height": 20},
                "rank_box": {"x": 500, "y": 220, "width": 400, "height": 300},
            }
        )
        self.assertEqual(rect, {"x": 80.0, "y": 196.0, "width": 840.0, "height": 360.0})

    def test_dongchedi_http_parser_matches_browser_metric_contract(self) -> None:
        """SSR直出页应按当前车型行解析三项指标，而不是误取同级均值。"""

        target = ReputationMappingTarget(
            vehicle_id="vehicle-1",
            platform_vehicle_id="24729",
            platform_url="https://www.dongchedi.com/auto/series/score/24729-x-x-x-x-x",
            platform_display_name="风云A9",
            mapping_hash="fixture",
        )
        content = """
        <html><body>
          <h1 class="tw-hidden">隐藏标题</h1><h1>风云A9</h1>
          <span>共 1,137 人评价</span>
          <div class="rank-wrapper"><ul>
            <li><span class="car-name">车型甲</span><span class="score-wrapper">4.08</span></li>
            <li class="tw-text-common-yellow">
              <span class="car-name">风云A9</span><span class="score-wrapper">3.90</span>
            </li>
          </ul></div>
        </body></html>
        """.encode()
        result = DongchediReputationAdapter._parse_http_page(
            target, target.platform_url, content, duration_ms=12
        )
        self.assertEqual(result.actual_name, "风云A9")
        self.assertEqual(result.score_raw, "3.90")
        self.assertEqual(result.rank_raw, "2")
        self.assertEqual(result.volume_raw, "1137")
        self.assertEqual(result.measurements[-1]["collection_method"], "http_ssr")
        self.assertIsNone(result.metric_region_path)

    def test_dongchedi_review_article_count_comes_from_score_page_next_data(self) -> None:
        """评价篇数读取评分页“全部评分”列表总量，不发起圈子页面请求。"""

        target = ReputationMappingTarget(
            vehicle_id="dcd-8985",
            platform_vehicle_id="8985",
            platform_url="https://www.dongchedi.com/auto/series/score/8985-x-x-x-x-x",
            platform_display_name="风云A9L",
            mapping_hash="fixture",
        )
        content = """
        <html><body><h1>懂风云A9L</h1>
          <span>共2,047人评价</span>
          <div class="rank-wrapper"><ul><li class="tw-text-common-yellow">
            <span class="car-name">风云A9L</span><span class="score-wrapper">3.76</span>
          </li></ul></div>
          <script id="__NEXT_DATA__" type="application/json">
            {"props":{"pageProps":{"reviewListData":{"total_count":2093}}}}
          </script>
        </body></html>
        """.encode()

        result = DongchediReputationAdapter._parse_http_page(
            target, target.platform_url, content, duration_ms=12
        )

        self.assertEqual(result.volume_raw, "2047")
        self.assertEqual(result.review_article_count_raw, "2093")
        self.assertEqual(result.review_article_count_url, target.platform_url)

    def test_dongchedi_presale_zero_reviews_is_confirmed_not_available(self) -> None:
        """未开售页零评分、零评价且全指标为空时应正常保存为平台暂无。"""

        target = ReputationMappingTarget(
            vehicle_id="rep-presale",
            platform_vehicle_id="26120",
            platform_url="https://www.dongchedi.com/auto/series/26120",
            platform_display_name="风云T7",
            mapping_hash="fixture",
        )
        content = """
        <html><body><h1>风云T7</h1>
          <span>暂无报价</span><span>预售价 10.99-12.99 万</span>
          <script id="__NEXT_DATA__" type="application/json">
            {"props":{"pageProps":{"seriesHomeHead":{
              "series_id":26120,
              "series_name":"风云T7",
              "business_status":2,
              "has_official_price":false,
              "has_pre_price":true,
              "total_score":0,
              "total_review_count":0
            }}}}
          </script>
        </body></html>
        """.encode()

        class PageResponse:
            status_code = 200
            url = target.platform_url

        PageResponse.content = content

        class RateResponse:
            status_code = 200
            url = "https://api.dcarapi.com/motor/car_score/api/v1/landing_page/get_detail/"

            @staticmethod
            def json():
                return {
                    "status": 0,
                    "data": {
                        "series_info": {"series_name": "风云T7"},
                        "review_count_info": {"total": 0},
                        "tag_info": None,
                        "tag_info_v2": None,
                    },
                }

        class Session:
            @staticmethod
            def get(url, **_kwargs):
                return RateResponse() if "dcarapi.com" in url else PageResponse()

        adapter = DongchediReputationAdapter(
            None,
            include_review_article_count=True,
            include_negative_rate=True,
        )
        adapter._http_session = lambda: Session()  # type: ignore[method-assign]

        result = adapter._visit_http(target)

        self.assertTrue(result.reputation_not_available)
        self.assertIsNone(result.score_raw)
        self.assertIsNone(result.rank_raw)
        self.assertIsNone(result.volume_raw)
        self.assertIsNone(result.review_article_count_raw)
        self.assertEqual(result.review_article_count_url, target.platform_url)
        self.assertIsNone(result.negative_rate_raw)
        self.assertEqual(result.negative_rate_positive_count, 0)
        self.assertEqual(result.negative_rate_negative_count, 0)
        metrics = self.client.app.state.container.reputation._official_metrics(result, None)
        self.assertTrue(
            all(
                metric["comparison_status"] == "not_available"
                for metric in metrics.values()
            )
        )

    def test_dongchedi_missing_review_count_with_metrics_remains_an_error(self) -> None:
        """已有口碑指标却缺评价篇数时仍按解析异常处理。"""

        target = ReputationMappingTarget(
            vehicle_id="rep-active",
            platform_vehicle_id="26120",
            platform_url="https://www.dongchedi.com/auto/series/26120",
            platform_display_name="风云T7",
            mapping_hash="fixture",
        )
        result = ReputationPageResult(
            vehicle_id=target.vehicle_id,
            platform_vehicle_id=target.platform_vehicle_id,
            mapping_hash=target.mapping_hash,
            final_url=target.platform_url,
            actual_name=target.platform_display_name,
            score_raw="3.90",
            rank_raw="1",
            volume_raw="12",
            review_article_count_raw=None,
            review_article_count_url=None,
            rank_scope="同级车评分",
            measurements=[],
            full_page_path=None,
            metric_region_path=None,
            full_page_sha256=None,
            metric_region_sha256=None,
            width=0,
            height=0,
            metric_rect={},
            duration_ms=1,
            reputation_not_available=True,
        )

        self.assertFalse(
            DongchediReputationAdapter._confirmed_no_reputation_data(
                score_raw=result.score_raw,
                rank_raw=result.rank_raw,
                volume_raw=result.volume_raw,
                review_article_count_raw=result.review_article_count_raw,
                page_not_available=result.reputation_not_available,
                negative_rate_positive_count=0,
                negative_rate_negative_count=0,
                require_negative_rate_confirmation=True,
            )
        )

    def test_dongchedi_negative_rate_reuses_series_id_and_app_counts(self) -> None:
        """正式巡检直接复用平台车型ID，并由优缺点数量计算整数差评率。"""

        requested: list[tuple[str, dict, dict]] = []

        class Response:
            status_code = 200
            url = "https://api.dcarapi.com/motor/car_score/api/v1/landing_page/get_detail/"

            @staticmethod
            def json():
                return {
                    "status": 0,
                    "data": {
                        "series_info": {"series_name": "零跑A10"},
                        "tag_info_v2": {
                            "hierarchical_tag_list": [
                                {"part_id": "3", "tag_name": "优点", "count": 214},
                                {"part_id": "4", "tag_name": "缺点", "count": 128},
                            ]
                        },
                    },
                }

        class Session:
            def get(self, url, **kwargs):
                requested.append((url, kwargs["params"], kwargs["headers"]))
                return Response()

        adapter = DongchediReputationAdapter(None, include_negative_rate=True)
        adapter._http_session = lambda: Session()  # type: ignore[method-assign]
        target = ReputationMappingTarget(
            vehicle_id="dcd-9267",
            platform_vehicle_id="9267",
            platform_url="https://www.dongchedi.com/auto/series/score/9267-x-x-x-x-x",
            platform_display_name="零跑A10",
            mapping_hash="fixture",
        )

        rate, source_url, positive_count, negative_count = adapter._visit_negative_rate(target)

        self.assertEqual(rate, "37%")
        self.assertEqual((positive_count, negative_count), (214, 128))
        self.assertIn("api.dcarapi.com", source_url)
        self.assertEqual(requested[0][1]["series_id"], "9267")
        self.assertEqual(requested[0][1]["aid"], "36")
        self.assertEqual(requested[0][2]["x-tt-appid"], "36")

    def test_dongchedi_negative_rate_marks_zero_reviews_not_available(self) -> None:
        """接口明确零人评分且无标签时保留来源并返回暂无差评率。"""

        class Response:
            status_code = 200
            url = "https://api.dcarapi.com/motor/car_score/api/v1/landing_page/get_detail/"

            @staticmethod
            def json():
                return {
                    "status": 0,
                    "data": {
                        "series_info": {"series_name": "风云T7"},
                        "review_count_info": {"total": 0},
                        "tag_info": None,
                        "tag_info_v2": None,
                    },
                }

        class Session:
            @staticmethod
            def get(_url, **_kwargs):
                return Response()

        adapter = DongchediReputationAdapter(None, include_negative_rate=True)
        adapter._http_session = lambda: Session()  # type: ignore[method-assign]
        target = ReputationMappingTarget(
            vehicle_id="dcd-26120",
            platform_vehicle_id="26120",
            platform_url="https://www.dongchedi.com/auto/series/26120",
            platform_display_name="风云T7",
            mapping_hash="fixture",
        )

        rate, source_url, positive_count, negative_count = adapter._visit_negative_rate(target)

        self.assertIsNone(rate)
        self.assertEqual((positive_count, negative_count), (0, 0))
        self.assertIn("api.dcarapi.com", source_url)

    def test_dongchedi_http_first_opens_browser_only_for_evidence_targets(self) -> None:
        """日常取数不启动浏览器，比较命中的目标才进入截图页面池。"""

        browser_targets: list[str] = []
        completed_targets: list[str] = []

        class FixtureAdapter(DongchediReputationAdapter):
            def _visit_http(self, target):
                measurement = {
                    "actual_name": target.platform_display_name,
                    "score_raw": "3.90" if target.vehicle_id == "vehicle-2" else "3.80",
                    "rank_raw": "2",
                    "volume_raw": "共 500 人评价",
                    "rank_scope": "同级车评分",
                    "collection_method": "http_ssr",
                }
                return ReputationPageResult(
                    vehicle_id=target.vehicle_id,
                    platform_vehicle_id=target.platform_vehicle_id,
                    mapping_hash=target.mapping_hash,
                    final_url=target.platform_url,
                    actual_name=target.platform_display_name,
                    score_raw=measurement["score_raw"],
                    rank_raw="2",
                    volume_raw="500",
                    review_article_count_raw=None,
                    review_article_count_url=None,
                    rank_scope="同级车评分",
                    measurements=[measurement],
                    full_page_path=None,
                    metric_region_path=None,
                    full_page_sha256=None,
                    metric_region_sha256=None,
                    width=0,
                    height=0,
                    metric_rect={},
                    duration_ms=10,
                )

            async def _validate_browser_targets(self, targets, output_dir, **kwargs):
                browser_targets.extend(target.vehicle_id for target in targets)
                values = []
                for index, target in enumerate(targets):
                    target_dir = output_dir / target.vehicle_id
                    target_dir.mkdir()
                    metric = target_dir / "region.png"
                    Image.new("RGB", (600, 240), "white").save(metric)
                    digest = hashlib.sha256(metric.read_bytes()).hexdigest()
                    result = ReputationPageResult(
                            vehicle_id=target.vehicle_id,
                            platform_vehicle_id=target.platform_vehicle_id,
                            mapping_hash=target.mapping_hash,
                            final_url=target.platform_url,
                            actual_name=target.platform_display_name,
                            score_raw="3.90",
                            rank_raw="2",
                            volume_raw="500",
                            review_article_count_raw=None,
                            review_article_count_url=None,
                            rank_scope="同级车评分",
                            measurements=[{"collection_method": "browser"}],
                            full_page_path=metric,
                            metric_region_path=metric,
                            full_page_sha256=digest,
                            metric_region_sha256=digest,
                            width=600,
                            height=240,
                            metric_rect={"x": 0, "y": 0, "width": 600, "height": 240},
                            duration_ms=20,
                        )
                    values.append(result)
                    if kwargs.get("on_result"):
                        kwargs["on_result"](index, target, result)
                return values

        targets = [
            ReputationMappingTarget(
                vehicle_id=f"vehicle-{index}",
                platform_vehicle_id=str(24000 + index),
                platform_url=f"https://www.dongchedi.com/auto/series/{24000 + index}",
                platform_display_name=f"车型{index}",
                mapping_hash=f"hash-{index}",
            )
            for index in range(1, 4)
        ]
        adapter = FixtureAdapter(
            None,
            concurrency=2,
            evidence_policy=lambda target, _measurement: target.vehicle_id == "vehicle-2",
            prefer_http_first=True,
        )
        results = adapter.validate_sync(
            targets,
            self.root / "http-first",
            on_result=lambda _index, target, _result: completed_targets.append(
                target.vehicle_id
            ),
        )
        self.assertEqual(browser_targets, ["vehicle-2"])
        self.assertCountEqual(completed_targets, ["vehicle-1", "vehicle-2", "vehicle-3"])
        self.assertEqual(len(completed_targets), 3)
        self.assertIsNone(results[0].metric_region_path)
        self.assertTrue(results[1].metric_region_path.is_file())
        self.assertIsNone(results[2].metric_region_path)


class OfficialReputationLifecycleTest(unittest.TestCase):
    """验证10:00正式批次、终态产物、补跑与删除的组合生命周期。"""

    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.settings = Settings(
            database_url=f"sqlite:///{(self.root / 'official.db').as_posix()}",
            data_dir=self.root / "data",
            start_background_services=False,
            runtime_mode="test",
        )
        self.client_context = TestClient(create_app(self.settings))
        self.client = self.client_context.__enter__()
        self.service = self.client.app.state.container.reputation
        self.service.adapter_factory = OfficialFakeAdapter
        OfficialFakeAdapter.score_overrides = {}
        OfficialFakeAdapter.negative_rate_overrides = {}
        OfficialFakeAdapter.failures = set()
        OfficialFakeAdapter.retry_once = set()
        OfficialFakeAdapter.progress_probe = None
        OfficialFakeAdapter.last_prefer_http_first = None
        OfficialFakeAdapter.last_include_negative_rate = None
        OfficialFakeAdapter.last_concurrency = None
        self._publish_scope()

    def tearDown(self) -> None:
        self.client_context.__exit__(None, None, None)
        self.temporary.cleanup()

    def _publish_scope(self) -> None:
        csv_path = self.root / "official-scope.csv"
        headers = [
            "schema_version",
            "seed_key",
            "series_name",
            "vehicle_name",
            "role",
            "role_order",
            "platform_code",
            "platform_vehicle_id",
            "platform_url",
            "platform_display_name",
        ]
        with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            for index in range(27):
                focus = index < 14
                platform_id = str(30000 + index)
                writer.writerow(
                    {
                        "schema_version": "reputation-scope-v1",
                        "seed_key": f"official-{index + 1:02d}",
                        "series_name": f"正式车系{index // 5 + 1}",
                        "vehicle_name": f"正式车型{index + 1:02d}",
                        "role": "focus" if focus else "competitor",
                        "role_order": index + 1 if focus else index - 13,
                        "platform_code": "dongchedi",
                        "platform_vehicle_id": platform_id,
                        "platform_url": (
                            "https://www.dongchedi.com/auto/series/score/"
                            f"{platform_id}-x-x-x-x-x"
                        ),
                        "platform_display_name": f"正式页面车型{index + 1:02d}",
                    }
                )
        scope = self.service.initialize_scope_csv(csv_path)
        self.client.app.state.container.session_store.import_state(
            "dongchedi",
            {
                "cookies": [
                    {
                        "name": "fixture",
                        "value": "session",
                        "domain": ".dongchedi.com",
                        "path": "/",
                    }
                ]
            },
        )
        validation = self.service.validate_mappings(
            type("Request", (), {"revision": scope["revision"], "vehicle_ids": None})()
        )
        self.service.publish_scope(
            type(
                "Publish",
                (),
                {
                    "revision": validation["scope"]["revision"],
                    "initial_review_acknowledged": True,
                },
            )()
        )
        with self.client.app.state.container.sessions.begin() as db:
            platform = db.get(PlatformConfig, "dongchedi")
            platform.enabled = True

    @staticmethod
    def _at(day: str, clock: str) -> datetime:
        return datetime.fromisoformat(f"{day}T{clock}:00").replace(
            tzinfo=ZoneInfo("Asia/Shanghai")
        ).astimezone(timezone.utc)

    def test_real_initialization_is_next_day_baseline(self) -> None:
        """真实初始化数据必须承担次日基线，不能因来源类型再次初始化。"""

        collected = self._at("2030-01-01", "15:30")
        with self.client.app.state.container.sessions.begin() as db:
            baseline = ReputationRun(
                number="RP-A-20300101-BASE",
                source_type="real_acceptance",
                run_type="baseline_initialization",
                planned_date="2030-01-01",
                status="success",
                platform_codes=["dongchedi"],
                planned_count=1,
                completed_count=1,
                report_status="success",
                created_at=collected,
                finished_at=collected,
            )
            db.add(baseline)
            db.flush()
            db.add(
                ReputationResult(
                    run_id=baseline.id,
                    vehicle_id="official-01",
                    series_name="正式车系1",
                    vehicle_name="正式车型01",
                    role="focus",
                    role_position=0,
                    vehicle_position=1,
                    platform_code="dongchedi",
                    platform_name="懂车帝",
                    status="success",
                    metrics={
                        "score": {"raw": "3.70", "value": "3.70"},
                        "rank": {"raw": "4", "value": "4", "scope": "同级车评分"},
                        "volume": {"raw": "500", "value": "500"},
                    },
                    collected_at=collected,
                )
            )

        due = self.service.check_schedule(self._at("2030-01-02", "10:00"))
        run_id = due["queued_run_ids"][0]
        queued = self.service.get_run(run_id)
        self.assertEqual(queued["run_type"], "daily")
        self.assertEqual(queued["baseline_source_run_id"], baseline.id)

        finished = self.service.execute_run(run_id)
        first = next(
            item for item in finished["results"] if item["vehicle_id"] == "official-01"
        )
        self.assertEqual(first["metrics"]["score"]["baseline_value"], "3.70")
        self.assertEqual(first["metrics"]["score"]["direction"], "up")

    def test_coordinator_generates_report_in_same_terminal_tick(self) -> None:
        coordinator = ReputationCoordinator(self.service)
        outcome = coordinator.tick(self._at("2030-01-02", "10:00"))
        run_id = outcome["queued_run_ids"][0]
        run = self.service.get_run(run_id)

        self.assertEqual(run["status"], "success")
        self.assertEqual(run["report_status"], "success")
        self.assertIsNone(run["report_planned_at"])
        self.assertIsNotNone(run["report_generated_at"])
        self.assertTrue(Path(run["downloads"]["txt"].split("/api/v1")[-1]).name)
        self.assertEqual(
            self.service.schedule_status()["last_event"]["message"],
            "正式口碑巡检已到达终态，汇报已生成。",
        )

    def test_official_run_concurrency_is_fixed_at_two(self) -> None:
        """帖子平台并发和异常排队快照都不得扩大巡检页面并发。"""

        with self.client.app.state.container.sessions.begin() as db:
            platform = db.get(PlatformConfig, "dongchedi")
            platform.internal_concurrency = 8

        due = self.service.check_schedule(self._at("2030-01-02", "10:00"))
        run_id = due["queued_run_ids"][0]
        self.assertEqual(2, self.service.get_run(run_id)["concurrency"])

        # 模拟修复前遗留或异常写入的排队快照；执行入口仍须收敛回固定值。
        with self.client.app.state.container.sessions.begin() as db:
            db.get(ReputationRun, run_id).concurrency = 8

        OfficialFakeAdapter.failures = {"official-01"}
        finished = self.service.execute_run(run_id)

        self.assertEqual(2, OfficialFakeAdapter.last_concurrency)
        self.assertEqual(2, finished["concurrency"])
        self.assertEqual("partial_success", finished["status"])
        retry = self.service.retry_failed(run_id)
        self.assertEqual(2, retry["concurrency"])

    def test_available_page_with_unknown_metric_stays_partial(self) -> None:
        """页面有口碑但缺可靠指标来源时应保留截图并显式标记部分成功。"""

        due = self.service.check_schedule(self._at("2030-01-02", "10:00"))
        run_id = due["queued_run_ids"][0]
        OfficialFakeAdapter.negative_rate_overrides = {"official-01": ""}

        finished = self.service.execute_run(run_id)
        first = next(
            item for item in finished["results"] if item["vehicle_id"] == "official-01"
        )

        self.assertEqual("partial_success", first["status"])
        self.assertEqual(
            "unknown", first["metrics"]["negative_rate"]["comparison_status"]
        )
        self.assertIsNotNone(first["evidence"])
        self.assertEqual(27, finished["complete_evidence_count"])
        self.assertEqual("partial_success", finished["status"])

    def test_official_run_persists_linear_progress_before_terminal_state(self) -> None:
        """每个车型终态都应先落库并发布进度，批次结束后才冻结汇报。"""

        due = self.service.check_schedule(self._at("2030-01-02", "10:00"))
        run_id = due["queued_run_ids"][0]
        observed: list[tuple[int, int, int, int, str]] = []

        def probe() -> None:
            current = self.service.get_run(run_id)
            observed.append(
                (
                    current["completed_count"] + current["failed_count"],
                    len(current["results"]),
                    current["required_evidence_count"],
                    current["complete_evidence_count"],
                    current["status"],
                )
            )

        OfficialFakeAdapter.progress_probe = probe
        finished = self.service.execute_run(run_id)

        self.assertEqual(list(range(1, 28)), [item[0] for item in observed])
        self.assertEqual(list(range(1, 28)), [item[1] for item in observed])
        self.assertTrue(all(item[2] == 27 for item in observed))
        self.assertEqual(list(range(1, 28)), [item[3] for item in observed])
        self.assertTrue(all(item[4] == "running" for item in observed))
        self.assertEqual("success", finished["status"])
        events = [
            event
            for event in self.client.app.state.container.events.wait_after(0, timeout=0)
            if event["type"] == "reputation.run.changed"
            and event["resource_id"] == run_id
        ]
        self.assertEqual(0, events[0]["summary"]["completed_count"])
        self.assertEqual(
            list(range(1, 28)),
            [event["summary"]["completed_count"] for event in events[1:28]],
        )
        self.assertEqual("success", events[-1]["summary"]["status"])

    def test_retryable_item_counts_only_after_second_attempt_terminal(self) -> None:
        """首轮暂时错误保持未完成，第二次尝试成功后才进入进度分子。"""

        due = self.service.check_schedule(self._at("2030-01-02", "10:00"))
        run_id = due["queued_run_ids"][0]
        OfficialFakeAdapter.retry_once = {"official-01"}

        finished = self.service.execute_run(run_id)

        retried = next(
            item for item in finished["results"] if item["vehicle_id"] == "official-01"
        )
        self.assertEqual(2, retried["attempt_count"])
        self.assertEqual("success", retried["status"])
        events = [
            event
            for event in self.client.app.state.container.events.wait_after(0, timeout=0)
            if event["type"] == "reputation.run.changed"
            and event["resource_id"] == run_id
        ]
        progress = events[1:-1]
        self.assertEqual(list(range(1, 28)), [item["summary"]["completed_count"] for item in progress])
        self.assertTrue(all(item["summary"]["failed_count"] == 0 for item in progress))

    def test_official_schedule_baseline_daily_retry_delete_and_missed_day(self) -> None:
        schedule = self.service.schedule_status()
        self.assertEqual(schedule["inspection_time"], "10:00:00")
        self.assertIsNone(schedule["report_time"])

        before = self.service.check_schedule(self._at("2030-01-02", "09:59"))
        self.assertIsNone(before["created_run_id"])

        due = self.service.check_schedule(self._at("2030-01-02", "10:00"))
        self.assertEqual(len(due["queued_run_ids"]), 1)
        baseline_id = due["queued_run_ids"][0]
        self.assertTrue(
            self.client.app.state.container.worker._official_reputation_waiting("dongchedi")
        )
        self.assertTrue(self.service.can_execute_official(baseline_id))
        duplicate = self.service.check_schedule(self._at("2030-01-02", "10:01"))
        self.assertEqual(duplicate["created_run_id"], baseline_id)
        baseline = self.service.execute_run(baseline_id)
        self.assertFalse(
            self.client.app.state.container.worker._official_reputation_waiting("dongchedi")
        )
        self.assertEqual(baseline["run_type"], "baseline_initialization")
        self.assertEqual(baseline["status"], "success")
        self.assertEqual(baseline["complete_evidence_count"], 27)
        self.assertFalse(OfficialFakeAdapter.last_prefer_http_first)
        self.assertTrue(OfficialFakeAdapter.last_include_negative_rate)
        self.assertIsNone(baseline["report_planned_at"])
        report_due = self.service.check_schedule(self._at("2030-01-02", "10:01"))
        self.assertIn(baseline_id, report_due["report_run_ids"])
        reported = self.service.generate_report(baseline_id, self._at("2030-01-02", "10:01"))
        self.assertEqual(reported["report_status"], "success")
        self.assertIsNone(reported["report_planned_at"])
        self.assertTrue(Path(reported["downloads"]["txt"].split("/api/v1")[-1]).name)

        OfficialFakeAdapter.score_overrides = {"official-01": "3.90"}
        OfficialFakeAdapter.negative_rate_overrides = {
            "official-01": "41%",
            "official-03": "32%",
        }
        daily_due = self.service.check_schedule(self._at("2030-01-03", "10:00"))
        daily_id = daily_due["queued_run_ids"][0]
        daily = self.service.execute_run(daily_id)
        self.assertEqual(daily["run_type"], "daily")
        self.assertEqual(daily["baseline_date"], "2030-01-02")
        self.assertEqual(daily["required_evidence_count"], 27)
        self.assertEqual(daily["complete_evidence_count"], 27)
        self.assertTrue(all(item["evidence_required"] for item in daily["results"]))
        self.assertTrue(all(item["evidence"] for item in daily["results"]))
        self.assertFalse(OfficialFakeAdapter.last_prefer_http_first)
        self.assertTrue(OfficialFakeAdapter.last_include_negative_rate)
        changed = next(item for item in daily["results"] if item["vehicle_id"] == "official-01")
        self.assertEqual(changed["metrics"]["score"]["direction"], "up")
        self.assertEqual(changed["metrics"]["review_article_count"]["raw"], "5000")
        self.assertEqual(changed["metrics"]["review_article_count"]["direction"], "same")
        self.assertEqual(changed["metrics"]["negative_rate"]["direction"], "up")
        self.assertEqual(changed["metrics"]["negative_rate"]["tone"], "negative")
        self.assertEqual(changed["metrics"]["negative_rate"]["positive_count"], 214)
        improved = next(item for item in daily["results"] if item["vehicle_id"] == "official-03")
        self.assertEqual(improved["metrics"]["negative_rate"]["direction"], "down")
        self.assertEqual(improved["metrics"]["negative_rate"]["tone"], "positive")

        OfficialFakeAdapter.failures = {"official-02"}
        failed_due = self.service.check_schedule(self._at("2030-01-04", "10:00"))
        failed_id = failed_due["queued_run_ids"][0]
        failed = self.service.execute_run(failed_id)
        self.assertEqual(failed["status"], "partial_success")
        incomplete_report = self.service.generate_report(
            failed_id, self._at("2030-01-04", "10:01")
        )
        self.assertIn("【不完整汇报】", incomplete_report["report_text"])
        self.assertNotIn("今日无口碑指标变化", incomplete_report["report_text"])
        retry = self.service.retry_failed(failed_id)
        self.assertEqual(retry["source_type"], "retry")
        self.assertEqual(retry["planned_count"], 1)
        OfficialFakeAdapter.failures = set()
        retry_done = self.service.execute_run(retry["id"])
        self.assertEqual(retry_done["status"], "success")
        linked = self.service.get_run(failed_id)
        self.assertEqual(linked["linked_status"], "success")
        self.assertEqual(linked["unresolved_count"], 0)
        self.assertEqual(linked["linked_complete_evidence_count"], 27)
        recovered = next(
            item for item in linked["results"] if item["vehicle_id"] == "official-02"
        )
        self.assertEqual(recovered["status"], "success")
        self.assertIsNotNone(recovered["evidence"])
        with self.assertRaisesRegex(Exception, "已经全部补跑成功"):
            self.service.retry_failed(failed_id)

        job = self.service.delete_official(failed_id)
        self.assertEqual(job["status"], "success", job)
        with self.client.app.state.container.sessions() as db:
            self.assertIsNone(db.get(ReputationRun, failed_id))
            self.assertEqual(
                db.scalar(select(func.count()).select_from(ReputationTombstone)), 1
            )
        same_day = self.service.check_schedule(self._at("2030-01-04", "13:00"))
        self.assertIsNone(same_day["created_run_id"])

        missed = self.service.check_schedule(self._at("2030-01-06", "09:00"))
        self.assertIn("2030-01-05", missed["missed_dates"])
        with self.client.app.state.container.sessions() as db:
            event = db.scalar(
                select(ReputationScheduleEvent).where(
                    ReputationScheduleEvent.planned_date == "2030-01-05"
                )
            )
            self.assertEqual(event.status, "missed")

        resumed = self.service.check_schedule(self._at("2030-01-06", "10:00"))
        resumed_run = self.service.get_run(resumed["queued_run_ids"][0])
        self.assertEqual(resumed_run["run_type"], "daily")
        self.assertEqual(resumed_run["baseline_date"], "2030-01-05")
        self.assertIsNone(resumed_run["baseline_source_run_id"])


if __name__ == "__main__":
    unittest.main()
