"""易车适配器复用公共配置、队列、计划、详情与导出闭环。"""

from __future__ import annotations

from dataclasses import replace
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from threadsnap.collectors.registry import PLATFORM_ADAPTERS
from threadsnap.collectors.yiche import _parse_time
from threadsnap.models import (
    Circle,
    CircleTask,
    ExtractionRun,
    PlatformConfig,
    ValidationJob,
)
from threadsnap.services import bootstrap_database

try:
    from .test_backend import AppCase
except ImportError:  # unittest discover 以顶层模块导入测试。
    from test_backend import AppCase


def yiche_record(post_id: str) -> dict:
    """构造遵循易车已确认公共映射的脱敏记录。"""

    return {
        "platform_post_id": post_id,
        "url": f"https://baa.yiche.com/sample/thread-{post_id}.html",
        "title": f"易车样本{post_id}",
        "author": "样本作者",
        "published_at": _parse_time("2026-08-27 10:30:00"),
        "content": "结构化正文",
        "image_urls": ["https://media.example.test/sample.jpg"],
        "video_urls": [],
        "reply_count": 1,
        "like_count": 2,
        "section": "dynamic",
        "visibility": "visible",
        "raw_status": {
            "forum_id": 9001,
            "post_type": 1,
            "document_http_status": 200,
            "document_classification": "content",
            "detail_identity_verified": True,
            "comment_api_status": "success",
            "comment_api_business_status": "1",
            "comment_identity_verified": True,
            "comment_termination": "have_next_false",
        },
        "comments": [
            {
                "platform_comment_id": f"c-{post_id}",
                "author": "样本评论者",
                "content": "一级评论",
                "published_at": datetime(2026, 8, 27, 2, 31, tzinfo=timezone.utc),
                "like_count": 0,
            }
        ],
    }


class FakeYicheCollector:
    """只替代平台网络边界，公共流程仍走真实服务与数据库。"""

    concurrency = 1
    supports_page_evidence = False

    def validate_circle(self, url: str) -> dict:
        order = "latest_publish" if "index-0-1-" in url else "latest_reply"
        return {
            "external_id": "sample",
            "forum_id": 9001,
            "seo_name": "sample",
            "forum_name": "样本社区",
            "name": "样本社区",
            "url": url,
            "sort": order,
            "adapter_version": "yiche-test-v1",
        }

    def collect_circle(
        self,
        _url: str,
        target: int,
        *,
        skip_post_ids: set[str] | None = None,
        on_progress=None,
    ) -> dict:
        skipped = skip_post_ids or set()
        records = []
        candidate = 7001
        while len(records) < target:
            post_id = str(candidate)
            candidate += 1
            if post_id in skipped:
                continue
            record = yiche_record(post_id)
            records.append(record)
            if on_progress:
                on_progress(record, None)
        return {"records": records, "failures": [], "stop_reason": "达到目标数量。"}

    def collect_urls(self, urls: list[str], *, on_progress=None) -> dict:
        records = [yiche_record(Path(url).stem.removeprefix("thread-")) for url in urls]
        if on_progress:
            for record in records:
                on_progress(record, None)
        return {"records": records, "failures": [], "stop_reason": "URL清单完成。"}

    def resolve_record_video_urls(
        self, _raw_status: dict, _stored_urls: list[str] | None = None
    ) -> None:
        return None


class YicheReleaseStateTests(AppCase):
    def test_released_adapter_is_available_but_default_disabled(self) -> None:
        platform = next(
            item for item in self.client.get("/api/v1/platforms").json() if item["code"] == "yiche"
        )
        self.assertEqual("available", platform["adapter_status"])
        self.assertFalse(platform["enabled"])
        self.assertFalse(platform["capabilities"]["page_evidence"])

        enabled = self.client.put(
            "/api/v1/platforms/yiche",
            json={"enabled": True, "internal_concurrency": 1},
        )
        self.assertEqual(200, enabled.status_code, enabled.text)
        self.assertTrue(enabled.json()["enabled"])

    def test_bootstrap_upgrades_existing_database_to_released_adapter(self) -> None:
        with self.container.sessions.begin() as db:
            platform = db.get(PlatformConfig, "yiche")
            assert platform is not None
            platform.adapter_status = "not_integrated"
            platform.enabled = False
        with self.container.sessions.begin() as db:
            bootstrap_database(db)
        with self.container.sessions() as db:
            platform = db.get(PlatformConfig, "yiche")
            assert platform is not None
            self.assertEqual("available", platform.adapter_status)
            self.assertFalse(platform.enabled)

    def test_bootstrap_withdrawal_closes_active_runs_and_preserves_history(self) -> None:
        with self.container.sessions.begin() as db:
            platform = db.get(PlatformConfig, "yiche")
            assert platform is not None
            platform.adapter_status = "available"
            platform.enabled = True

        def create_run(key: str) -> str:
            response = self.client.post(
                "/api/v1/runs/manual",
                json={
                    "platform_code": "yiche",
                    "circle_urls": ["https://baa.yiche.com/sample/"],
                    "quantity": 1,
                    "screenshot_enabled": False,
                    "idempotency_key": key,
                },
            )
            self.assertEqual(202, response.status_code, response.text)
            return str(response.json()["id"])

        queued_run_id = create_run("withdrawn-yiche-queued")
        running_run_id = create_run("withdrawn-yiche-running")
        history_run_id = create_run("withdrawn-yiche-success-history")
        completed_at = datetime.now(timezone.utc)
        with self.container.sessions.begin() as db:
            running_run = db.get(ExtractionRun, running_run_id)
            running_task = db.scalar(
                select(CircleTask).where(CircleTask.run_id == running_run_id)
            )
            history_run = db.get(ExtractionRun, history_run_id)
            history_task = db.scalar(
                select(CircleTask).where(CircleTask.run_id == history_run_id)
            )
            assert running_run is not None and running_task is not None
            assert history_run is not None and history_task is not None
            running_run.status = "running"
            running_run.started_at = completed_at
            running_task.status = "running"
            running_task.started_at = completed_at
            history_run.status = "success"
            history_run.completed_count = 1
            history_run.finished_at = completed_at
            history_task.status = "success"
            history_task.completed_count = 1
            history_task.finished_at = completed_at

        dormant_spec = replace(PLATFORM_ADAPTERS["yiche"], adapter_status="not_integrated")
        with patch.dict(PLATFORM_ADAPTERS, {"yiche": dormant_spec}):
            with self.container.sessions.begin() as db:
                bootstrap_database(db)

        with self.container.sessions() as db:
            platform = db.get(PlatformConfig, "yiche")
            assert platform is not None
            self.assertEqual("not_integrated", platform.adapter_status)
            self.assertFalse(platform.enabled)
            for run_id in (queued_run_id, running_run_id):
                run = db.get(ExtractionRun, run_id)
                task = db.scalar(select(CircleTask).where(CircleTask.run_id == run_id))
                assert run is not None and task is not None
                self.assertEqual("failed", task.status)
                self.assertEqual("PLATFORM_NOT_INTEGRATED", task.error_code)
                self.assertIsNotNone(task.finished_at)
                self.assertEqual("failed", run.status)
                self.assertIsNotNone(run.finished_at)
            active_tasks = list(
                db.scalars(
                    select(CircleTask).where(
                        CircleTask.platform_code == "yiche",
                        CircleTask.status.in_(["queued", "running"]),
                    )
                )
            )
            self.assertEqual([], active_tasks)
            history_run = db.get(ExtractionRun, history_run_id)
            history_task = db.scalar(
                select(CircleTask).where(CircleTask.run_id == history_run_id)
            )
            assert history_run is not None and history_task is not None
            self.assertEqual("success", history_run.status)
            self.assertEqual("success", history_task.status)
            self.assertEqual(completed_at, history_run.finished_at)
            self.assertEqual(completed_at, history_task.finished_at)

    def test_worker_fails_validation_job_after_adapter_withdrawal(self) -> None:
        with self.container.sessions.begin() as db:
            platform = db.get(PlatformConfig, "yiche")
            assert platform is not None
            platform.adapter_status = "not_integrated"
            platform.enabled = False
            circle = Circle(
                platform_code="yiche",
                external_id="sample",
                url="https://baa.yiche.com/sample/",
                source_kind="configured",
            )
            db.add(circle)
            db.flush()
            job = ValidationJob(circle_id=circle.id)
            db.add(job)
            db.flush()
            job_id = job.id

        dormant_spec = replace(PLATFORM_ADAPTERS["yiche"], adapter_status="not_integrated")
        with patch.dict(PLATFORM_ADAPTERS, {"yiche": dormant_spec}):
            self.assertTrue(self.container.worker.process_once())
        with self.container.sessions() as db:
            job = db.get(ValidationJob, job_id)
            assert job is not None
            self.assertEqual("failed", job.status)
            self.assertEqual("PLATFORM_NOT_INTEGRATED", job.error_code)


class YichePublicFlowTests(AppCase):
    def setUp(self) -> None:
        super().setUp()
        enabled = self.client.put(
            "/api/v1/platforms/yiche",
            json={"enabled": True, "internal_concurrency": 8},
        )
        self.assertEqual(200, enabled.status_code, enabled.text)
        self.assertEqual(1, enabled.json()["internal_concurrency"])
        self.container.worker._collector = lambda *_args: FakeYicheCollector()  # type: ignore[method-assign]

    def save_yiche_sources(self) -> list[Circle]:
        response = self.client.put(
            "/api/v1/circles/batch",
            json={
                "rows": [
                    {
                        "platform_code": "yiche",
                        "url": "https://baa.yiche.com/sample/index-0-0-2.html",
                        "vehicle_name": "样本最新回复",
                    },
                    {
                        "platform_code": "yiche",
                        "url": "https://baa.yiche.com/sample/index-0-1-1.html?tag=-1",
                        "vehicle_name": "样本最新发布",
                    },
                ],
                "deleted_ids": [],
            },
        )
        self.assertEqual(200, response.status_code, response.text)
        ids = [item["id"] for item in response.json()["items"]]
        with self.container.sessions.begin() as db:
            items = [db.get(Circle, circle_id) for circle_id in ids]
            for item in items:
                assert item is not None
                item.validation_status = "verified"
                item.auto_enabled = True
            return [item for item in items if item is not None]

    def test_manual_detail_and_xlsx_close_over_yiche_without_new_storage_schema(self) -> None:
        circles = self.save_yiche_sources()
        self.assertIn(
            self.container.auth._validation_probe_url("yiche"),
            {circle.url for circle in circles},
        )
        auth_task = self.client.post("/api/v1/platforms/yiche/auth/tasks")
        self.assertEqual(202, auth_task.status_code, auth_task.text)
        self.assertEqual("yiche", auth_task.json()["platform_code"])
        created = self.client.post(
            "/api/v1/runs/manual",
            json={
                "platform_code": "yiche",
                "circle_ids": [circles[0].id],
                "quantity": 1,
                "ai_analysis_enabled": False,
                "screenshot_enabled": False,
                "idempotency_key": "yiche-manual-close-loop",
            },
        )
        self.assertEqual(202, created.status_code, created.text)
        self.assertTrue(self.container.worker.process_once())
        run = self.client.get(f"/api/v1/runs/{created.json()['id']}").json()
        self.assertEqual("success", run["status"])
        posts = self.client.get(f"/api/v1/runs/{run['id']}/posts").json()["items"]
        self.assertEqual(1, len(posts))
        detail = self.client.get(f"/api/v1/runs/{run['id']}/posts/{posts[0]['id']}").json()
        self.assertEqual("visible", detail["visibility"])
        self.assertTrue(detail["published_at"].startswith("2026-08-27T02:30:00"))
        self.assertEqual("一级评论", detail["comments"][0]["content"])

        tag = next(
            item["tag"]
            for item in self.container.templates.field_tags(circles[0].id)
            if item["field"] == "post.title"
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = tag
        source = Path(self.temp.name) / "yiche-template.xlsx"
        workbook.save(source)
        version = self.container.templates.upload("易车模板", source.name, source.read_bytes())
        exported = self.container.templates.create_export(run["id"], version["version_id"])
        output = load_workbook(self.container.templates.export_path(exported["id"])).active
        self.assertEqual("易车样本7001", output["A1"].value)

    def test_yiche_normalizes_unsupported_page_evidence_for_manual_and_plan(self) -> None:
        circle = self.save_yiche_sources()[0]
        manual = self.client.post(
            "/api/v1/runs/manual",
            json={
                "platform_code": "yiche",
                "circle_ids": [circle.id],
                "quantity": 1,
                "screenshot_enabled": True,
                "idempotency_key": "yiche-page-evidence-manual",
            },
        )
        self.assertEqual(202, manual.status_code, manual.text)
        self.assertFalse(manual.json()["screenshot_enabled"])
        with self.container.sessions() as db:
            stored = db.get(ExtractionRun, manual.json()["id"])
            assert stored is not None
            self.assertTrue(stored.config_snapshot["requested_screenshot_enabled"])
            self.assertFalse(stored.config_snapshot["screenshot_enabled"])

        plan = self.client.put(
            "/api/v1/extraction-plan",
            json={
                "revision": 1,
                "rules": [
                    {
                        "id": "yiche-evidence-rule",
                        "name": "易车截图边界",
                        "platform_quantities": {"yiche": 1},
                        "circle_ids": [circle.id],
                        "screenshot_enabled": True,
                    }
                ],
                "nodes": [
                    {
                        "id": "yiche-evidence-node",
                        "weekdays": [4],
                        "time": "11:00:00",
                        "enabled": True,
                        "rule_ids": ["yiche-evidence-rule"],
                    }
                ],
                "recurring_nodes": [],
            },
        )
        self.assertEqual(200, plan.status_code, plan.text)
        self.assertTrue(plan.json()["rules"][0]["screenshot_enabled"])
        scheduled = self.container.runs.create_scheduled(
            datetime(2026, 8, 14, 3, 0, tzinfo=timezone.utc),
            "yiche-evidence-node",
            plan.json()["revision"],
        )
        assert scheduled is not None
        detail = self.container.runs.get_run(scheduled["id"])
        self.assertFalse(detail["screenshot_enabled"])
        self.assertFalse(detail["tasks"][0]["screenshot_enabled"])

    def test_weekly_and_recurring_runs_share_yiche_fifo(self) -> None:
        circle = self.save_yiche_sources()[0]
        saved = self.client.put(
            "/api/v1/extraction-plan",
            json={
                "revision": 1,
                "rules": [
                    {
                        "id": "yiche-rule",
                        "name": "易车规则",
                        "platform_quantities": {"yiche": 1},
                        "circle_ids": [circle.id],
                        "ai_analysis_enabled": False,
                        "screenshot_enabled": False,
                    }
                ],
                "nodes": [
                    {
                        "id": "yiche-weekly",
                        "weekdays": [4],
                        "time": "10:00:00",
                        "enabled": True,
                        "rule_ids": ["yiche-rule"],
                    }
                ],
                "recurring_nodes": [
                    {
                        "id": "yiche-recurring",
                        "weekdays": [4],
                        "start_time": "10:00:00",
                        "end_time": "10:05:00",
                        "interval_minutes": 5,
                        "enabled": True,
                        "rule_ids": ["yiche-rule"],
                    }
                ],
            },
        )
        self.assertEqual(200, saved.status_code, saved.text)
        self.container.scheduler.tick(datetime(2026, 8, 14, 2, 0, tzinfo=timezone.utc))
        runs = self.client.get("/api/v1/runs?trigger_types=scheduled&trigger_types=recurring").json()
        self.assertEqual(2, runs["total"])
        self.assertTrue(self.container.worker.process_once())
        states = self.client.get(
            "/api/v1/runs?trigger_types=scheduled&trigger_types=recurring"
        ).json()["items"]
        self.assertEqual(1, sum(item["status"] == "success" for item in states))
        self.assertEqual(1, sum(item["status"] == "queued" for item in states))
        self.assertTrue(self.container.worker.process_once())
        states = self.client.get(
            "/api/v1/runs?trigger_types=scheduled&trigger_types=recurring"
        ).json()["items"]
        self.assertTrue(all(item["status"] == "success" for item in states))
