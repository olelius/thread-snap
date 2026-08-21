"""ThreadSnap 第一版后端的领域、接口、队列和导出验证。"""

from __future__ import annotations

import asyncio
import base64
import hashlib
import json
import tempfile
import unittest
from datetime import datetime, timedelta, timezone
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from patchright.async_api import Error as PlaywrightError
from sqlalchemy import select

from threadsnap.app import create_app, require_internal_loopback
from threadsnap.auth import AuthPageLoadError, AuthTask
from threadsnap.collectors.dongchedi import (
    ADAPTER_VERSION,
    AuthenticationRequired,
    DongchediCollector,
    parse_circle_url,
)
from threadsnap.config import Settings
from threadsnap.errors import DomainError
from threadsnap.models import (
    Circle,
    CircleTask,
    CommentSnapshot,
    ExtractionRun,
    ExtractionRunRule,
    PlatformConfig,
    PlatformSession,
    PostSnapshot,
    ScheduleEvent,
    ScheduleNodeRule,
    SentimentAnalysis,
    Vehicle,
)
from threadsnap.schemas import (
    CircleBatchUpdate,
    CircleRow,
    ManualRunCreate,
)
from threadsnap.sentiment import (
    DEEPSEEK_MODEL_CODE,
    HOSTED_MODEL_CODE,
    SentimentFeedback,
    build_request,
    deduplicate_media_urls,
    normalize_feedback_payload,
    normalize_text_only_feedback_payload,
    sentiment_input_hash,
    validate_modality_identity,
)
from threadsnap.services import bootstrap_database


def sample_record(post_id: str) -> dict:
    """构造一条包含主评论的标准帖子快照。"""

    return {
        "platform_post_id": post_id,
        "url": f"https://www.dongchedi.com/ugc/article/{post_id}",
        "title": f"标题{post_id}",
        "author": "作者甲",
        "published_at": datetime(2026, 8, 14, 2, 30, tzinfo=timezone.utc),
        "content": f"正文{post_id}",
        "image_urls": ["https://example.test/a.jpg"],
        "video_urls": [],
        "reply_count": 1,
        "like_count": 3,
        "section": "动态",
        "visibility": "visible",
        "raw_status": {"operation_status": 0},
        "comments": [
            {
                "platform_comment_id": f"c-{post_id}",
                "author": "评论者乙",
                "content": "一级评论",
                "published_at": datetime(2026, 8, 14, 2, 35, tzinfo=timezone.utc),
                "like_count": 5,
            }
        ],
    }


class FakeCollector:
    """可控的平台采集器，用于验证持久队列而不访问网络。"""

    def __init__(self, *, auth: bool = False):
        self.auth = auth
        self.concurrency = 1

    def validate_circle(self, url: str) -> dict:
        return {
            "platform_code": "dongchedi",
            "external_id": url.rstrip("/").split("/")[-1],
            "name": "测试车友圈",
            "url": url,
            "section": "dynamic",
            "sort": "latest_reply",
            "sample_post_id": "1001",
            "adapter_version": "fake-v1",
        }

    def collect_circle(
        self,
        _url: str,
        target: int,
        skip_post_ids: set[str] | None = None,
        on_progress=None,
    ) -> dict:
        if self.auth:
            raise AuthenticationRequired("平台会话需要刷新。", trigger_url=_url)
        skip = skip_post_ids or set()
        records = []
        candidate = 1001
        while len(records) < target:
            post_id = str(candidate)
            candidate += 1
            if post_id in skip:
                continue
            record = sample_record(post_id)
            records.append(record)
            if on_progress:
                on_progress(record, None)
        return {
            "records": records,
            "failures": [],
            "stop_reason": "已经取得配置数量的有效帖子。",
        }

    def collect_urls(self, urls: list[str], on_progress=None) -> dict:
        records = [sample_record(url.rstrip("/").split("/")[-1]) for url in urls]
        if on_progress:
            for record in records:
                on_progress(record, None)
        return {
            "records": records,
            "failures": [],
            "stop_reason": "已处理全部导入帖子链接。",
        }


class FakeAuthResponse:
    def __init__(self, *, status: int = 200, headers: dict[str, str] | None = None):
        self.status = status
        self.headers = headers or {}

    async def all_headers(self) -> dict[str, str]:
        return self.headers


class FakeAuthLocator:
    def __init__(self, *, text: str = "", count: int = 0):
        self.text = text
        self.value_count = count

    async def inner_text(self, **_kwargs: object) -> str:
        return self.text

    async def count(self) -> int:
        return self.value_count


class FakeAuthPage:
    url = "https://www.dongchedi.com/login-required"

    def __init__(self, *, html: str = "<html><body>登录</body></html>", controls: int = 1):
        self.html = html
        self.controls = controls

    def is_closed(self) -> bool:
        return False

    async def content(self) -> str:
        return self.html

    def locator(self, selector: str) -> FakeAuthLocator:
        if selector == "body":
            return FakeAuthLocator(text="登录" if "登录" in self.html else "")
        return FakeAuthLocator(count=self.controls)

    async def wait_for_timeout(self, _milliseconds: float) -> None:
        return None


class FakeAuthContext:
    def __init__(self, state: dict):
        self.state = state
        self.closed = False

    async def storage_state(self) -> dict:
        return self.state

    async def close(self) -> None:
        self.closed = True


class AlreadyClosedAuthContext(FakeAuthContext):
    async def close(self) -> None:
        raise PlaywrightError(
            "BrowserContext.close: Target page, context or browser has been closed"
        )


class FakePlaywright:
    def __init__(self):
        self.stopped = False

    async def stop(self) -> None:
        self.stopped = True


class FakeAuthSocket:
    def __init__(self):
        self.messages: list[dict] = []

    async def send_json(self, message: dict) -> None:
        self.messages.append(message)


class FakeCDPSession:
    def __init__(self):
        self.calls: list[tuple[str, dict | None]] = []
        self.listeners: dict[str, object] = {}
        self.detached = False

    def on(self, event: str, listener: object) -> None:
        self.listeners[event] = listener

    async def send(self, method: str, params: dict | None = None) -> dict:
        self.calls.append((method, params))
        if method == "Page.startScreencast":
            listener = self.listeners["Page.screencastFrame"]
            asyncio.get_running_loop().call_soon(
                listener,  # type: ignore[arg-type]
                {"data": "jpeg-frame", "sessionId": 7},
            )
        return {}

    async def detach(self) -> None:
        self.detached = True


class FakeCDPContext(FakeAuthContext):
    def __init__(self, state: dict, cdp: FakeCDPSession):
        super().__init__(state)
        self.cdp = cdp

    async def new_cdp_session(self, _page: object) -> FakeCDPSession:
        return self.cdp


class FakeStreamSocket(FakeAuthSocket):
    def __init__(self):
        super().__init__()
        self.frame_sent = asyncio.Event()
        self.accepted = False
        self.accepted_subprotocol: str | None = None
        self.closed = False

    async def accept(self, *, subprotocol: str | None = None) -> None:
        self.accepted = True
        self.accepted_subprotocol = subprotocol

    async def receive_json(self) -> dict:
        await self.frame_sent.wait()
        return {"type": "close"}

    async def send_json(self, message: dict) -> None:
        await super().send_json(message)
        if message.get("type") == "frame":
            self.frame_sent.set()

    async def close(self, **_kwargs: object) -> None:
        self.closed = True


def auth_state(value: str) -> dict:
    return {
        "cookies": [
            {
                "name": "sessionid",
                "value": value,
                "domain": ".dongchedi.com",
                "path": "/",
            }
        ],
        "origins": [],
    }


class AppCase(unittest.TestCase):
    """每个用例使用独立数据库和持久文件目录。"""

    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        root = Path(self.temp.name)
        settings = Settings(
            database_url=f"sqlite:///{(root / 'test.db').as_posix()}",
            data_dir=root / "data",
            start_background_services=False,
        )
        self.app = create_app(settings)
        self.client_context = TestClient(self.app)
        self.client = self.client_context.__enter__()
        self.container = self.app.state.container

    def tearDown(self) -> None:
        self.client_context.__exit__(None, None, None)
        self.temp.cleanup()

    def save_verified_circle(self, *, external_id: str = "24729", name: str = "风云A9") -> Circle:
        with self.container.sessions.begin() as db:
            vehicle = Vehicle(name=name)
            db.add(vehicle)
            db.flush()
            circle = Circle(
                platform_code="dongchedi",
                external_id=external_id,
                name=f"{name}车友圈",
                url=f"https://www.dongchedi.com/community/{external_id}",
                vehicle_id=vehicle.id,
                source_kind="configured",
                validation_status="verified",
            )
            db.add(circle)
            db.flush()
            return circle


class ApiAndConfigTests(AppCase):
    def test_historical_results_use_current_source_name_and_include_list_order(self) -> None:
        circle = self.save_verified_circle(name="A9L")
        run = self.container.runs.create_manual(
            ManualRunCreate(platform_code="dongchedi", circle_ids=[circle.id], quantity=1),
            scope="api",
            header_key="live-source-name-test-0001",
        )
        with self.container.sessions.begin() as db:
            task = db.scalar(select(CircleTask).where(CircleTask.run_id == run["id"]))
            self.assertIsNotNone(task)
            assert task is not None
            record = sample_record("live-source-name-post")
            db.add(
                PostSnapshot(
                    run_id=run["id"],
                    circle_task_id=task.id,
                    platform_post_id=record["platform_post_id"],
                    url=record["url"],
                    title=record["title"],
                    visibility="visible",
                    order_index=0,
                )
            )
            configured_circle = db.get(Circle, circle.id)
            assert configured_circle is not None and configured_circle.vehicle_id
            db.get(Vehicle, configured_circle.vehicle_id).name = "A9"

        current_run = self.client.get(f"/api/v1/runs/{run['id']}").json()
        posts = self.client.get(f"/api/v1/runs/{run['id']}/posts").json()["items"]
        self.assertEqual(["A9"], current_run["source_names"])
        self.assertEqual("A9", current_run["tasks"][0]["source_name"])
        self.assertEqual("A9", posts[0]["source_name"])
        self.assertEqual("latest_reply", posts[0]["list_order"])
        self.assertEqual("最新回复", posts[0]["list_order_name"])

    def test_media_resolve_returns_fresh_deduplicated_urls_without_mutating_snapshot(self) -> None:
        """按需刷新只返回临时 URL；路径签名不同的同一视频只保留一条。"""

        circle = self.save_verified_circle(name="A9L")
        run = self.container.runs.create_manual(
            ManualRunCreate(platform_code="dongchedi", circle_ids=[circle.id], quantity=1),
            scope="api",
            header_key="media-resolve-test-0001",
        )
        stable_tail = "/video/tos/cn/tos-cn-v-4eff5f/video-content-id/"
        old_urls = [
            f"https://v26-microapp-dcar.dcarvod.com/{signature}/{expiry}{stable_tail}?token={index}"
            for index, (signature, expiry) in enumerate(
                (("a" * 32, "6a86c4ac"), ("b" * 32, "6a86c4ab")),
                start=1,
            )
        ]
        fresh_urls = [
            f"https://v26-microapp-dcar.dcarvod.com/{signature}/7fffffff{stable_tail}?token={index}"
            for index, signature in enumerate(("c" * 32, "d" * 32), start=1)
        ]
        record = sample_record("media-resolve-1")
        record["video_urls"] = old_urls
        record["raw_status"] = {"video_id": "video-content-id"}
        with self.container.sessions.begin() as db:
            task = db.scalar(select(CircleTask).where(CircleTask.run_id == run["id"]))
            assert task is not None
            self.container.worker._store_records(db, task, [record])
            post = db.scalar(select(PostSnapshot).where(PostSnapshot.run_id == run["id"]))
            assert post is not None
            post_id = post.id

        with (
            patch(
                "threadsnap.collectors.dongchedi.DongchediCollector.resolve_video_urls",
                return_value=fresh_urls,
            ) as resolved,
            patch("threadsnap.worker.sync_playwright") as browser_started,
        ):
            response = self.client.post(f"/api/v1/runs/{run['id']}/posts/{post_id}/media/resolve")
            cached = self.client.post(f"/api/v1/runs/{run['id']}/posts/{post_id}/media/resolve")

        self.assertEqual(200, response.status_code, response.text)
        self.assertEqual([fresh_urls[0]], response.json()["video_urls"])
        self.assertEqual(
            [f"/api/v1/runs/{run['id']}/posts/{post_id}/media/play/0"],
            response.json()["playback_urls"],
        )
        self.assertEqual("live_url", response.json()["source"])
        self.assertIsNotNone(response.json()["expires_at"])
        self.assertEqual(response.json(), cached.json())
        self.assertEqual(1, resolved.call_count)
        browser_started.assert_not_called()
        playback = self.client.get(response.json()["playback_urls"][0], follow_redirects=False)
        self.assertEqual(307, playback.status_code)
        self.assertEqual(fresh_urls[0], playback.headers["location"])
        self.assertEqual("no-referrer", playback.headers["referrer-policy"])
        self.assertEqual("no-store", playback.headers["cache-control"])
        with self.container.sessions() as db:
            stored = db.get(PostSnapshot, post_id)
            assert stored is not None
            self.assertEqual(old_urls, stored.video_urls)

    def test_sentiment_normalizes_provider_shape_and_deduplicates_signed_media(self) -> None:
        """兼容已观察到的提供方形状偏差，同时只提交一次同一稳定媒体。"""

        first = "https://media.example.test/video.mp4?token=first"
        second = "https://media.example.test/video.mp4?token=second"
        video_hash = hashlib.sha256(b"https://media.example.test/video.mp4").hexdigest()
        post = PostSnapshot(
            title="提车记录",
            content="风云A9 提车仪式",
            image_urls=[],
            video_urls=[first, second],
        )
        payload = {
            "subject_relevance": True,
            "matched_subjects": [{"brand": "奇瑞", "product": "风云A9"}],
            "sentiment": "non_negative",
            "primary_category": "product_showcase",
            "secondary_categories": ["other"],
            "modalities": {
                "text": {"status": "processed", "evidence": "标题与正文描述提车"},
                "image": {
                    "status": "skipped",
                    "expected_count": 0,
                    "processed_count": 0,
                    "items": [],
                },
                "video_visual": {
                    "status": "processed",
                    "expected_count": 1,
                    "processed_count": 1,
                    "items": [
                        {
                            "input_index": 1,
                            "url_hash": video_hash,
                            "status": "processed",
                            "evidence": "视频展示提车仪式",
                        }
                    ],
                },
                "video_audio": {
                    "status": "processed",
                    "expected_count": 1,
                    "processed_count": 1,
                    "items": [
                        {
                            "input_index": 1,
                            "url_hash": video_hash,
                            "status": "processed",
                            "evidence": "音频已处理",
                        }
                    ],
                },
            },
            "summary": "内容为风云A9提车分享。",
        }

        normalized, changed = normalize_feedback_payload(payload, post)
        feedback = SentimentFeedback.model_validate(normalized)
        validate_modality_identity(feedback, post)

        self.assertTrue(changed)
        self.assertEqual(["标题与正文描述提车"], feedback.modalities.text.evidence)
        self.assertEqual("absent", feedback.modalities.image.status)
        self.assertEqual(0, feedback.modalities.video_visual.items[0].input_index)
        self.assertEqual([first], deduplicate_media_urls(post.video_urls))

        config = SimpleNamespace(
            model_code="qwen3.5-omni-plus-2026-03-15",
            brand="奇瑞",
            products=["风云A9"],
            supplement=None,
        )
        request = build_request(post, config)
        videos = [item for item in request["messages"][0]["content"] if item["type"] == "video_url"]
        self.assertEqual([first], [item["video_url"]["url"] for item in videos])

    def test_sentiment_normalizes_relevant_media_item_as_processed(self) -> None:
        """汇总计数明确完成时，兼容提供方把逐图状态写成 relevant。"""

        image_url = "https://media.example.test/a.jpg"
        image_hash = hashlib.sha256(image_url.encode()).hexdigest()
        post = PostSnapshot(title="风云A9", image_urls=[image_url], video_urls=[])
        payload = {
            "subject_relevance": True,
            "matched_subjects": ["风云A9"],
            "sentiment": "non_negative",
            "primary_category": None,
            "secondary_categories": [],
            "modalities": {
                "text": {"status": "present", "evidence": ["标题提及车型"]},
                "image": {
                    "status": "present",
                    "expected_count": 1,
                    "processed_count": 1,
                    "items": [
                        {
                            "input_index": 0,
                            "url_hash": "model-copied-wrong-hash",
                            "status": "relevant",
                            "evidence": ["图片展示车辆"],
                        }
                    ],
                },
                "video_visual": {
                    "status": "absent",
                    "expected_count": 0,
                    "processed_count": 0,
                    "items": [],
                },
                "video_audio": {
                    "status": "absent",
                    "expected_count": 0,
                    "processed_count": 0,
                    "items": [],
                },
            },
            "summary": "内容为车型展示。",
        }

        normalized, changed = normalize_feedback_payload(payload, post)
        feedback = SentimentFeedback.model_validate(normalized)
        validate_modality_identity(feedback, post)

        self.assertTrue(changed)
        self.assertEqual(["风云A9"], feedback.matched_subjects)
        self.assertIsNone(feedback.primary_category)
        self.assertEqual([], feedback.secondary_categories)
        self.assertEqual("processed", feedback.modalities.text.status)
        self.assertEqual("processed", feedback.modalities.image.status)
        self.assertEqual("processed", feedback.modalities.image.items[0].status)
        self.assertEqual(image_hash, feedback.modalities.image.items[0].url_hash)

    def test_sentiment_worker_starts_two_bounded_consumers(self) -> None:
        worker = self.container.sentiment_worker
        worker.start()
        try:
            self.assertEqual(2, worker.concurrency)
            self.assertEqual(2, len(worker.threads))
            self.assertTrue(all(thread.is_alive() for thread in worker.threads))
        finally:
            worker.stop()

    def test_local_text_sentiment_skips_cloud_and_media_resolution(self) -> None:
        """验证本地模型可无密钥启用，且只消费标题和正文。"""

        class FakeLocalAnalyzer:
            def __init__(self) -> None:
                self.validations = 0
                self.analyses = 0

            def validate(self, subject: dict) -> int:
                self.validations += 1
                self.assert_subject = subject
                return 12

            def analyze(self, **values):
                self.analyses += 1
                self.assert_values = values

                def skipped(count: int) -> dict:
                    return {
                        "status": "not_requested" if count else "absent",
                        "expected_count": count,
                        "processed_count": 0,
                        "items": [],
                    }

                return (
                    {
                        "subject_relevance": True,
                        "matched_subjects": ["风云A9"],
                        "sentiment": "negative",
                        "primary_category": "product_criticism",
                        "secondary_categories": [],
                        "modalities": {
                            "text": {"status": "processed", "evidence": ["车机太卡了。"]},
                            "image": skipped(values["image_count"]),
                            "video_visual": skipped(values["video_count"]),
                            "video_audio": skipped(values["video_count"]),
                        },
                        "summary": "本地文字模型识别到风云A9车机卡顿的负面反馈。",
                    },
                    '{"provider":"local-test"}',
                    23,
                )

        fake_local = FakeLocalAnalyzer()
        self.container.sentiment.local_analyzer = fake_local
        config = self.client.get("/api/v1/sentiment/config").json()
        saved = self.client.put(
            "/api/v1/sentiment/config",
            json={
                "revision": config["revision"],
                "enabled": False,
                "api_base_url": "",
                "model_code": "paddlenlp-local-text-nano-v1",
                "subject": {
                    "brand": "奇瑞",
                    "products": ["风云A9"],
                    "supplement": "",
                },
            },
        )
        self.assertEqual(200, saved.status_code, saved.text)
        self.assertEqual("local", saved.json()["model_provider"])
        self.assertEqual("text_only", saved.json()["model_input_mode"])
        self.assertFalse(saved.json()["api_key_configured"])

        tested = self.client.post("/api/v1/sentiment/config/test")
        self.assertEqual(200, tested.status_code, tested.text)
        self.assertEqual(1, fake_local.validations)
        current = self.client.get("/api/v1/sentiment/config").json()
        enabled = self.client.put(
            "/api/v1/sentiment/config",
            json={
                "revision": current["revision"],
                "enabled": True,
                "api_base_url": current["api_base_url"],
                "model_code": current["model_code"],
                "subject": {
                    "brand": current["subject"]["brand"],
                    "products": current["subject"]["products"],
                    "supplement": current["subject"]["supplement"],
                },
            },
        )
        self.assertEqual(200, enabled.status_code, enabled.text)
        self.assertTrue(enabled.json()["enabled"])

        circle = self.save_verified_circle(name="风云A9")
        run = self.client.post(
            "/api/v1/runs/manual",
            json={
                "platform_code": "dongchedi",
                "circle_ids": [circle.id],
                "quantity": 1,
                "idempotency_key": "local-sentiment-test-run",
            },
        ).json()
        record = sample_record("local-sentiment-1")
        record.update(
            title="风云A9车机太卡",
            content="风云A9的车机太卡了。",
            image_urls=["https://example.test/local.jpg"],
            video_urls=["https://example.test/local.mp4"],
        )
        with self.container.sessions.begin() as db:
            task = db.scalar(select(CircleTask).where(CircleTask.run_id == run["id"]))
            assert task is not None
            self.container.worker._store_records(db, task, [record])

        def unexpected_media_resolution(*_args, **_kwargs):
            raise AssertionError("本地文字模型不应解析视频播放地址")

        class UnexpectedCloudClient:
            def request(self, *_args, **_kwargs):
                raise AssertionError("本地文字模型不应请求云端 API")

        self.container.sentiment.client = UnexpectedCloudClient()
        self.container.sentiment_worker.media_resolver = unexpected_media_resolution
        self.assertTrue(self.container.sentiment_worker.process_once())
        self.assertEqual(1, fake_local.analyses)
        self.assertEqual(1, fake_local.assert_values["image_count"])
        self.assertEqual(1, fake_local.assert_values["video_count"])
        with self.container.sessions() as db:
            analysis = db.scalar(select(SentimentAnalysis))
            assert analysis is not None
            self.assertEqual("analysis_completed", analysis.status)
            self.assertEqual("negative", analysis.result)
            self.assertEqual("not_requested", analysis.modalities["image"]["status"])
            self.assertEqual("not_requested", analysis.modalities["video_visual"]["status"])
            self.assertEqual({"provider": "local", "billable_tokens": 0}, analysis.usage)

    def test_local_sentiment_input_hash_excludes_media(self) -> None:
        post = PostSnapshot(
            title="风云A9",
            content="车机卡顿",
            image_urls=["https://example.test/first.jpg"],
            video_urls=["https://example.test/first.mp4"],
        )
        local_before = sentiment_input_hash(post, "paddlenlp-local-text-nano-v1")
        hosted_before = sentiment_input_hash(post, HOSTED_MODEL_CODE)
        post.image_urls = ["https://example.test/second.jpg"]
        post.video_urls = ["https://example.test/second.mp4"]
        self.assertEqual(
            local_before,
            sentiment_input_hash(post, "paddlenlp-local-text-nano-v1"),
        )
        self.assertNotEqual(hosted_before, sentiment_input_hash(post, HOSTED_MODEL_CODE))

    def test_deepseek_text_model_uses_independent_connection_and_skips_media(self) -> None:
        """验证 DeepSeek 使用独立凭证、纯文字请求和统一结果合同。"""

        initial = self.client.get("/api/v1/sentiment/config").json()
        qwen_saved = self.client.put(
            "/api/v1/sentiment/config",
            json={
                "revision": initial["revision"],
                "enabled": False,
                "api_base_url": "https://qwen.example.test/compatible-mode/v1",
                "api_key": "qwen-test-secret",
                "model_code": HOSTED_MODEL_CODE,
                "subject": {
                    "brand": "奇瑞",
                    "products": ["风云A9"],
                    "supplement": "",
                },
            },
        )
        self.assertEqual(200, qwen_saved.status_code, qwen_saved.text)
        self.assertTrue(
            qwen_saved.json()["model_connections"][HOSTED_MODEL_CODE]["api_key_configured"]
        )

        deepseek_saved = self.client.put(
            "/api/v1/sentiment/config",
            json={
                "revision": qwen_saved.json()["revision"],
                "enabled": False,
                "api_base_url": "https://api.deepseek.com",
                "api_key": "deepseek-test-secret",
                "model_code": DEEPSEEK_MODEL_CODE,
                "subject": {
                    "brand": "奇瑞",
                    "products": ["风云A9"],
                    "supplement": "",
                },
            },
        )
        self.assertEqual(200, deepseek_saved.status_code, deepseek_saved.text)
        deepseek_config = deepseek_saved.json()
        self.assertEqual("hosted", deepseek_config["model_provider"])
        self.assertEqual("text_only", deepseek_config["model_input_mode"])
        self.assertEqual("https://api.deepseek.com", deepseek_config["api_base_url"])
        self.assertTrue(deepseek_config["api_key_configured"])
        self.assertTrue(
            deepseek_config["model_connections"][HOSTED_MODEL_CODE]["api_key_configured"]
        )
        self.assertNotIn("qwen-test-secret", deepseek_saved.text)
        self.assertNotIn("deepseek-test-secret", deepseek_saved.text)

        class FakeDeepSeekClient:
            def __init__(self) -> None:
                self.requests: list[dict] = []

            def request(self, base_url: str, api_key: str, body: dict):
                self.assert_connection = (base_url, api_key)
                self.requests.append(body)
                content = body["messages"][0]["content"]
                if '"ok":true' in content:
                    return '{"ok":true}', {"total_tokens": 4}, "deepseek-test", 6
                payload = {
                    "subject_relevance": True,
                    "matched_subjects": ["风云A9"],
                    "sentiment": "negative",
                    "primary_category": "product_criticism",
                    "secondary_categories": [],
                    "modalities": {
                        "text": {
                            "status": "processed",
                            "evidence": ["标题和正文反馈风云A9车机卡顿。"],
                        },
                        # 即使提供方声称处理了媒体，后端也必须按真实输入覆盖为未参与。
                        "image": {
                            "status": "processed",
                            "expected_count": 1,
                            "processed_count": 1,
                            "items": [],
                        },
                        "video_visual": {
                            "status": "processed",
                            "expected_count": 1,
                            "processed_count": 1,
                            "items": [],
                        },
                        "video_audio": {
                            "status": "speech",
                            "expected_count": 1,
                            "processed_count": 1,
                            "items": [],
                        },
                    },
                    "summary": "文字内容反馈风云A9车机卡顿，情感为负面。",
                }
                return (
                    json.dumps(payload, ensure_ascii=False),
                    {"prompt_tokens": 80, "completion_tokens": 60, "total_tokens": 140},
                    "deepseek-analysis",
                    16,
                )

        fake_client = FakeDeepSeekClient()
        self.container.sentiment.client = fake_client
        with patch(
            "threadsnap.sentiment.validate_public_https_base_url",
            side_effect=lambda value, resolve: value.rstrip("/"),
        ):
            tested = self.client.post("/api/v1/sentiment/config/test")
        self.assertEqual(200, tested.status_code, tested.text)
        self.assertEqual(
            ("https://api.deepseek.com", "deepseek-test-secret"),
            fake_client.assert_connection,
        )
        test_body = fake_client.requests[0]
        self.assertIsInstance(test_body["messages"][0]["content"], str)
        self.assertEqual({"type": "disabled"}, test_body["thinking"])
        self.assertNotIn("modalities", test_body)

        current = self.client.get("/api/v1/sentiment/config").json()
        enabled = self.client.put(
            "/api/v1/sentiment/config",
            json={
                "revision": current["revision"],
                "enabled": True,
                "api_base_url": current["api_base_url"],
                "model_code": current["model_code"],
                "subject": {
                    "brand": current["subject"]["brand"],
                    "products": current["subject"]["products"],
                    "supplement": current["subject"]["supplement"],
                },
            },
        )
        self.assertEqual(200, enabled.status_code, enabled.text)
        self.assertTrue(enabled.json()["enabled"])

        circle = self.save_verified_circle(name="风云A9")
        run = self.client.post(
            "/api/v1/runs/manual",
            json={
                "platform_code": "dongchedi",
                "circle_ids": [circle.id],
                "quantity": 1,
                "idempotency_key": "deepseek-text-sentiment-run",
            },
        ).json()
        record = sample_record("deepseek-text-1")
        record.update(
            title="风云A9车机卡顿",
            content="风云A9车机升级后依然卡顿。",
            image_urls=["https://media.example.test/secret-image.jpg"],
            video_urls=["https://media.example.test/secret-video.mp4"],
        )
        with self.container.sessions.begin() as db:
            task = db.scalar(select(CircleTask).where(CircleTask.run_id == run["id"]))
            assert task is not None
            self.container.worker._store_records(db, task, [record])

        def unexpected_media_resolution(*_args, **_kwargs):
            raise AssertionError("DeepSeek 文字模型不应解析视频播放地址")

        self.container.sentiment_worker.media_resolver = unexpected_media_resolution
        with patch(
            "threadsnap.sentiment.validate_public_https_base_url",
            side_effect=lambda value, resolve: value.rstrip("/"),
        ):
            self.assertTrue(self.container.sentiment_worker.process_once())
        analysis_body = fake_client.requests[-1]
        text_input = analysis_body["messages"][0]["content"]
        self.assertIsInstance(text_input, str)
        self.assertNotIn("secret-image.jpg", text_input)
        self.assertNotIn("secret-video.mp4", text_input)
        self.assertNotIn("image_url", json.dumps(analysis_body))
        self.assertNotIn("video_url", json.dumps(analysis_body))
        with self.container.sessions() as db:
            analysis = db.scalar(select(SentimentAnalysis))
            assert analysis is not None
            self.assertEqual("analysis_completed", analysis.status)
            self.assertEqual("negative", analysis.result)
            self.assertEqual("not_requested", analysis.modalities["image"]["status"])
            self.assertEqual("not_requested", analysis.modalities["video_visual"]["status"])
            self.assertEqual("not_requested", analysis.modalities["video_audio"]["status"])

        latest = self.client.get("/api/v1/sentiment/config").json()
        qwen_restored = self.client.put(
            "/api/v1/sentiment/config",
            json={
                "revision": latest["revision"],
                "enabled": False,
                "api_base_url": latest["model_connections"][HOSTED_MODEL_CODE][
                    "api_base_url"
                ],
                "model_code": HOSTED_MODEL_CODE,
                "subject": {
                    "brand": latest["subject"]["brand"],
                    "products": latest["subject"]["products"],
                    "supplement": latest["subject"]["supplement"],
                },
            },
        )
        self.assertEqual(200, qwen_restored.status_code, qwen_restored.text)
        self.assertEqual(
            "https://qwen.example.test/compatible-mode/v1",
            qwen_restored.json()["api_base_url"],
        )
        self.assertTrue(qwen_restored.json()["api_key_configured"])

    def test_deepseek_text_input_hash_excludes_media(self) -> None:
        post = PostSnapshot(
            title="风云A9",
            content="车机卡顿",
            image_urls=["https://example.test/first.jpg"],
            video_urls=["https://example.test/first.mp4"],
        )
        before = sentiment_input_hash(post, DEEPSEEK_MODEL_CODE)
        post.image_urls = ["https://example.test/second.jpg"]
        post.video_urls = ["https://example.test/second.mp4"]
        self.assertEqual(before, sentiment_input_hash(post, DEEPSEEK_MODEL_CODE))

    def test_deepseek_text_normalization_marks_media_not_requested(self) -> None:
        post = PostSnapshot(
            title="风云A9",
            image_urls=["https://example.test/a.jpg"],
            video_urls=["https://example.test/a.mp4"],
        )
        payload = {
            "modalities": {
                "text": {"status": "processed", "evidence": ["标题提及风云A9。"]},
            }
        }
        normalized, changed = normalize_text_only_feedback_payload(payload, post)
        self.assertTrue(changed)
        self.assertEqual("not_requested", normalized["modalities"]["image"]["status"])
        self.assertEqual(1, normalized["modalities"]["image"]["expected_count"])
        self.assertEqual([], normalized["modalities"]["video_audio"]["items"])

    def test_bootstrap_refreshes_available_adapter_version(self) -> None:
        with self.container.sessions.begin() as db:
            platform = db.get(PlatformConfig, "dongchedi")
            assert platform is not None
            platform.adapter_version = "dongchedi-dynamic-v1"

        with self.container.sessions.begin() as db:
            bootstrap_database(db)

        with self.container.sessions() as db:
            platform = db.get(PlatformConfig, "dongchedi")
            assert platform is not None
        self.assertEqual(ADAPTER_VERSION, platform.adapter_version)

    def test_sentiment_config_worker_filters_detail_and_manual_revision(self) -> None:
        """覆盖配置、单次分析、三态列表、详情依据和人工优先的组合路径。"""

        class FakeSentimentClient:
            def __init__(self) -> None:
                self.analysis_requests = 0

            def request(self, _base_url: str, _key: str, body: dict):
                text = body["messages"][0]["content"][-1]["text"]
                if "ok" in text:
                    return '{"ok":true}', None, "test-request", 5
                self.analysis_requests += 1
                if self.analysis_requests == 1:
                    return (
                        '{"subject_relevance":',
                        {"total_tokens": 17},
                        "truncated-request",
                        7,
                    )
                feedback = {
                    "subject_relevance": True,
                    "matched_subjects": ["A9L"],
                    "sentiment": "negative",
                    "primary_category": "product_complaint",
                    "secondary_categories": [],
                    "modalities": {
                        "text": {"status": "processed", "evidence": "正文描述产品故障"},
                        "image": {
                            "status": "processed",
                            "expected_count": 1,
                            "processed_count": 1,
                            "items": [
                                {
                                    "input_index": 1,
                                    "url_hash": hashlib.sha256(
                                        b"https://example.test/a.jpg"
                                    ).hexdigest(),
                                    "status": "processed",
                                    "evidence": "图片显示故障提示",
                                }
                            ],
                        },
                        "video_visual": {
                            "status": "skipped",
                            "expected_count": 0,
                            "processed_count": 0,
                            "items": [],
                        },
                        "video_audio": {
                            "status": "skipped",
                            "expected_count": 0,
                            "processed_count": 0,
                            "items": [],
                        },
                    },
                    "summary": "帖子反馈 A9L 产品故障。",
                }
                return (
                    json.dumps(feedback, ensure_ascii=False),
                    {"total_tokens": 123},
                    "model-request",
                    18,
                )

        config = self.client.get("/api/v1/sentiment/config").json()
        self.assertFalse(config["api_key_configured"])
        saved = self.client.put(
            "/api/v1/sentiment/config",
            json={
                "revision": config["revision"],
                "enabled": False,
                "api_base_url": "https://api.example.test/v1",
                "api_key": "test-secret-key",
                "model_code": "qwen3.5-omni-plus-2026-03-15",
                "subject": {
                    "brand": "奇瑞",
                    "products": ["A9L"],
                    "supplement": "包含品牌服务反馈",
                },
            },
        )
        self.assertEqual(200, saved.status_code)
        self.assertNotIn("test-secret-key", saved.text)
        fake_sentiment_client = FakeSentimentClient()
        self.container.sentiment.client = fake_sentiment_client
        with patch(
            "threadsnap.sentiment.validate_public_https_base_url",
            side_effect=lambda value, resolve: value.rstrip("/"),
        ):
            tested = self.client.post("/api/v1/sentiment/config/test")
        self.assertEqual(200, tested.status_code)
        current = self.client.get("/api/v1/sentiment/config").json()
        enabled = self.client.put(
            "/api/v1/sentiment/config",
            json={
                "revision": current["revision"],
                "enabled": True,
                "api_base_url": current["api_base_url"],
                "model_code": current["model_code"],
                "subject": {
                    "brand": current["subject"]["brand"],
                    "products": current["subject"]["products"],
                    "supplement": current["subject"]["supplement"],
                },
            },
        )
        self.assertEqual(200, enabled.status_code)

        circle = self.save_verified_circle(name="A9L")
        run_response = self.client.post(
            "/api/v1/runs/manual",
            json={
                "platform_code": "dongchedi",
                "circle_ids": [circle.id],
                "quantity": 1,
                "idempotency_key": "sentiment-test-run",
            },
        )
        self.assertEqual(202, run_response.status_code, run_response.text)
        run = run_response.json()
        with self.container.sessions.begin() as db:
            task = db.scalar(select(CircleTask).where(CircleTask.run_id == run["id"]))
            assert task is not None
            self.container.worker._store_records(db, task, [sample_record("sentiment-1")])
            analysis = db.scalar(select(SentimentAnalysis))
            assert analysis is not None
            self.assertEqual("analysis_queued", analysis.status)
        with patch(
            "threadsnap.sentiment.validate_public_https_base_url",
            side_effect=lambda value, resolve: value.rstrip("/"),
        ):
            self.assertTrue(self.container.sentiment_worker.process_once())
        self.assertEqual(2, fake_sentiment_client.analysis_requests)
        with self.container.sessions() as db:
            retried_analysis = db.scalar(select(SentimentAnalysis))
            assert retried_analysis is not None
            self.assertEqual(1, retried_analysis.retry_count)
            self.assertEqual(1, len(retried_analysis.attempt_failures))
            self.assertEqual(
                "MODEL_STREAM_INCOMPLETE",
                retried_analysis.attempt_failures[0]["error_code"],
            )

        listed = self.client.get(
            f"/api/v1/runs/{run['id']}/posts",
            params={"sentiment_result": "negative", "analysis_status": "analysis_completed"},
        ).json()
        self.assertEqual(1, listed["total"])
        post = listed["items"][0]
        self.assertEqual("negative", post["sentiment_result"])
        self.assertEqual("ai", post["sentiment_source"])
        detail = self.client.get(f"/api/v1/runs/{run['id']}/posts/{post['id']}").json()
        self.assertEqual("帖子反馈 A9L 产品故障。", detail["sentiment"]["summary"])
        self.assertNotIn("raw_response", detail["sentiment"])

        revised = self.client.post(
            f"/api/v1/runs/{run['id']}/posts/{post['id']}/sentiment/manual-revisions",
            json={
                "action": "set_result",
                "result": "non_negative",
                "secondary_categories": [],
                "note": "人工核对",
            },
        )
        self.assertEqual(200, revised.status_code)
        self.assertEqual("manual", revised.json()["source"])
        restored = self.client.post(
            f"/api/v1/runs/{run['id']}/posts/{post['id']}/sentiment/manual-revisions",
            json={"action": "restore_ai", "secondary_categories": []},
        )
        self.assertEqual(200, restored.status_code)
        self.assertEqual("ai", restored.json()["source"])

        # 启用中替换连接配置必须先保存并自动关闭，避免旧验证状态继续消费模型。
        active_config = self.client.get("/api/v1/sentiment/config").json()
        rotated = self.client.put(
            "/api/v1/sentiment/config",
            json={
                "revision": active_config["revision"],
                "enabled": True,
                "api_base_url": active_config["api_base_url"],
                "api_key": "rotated-test-secret",
                "model_code": active_config["model_code"],
                "subject": {
                    "brand": active_config["subject"]["brand"],
                    "products": active_config["subject"]["products"],
                    "supplement": active_config["subject"]["supplement"],
                },
            },
        )
        self.assertEqual(200, rotated.status_code)
        self.assertFalse(rotated.json()["enabled"])
        self.assertEqual("unverified", rotated.json()["validation_status"])
        self.assertNotIn("rotated-test-secret", rotated.text)

        # 关闭期间的新快照保持禁用，不得绕过开关复用同内容的历史 AI 或人工结论。
        second_run_response = self.client.post(
            "/api/v1/runs/manual",
            json={
                "platform_code": "dongchedi",
                "circle_ids": [circle.id],
                "quantity": 1,
                "idempotency_key": "sentiment-disabled-run",
            },
        )
        self.assertEqual(202, second_run_response.status_code)
        second_run = second_run_response.json()
        with self.container.sessions.begin() as db:
            second_task = db.scalar(select(CircleTask).where(CircleTask.run_id == second_run["id"]))
            assert second_task is not None
            self.container.worker._store_records(db, second_task, [sample_record("sentiment-1")])
            second_post = db.scalar(
                select(PostSnapshot).where(PostSnapshot.run_id == second_run["id"])
            )
            assert second_post is not None
            second_analysis = db.scalar(
                select(SentimentAnalysis).where(SentimentAnalysis.post_id == second_post.id)
            )
            assert second_analysis is not None
            self.assertEqual("analysis_disabled", second_analysis.status)
            self.assertIsNone(second_post.sentiment_result)

        # 运行配置失效后继续入库的快照必须暂停，修复并启用后才能自动恢复；
        # 不能把它们误记为用户主动关闭分析时产生的禁用任务。
        with self.container.sessions.begin() as db:
            invalid_config = self.container.sentiment.ensure_default(db)
            invalid_config.validation_status = "invalid"
            invalid_config.validation_error = "模拟运行期间配置失效"
        third_run_response = self.client.post(
            "/api/v1/runs/manual",
            json={
                "platform_code": "dongchedi",
                "circle_ids": [circle.id],
                "quantity": 1,
                "idempotency_key": "sentiment-paused-run",
            },
        )
        self.assertEqual(202, third_run_response.status_code)
        third_run = third_run_response.json()
        with self.container.sessions.begin() as db:
            third_task = db.scalar(select(CircleTask).where(CircleTask.run_id == third_run["id"]))
            assert third_task is not None
            self.container.worker._store_records(db, third_task, [sample_record("sentiment-2")])
            third_post = db.scalar(
                select(PostSnapshot).where(PostSnapshot.run_id == third_run["id"])
            )
            assert third_post is not None
            third_analysis = db.scalar(
                select(SentimentAnalysis).where(SentimentAnalysis.post_id == third_post.id)
            )
            assert third_analysis is not None
            self.assertEqual("analysis_paused", third_analysis.status)
            self.assertEqual("MODEL_CONFIG_ERROR", third_analysis.error_code)
            self.assertEqual("analysis_paused", third_post.analysis_status)

    def test_internal_api_rejects_non_loopback_client(self) -> None:
        request = SimpleNamespace(client=SimpleNamespace(host="10.20.30.40"))
        with self.assertRaises(DomainError) as raised:
            require_internal_loopback(request)
        self.assertEqual(403, raised.exception.status_code)
        self.assertIn("本机", raised.exception.message)

    def test_platform_clamping_plan_and_chinese_validation(self) -> None:
        circle = self.save_verified_circle()
        response = self.client.put(
            "/api/v1/platforms/dongchedi",
            json={"enabled": True, "internal_concurrency": 99},
        )
        self.assertEqual(200, response.status_code)
        self.assertEqual(8, response.json()["internal_concurrency"])
        self.assertEqual(1, len(response.json()["notes"]))

        disabled = self.client.put(
            "/api/v1/platforms/autohome",
            json={"enabled": True, "internal_concurrency": 1},
        )
        self.assertEqual(409, disabled.status_code)
        self.assertIn("暂未接入", disabled.json()["message"])

        plan = self.client.put(
            "/api/v1/extraction-plan",
            json={
                "revision": 1,
                "rules": [
                    {
                        "id": "rule-0001",
                        "name": "工作日规则",
                        "platform_quantities": {"dongchedi": 30},
                        "circle_ids": [circle.id],
                    }
                ],
                "nodes": [
                    {
                        "id": "node-0001",
                        "weekdays": [4, 0, 0],
                        "time": "17:00:05",
                        "enabled": True,
                        "rule_ids": ["rule-0001"],
                    }
                ],
            },
        )
        self.assertEqual(200, plan.status_code)
        self.assertEqual([0, 4], plan.json()["nodes"][0]["weekdays"])
        self.assertEqual("17:00:05", plan.json()["nodes"][0]["time"])
        self.assertEqual([circle.id], plan.json()["rules"][0]["circle_ids"])
        invalid = self.client.put(
            "/api/v1/extraction-plan",
            json={
                "revision": 2,
                "rules": plan.json()["rules"],
                "nodes": [{**plan.json()["nodes"][0], "time": "25:99"}],
            },
        )
        self.assertEqual(422, invalid.status_code)
        self.assertEqual("请求参数校验失败。", invalid.json()["message"])
        self.assertTrue(
            all(item["reason"] == "字段格式或取值无效" for item in invalid.json()["details"])
        )

    def test_circle_batch_is_atomic_and_requires_validation_before_auto_enable(
        self,
    ) -> None:
        duplicate = CircleBatchUpdate(
            rows=[
                CircleRow(
                    platform_code="dongchedi",
                    url="https://www.dongchedi.com/community/24729",
                    vehicle_name="A9",
                ),
                CircleRow(
                    platform_code="dongchedi",
                    url="https://www.dongchedi.com/community/24729",
                    vehicle_name="A9L",
                ),
            ]
        )
        with self.assertRaises(DomainError):
            self.container.config.save_circle_batch(duplicate)
        with self.container.sessions() as db:
            self.assertEqual(0, len(list(db.scalars(select(Circle)))))
            self.assertEqual(0, len(list(db.scalars(select(Vehicle)))))

        created = self.container.config.save_circle_batch(
            CircleBatchUpdate(
                rows=[
                    CircleRow(
                        platform_code="dongchedi",
                        url="https://www.dongchedi.com/community/24729",
                        vehicle_name="A9",
                    )
                ]
            )
        )
        circle_id = created["items"][0]["id"]
        with self.assertRaises(DomainError) as raised:
            self.container.config.save_circle_batch(
                CircleBatchUpdate(
                    rows=[
                        CircleRow(
                            id=circle_id,
                            platform_code="dongchedi",
                            url="https://www.dongchedi.com/community/24729",
                            vehicle_id=created["items"][0]["vehicle_id"],
                            auto_enabled=True,
                        )
                    ]
                )
            )
        self.assertIn("验证通过", raised.exception.details[0]["reason"])

    def test_same_circle_can_save_latest_reply_and_latest_publish_sources(self) -> None:
        response = self.client.put(
            "/api/v1/circles/batch",
            json={
                "rows": [
                    {
                        "platform_code": "dongchedi",
                        "url": "https://www.dongchedi.com/community/24729",
                        "vehicle_name": "风云A9最新回复",
                    },
                    {
                        "platform_code": "dongchedi",
                        "url": "https://www.dongchedi.com/community/24729/dongtai-release",
                        "vehicle_name": "风云A9最新发布",
                    },
                ],
                "deleted_ids": [],
            },
        )

        self.assertEqual(200, response.status_code)
        items = response.json()["items"]
        self.assertEqual(2, len(items))
        self.assertEqual({"latest_reply", "latest_publish"}, {item["list_order"] for item in items})
        self.assertEqual(
            {
                "https://www.dongchedi.com/community/24729",
                "https://www.dongchedi.com/community/24729/dongtai-release",
            },
            {item["url"] for item in items},
        )

        plan = self.client.get("/api/v1/extraction-plan").json()
        saved_plan = self.client.put(
            "/api/v1/extraction-plan",
            json={
                "revision": plan["revision"],
                "rules": [
                    {
                        "id": "same-circle-two-feeds",
                        "name": "风云A9统一规则",
                        "platform_quantities": {"dongchedi": 1},
                        "circle_ids": [item["id"] for item in items],
                    }
                ],
                "nodes": [],
            },
        )
        self.assertEqual(200, saved_plan.status_code)
        self.assertEqual(
            {item["id"] for item in items},
            set(saved_plan.json()["rules"][0]["circle_ids"]),
        )

        self.client.put(
            "/api/v1/platforms/dongchedi",
            json={"enabled": True, "internal_concurrency": 1},
        )
        manual_runs = []
        for index, item in enumerate(items):
            response = self.client.post(
                "/api/v1/runs/manual",
                headers={"Idempotency-Key": f"one-feed-source-{index}"},
                json={
                    "platform_code": "dongchedi",
                    "circle_ids": [item["id"]],
                    "circle_urls": [],
                    "known_post_urls": [],
                    "quantity": 1,
                    "idempotency_key": f"one-feed-source-{index}",
                },
            )
            self.assertEqual(202, response.status_code)
            manual_runs.append(response.json())
        self.assertEqual(
            {"风云A9最新回复", "风云A9最新发布"},
            {run["source_names"][0] for run in manual_runs},
        )

    def test_bulk_first_validation_auto_enables_and_revalidation_preserves_manual_disable(
        self,
    ) -> None:
        created = [
            self.client.post(
                "/api/v1/circles",
                json={
                    "platform_code": "dongchedi",
                    "url": f"https://www.dongchedi.com/community/{external_id}",
                    "vehicle_name": vehicle_name,
                },
            ).json()
            for external_id, vehicle_name in [("24729", "A9"), ("8985", "A9L")]
        ]
        self.assertTrue(all(item["first_validated_at"] is None for item in created))
        self.assertTrue(all(not item["auto_enabled"] for item in created))

        queued = self.client.post("/api/v1/circles/validate-unverified")
        self.assertEqual(202, queued.status_code)
        self.assertEqual(2, queued.json()["queued_count"])
        self.assertEqual(2, queued.json()["total_count"])
        reused = self.client.post("/api/v1/circles/validate-unverified")
        self.assertEqual(0, reused.json()["queued_count"])
        self.assertEqual(2, reused.json()["reused_count"])

        with patch.object(self.container.worker, "_collector", return_value=FakeCollector()):
            self.assertTrue(self.container.worker.process_once())
            self.assertTrue(self.container.worker.process_once())
        verified = {item["id"]: item for item in self.client.get("/api/v1/circles").json()}
        self.assertTrue(all(item["validation_status"] == "verified" for item in verified.values()))
        self.assertTrue(all(item["auto_enabled"] for item in verified.values()))
        self.assertTrue(all(item["first_validated_at"] for item in verified.values()))

        target = verified[created[0]["id"]]
        first_validated_at = target["first_validated_at"]
        disabled = self.client.put(
            f"/api/v1/circles/{target['id']}",
            json={
                "platform_code": "dongchedi",
                "url": target["url"],
                "vehicle_id": target["vehicle_id"],
                "auto_enabled": False,
            },
        )
        self.assertEqual(200, disabled.status_code)
        self.assertFalse(disabled.json()["auto_enabled"])

        retried = self.client.post(f"/api/v1/circles/{target['id']}/validate")
        self.assertEqual(202, retried.status_code)
        with patch.object(self.container.worker, "_collector", return_value=FakeCollector()):
            self.assertTrue(self.container.worker.process_once())
        revalidated = self.client.get(f"/api/v1/circles/{target['id']}").json()
        self.assertEqual("verified", revalidated["validation_status"])
        self.assertEqual(first_validated_at, revalidated["first_validated_at"])
        self.assertFalse(revalidated["auto_enabled"])

    def test_circle_crud_and_explicit_batch_delete(self) -> None:
        created = self.client.post(
            "/api/v1/circles",
            json={
                "platform_code": "dongchedi",
                "url": "https://www.dongchedi.com/community/24729",
                "vehicle_name": "A9",
            },
        )
        self.assertEqual(201, created.status_code)
        circle_id = created.json()["id"]
        self.assertEqual("A9", created.json()["vehicle_name"])

        listed = self.client.get("/api/v1/circles")
        self.assertEqual(200, listed.status_code)
        self.assertEqual([circle_id], [item["id"] for item in listed.json()])
        self.assertEqual(circle_id, self.client.get(f"/api/v1/circles/{circle_id}").json()["id"])

        updated = self.client.put(
            f"/api/v1/circles/{circle_id}",
            json={
                "platform_code": "dongchedi",
                "url": "https://www.dongchedi.com/community/24729",
                "vehicle_name": "A9 Pro",
            },
        )
        self.assertEqual(200, updated.status_code)
        self.assertEqual("A9 Pro", updated.json()["vehicle_name"])

        conflict = self.client.put(
            "/api/v1/circles/batch",
            json={
                "rows": [
                    {
                        "id": circle_id,
                        "platform_code": "dongchedi",
                        "url": "https://www.dongchedi.com/community/24729",
                        "vehicle_id": updated.json()["vehicle_id"],
                    }
                ],
                "deleted_ids": [circle_id],
            },
        )
        self.assertEqual(400, conflict.status_code)
        self.assertEqual(1, len(self.client.get("/api/v1/circles").json()))

        removed = self.client.put(
            "/api/v1/circles/batch",
            json={"rows": [], "deleted_ids": [circle_id]},
        )
        self.assertEqual(200, removed.status_code)
        self.assertEqual(1, removed.json()["deleted_count"])
        self.assertEqual([], self.client.get("/api/v1/circles").json())
        self.assertTrue(
            all(not vehicle["circles"] for vehicle in self.client.get("/api/v1/vehicles").json())
        )

        second = self.client.post(
            "/api/v1/circles",
            json={
                "platform_code": "dongchedi",
                "url": "https://www.dongchedi.com/community/24730",
                "vehicle_name": "A9",
            },
        )
        self.assertEqual(201, second.status_code)
        deleted = self.client.delete(f"/api/v1/circles/{second.json()['id']}")
        self.assertEqual(200, deleted.status_code)
        self.assertEqual([], self.client.get("/api/v1/circles").json())

    def test_active_rule_scope_blocks_circle_delete_until_new_version_removes_it(self) -> None:
        circle = self.save_verified_circle()
        plan = self.client.put(
            "/api/v1/extraction-plan",
            json={
                "revision": 1,
                "rules": [
                    {
                        "id": "rule-scope-0001",
                        "name": "指定圈子规则",
                        "platform_quantities": {"dongchedi": 30},
                        "circle_ids": [circle.id],
                    }
                ],
                "nodes": [],
            },
        )
        self.assertEqual(200, plan.status_code)

        blocked = self.client.delete(f"/api/v1/circles/{circle.id}")
        self.assertEqual(400, blocked.status_code)
        self.assertIn("指定圈子规则", blocked.json()["details"][0]["reason"])

        cleared = self.client.put(
            "/api/v1/extraction-plan",
            json={
                "revision": plan.json()["revision"],
                "rules": [
                    {
                        "id": "rule-scope-0001",
                        "name": "指定圈子规则",
                        "platform_quantities": {},
                        "circle_ids": [],
                    }
                ],
                "nodes": [],
            },
        )
        self.assertEqual(200, cleared.status_code)
        self.assertEqual(2, cleared.json()["rules"][0]["version"])
        self.assertEqual(200, self.client.delete(f"/api/v1/circles/{circle.id}").status_code)

    def test_rule_scope_can_select_circles_from_multiple_integrated_platforms(self) -> None:
        dongchedi_circle = self.save_verified_circle()
        with self.container.sessions.begin() as db:
            stored = db.get(Circle, dongchedi_circle.id)
            stored.auto_enabled = True
            autohome = db.get(PlatformConfig, "autohome")
            autohome.adapter_status = "available"
            autohome.enabled = True
            vehicle = Vehicle(name="多平台车型")
            db.add(vehicle)
            db.flush()
            autohome_circle = Circle(
                platform_code="autohome",
                external_id="88001",
                name="汽车之家测试圈",
                url="https://example.test/forum/88001",
                vehicle_id=vehicle.id,
                source_kind="configured",
                auto_enabled=True,
                validation_status="verified",
            )
            db.add(autohome_circle)
            db.flush()

        plan = self.client.put(
            "/api/v1/extraction-plan",
            json={
                "revision": 1,
                "rules": [
                    {
                        "id": "rule-multi-0001",
                        "name": "多平台规则",
                        "platform_quantities": {"dongchedi": 10, "autohome": 20},
                        "circle_ids": [dongchedi_circle.id, autohome_circle.id],
                    }
                ],
                "nodes": [
                    {
                        "id": "node-multi-0001",
                        "weekdays": [0],
                        "time": "09:30:00",
                        "enabled": True,
                        "rule_ids": ["rule-multi-0001"],
                    }
                ],
            },
        )
        self.assertEqual(200, plan.status_code)
        run = self.container.runs.create_scheduled(
            datetime(2026, 8, 17, 1, 30, tzinfo=timezone.utc),
            "node-multi-0001",
            plan.json()["revision"],
        )
        self.assertIsNotNone(run)
        detail = self.container.runs.get_run(run["id"])
        self.assertEqual(
            {"autohome": 20, "dongchedi": 10},
            {task["platform_code"]: task["target_count"] for task in detail["tasks"]},
        )

    def test_schedule_node_merges_multiple_rules_and_uses_maximum_circle_target(self) -> None:
        circle = self.save_verified_circle()
        with self.container.sessions.begin() as db:
            stored = db.get(Circle, circle.id)
            assert stored is not None
            stored.auto_enabled = True

        plan = self.client.put(
            "/api/v1/extraction-plan",
            json={
                "revision": 1,
                "rules": [
                    {
                        "id": "rule-merge-a",
                        "name": "合并规则 A",
                        "platform_quantities": {"dongchedi": 10},
                        "circle_ids": [circle.id],
                    },
                    {
                        "id": "rule-merge-b",
                        "name": "合并规则 B",
                        "platform_quantities": {"dongchedi": 30},
                        "circle_ids": [circle.id],
                    },
                ],
                "nodes": [
                    {
                        "id": "node-merge-0001",
                        "weekdays": [0],
                        "time": "09:45:00",
                        "enabled": True,
                        "rule_ids": ["rule-merge-a", "rule-merge-b"],
                    }
                ],
            },
        )
        self.assertEqual(200, plan.status_code)
        self.assertEqual(["rule-merge-a", "rule-merge-b"], plan.json()["nodes"][0]["rule_ids"])

        run = self.container.runs.create_scheduled(
            datetime(2026, 8, 17, 1, 45, tzinfo=timezone.utc),
            "node-merge-0001",
            plan.json()["revision"],
        )
        assert run is not None
        repeated = self.container.runs.create_scheduled(
            datetime(2026, 8, 17, 1, 45, tzinfo=timezone.utc),
            "node-merge-0001",
            plan.json()["revision"],
        )
        assert repeated is not None
        self.assertEqual(run["id"], repeated["id"])
        detail = self.container.runs.get_run(run["id"])
        self.assertEqual(1, len(detail["tasks"]))
        self.assertEqual(30, detail["tasks"][0]["target_count"])
        self.assertEqual(30, detail["planned_count"])
        self.assertEqual(
            ["rule-merge-a", "rule-merge-b"],
            [item["id"] for item in detail["extraction_rules"]],
        )
        with self.container.sessions() as db:
            stored_run = db.get(ExtractionRun, run["id"])
            assert stored_run is not None
            self.assertIsNone(stored_run.extraction_rule_id)
            self.assertEqual(2, len(stored_run.config_snapshot["rules"]))
            self.assertEqual(
                ["rule-merge-a", "rule-merge-b"],
                [
                    item.rule_id
                    for item in db.scalars(
                        select(ExtractionRunRule)
                        .where(ExtractionRunRule.run_id == run["id"])
                        .order_by(ExtractionRunRule.position)
                    )
                ],
            )
            self.assertEqual(
                ["rule-merge-a", "rule-merge-b"],
                [
                    item.rule_id
                    for item in db.scalars(
                        select(ScheduleNodeRule)
                        .where(ScheduleNodeRule.schedule_node_id == "node-merge-0001")
                        .order_by(ScheduleNodeRule.position)
                    )
                ],
            )
            event = db.scalar(
                select(ScheduleEvent).where(ScheduleEvent.schedule_node_id == "node-merge-0001")
            )
            assert event is not None
            self.assertEqual(2, len(event.rule_snapshots))

    def test_manual_idempotency_and_dual_api_contract(self) -> None:
        payload = {
            "platform_code": "dongchedi",
            "circle_urls": ["https://www.dongchedi.com/community/24729"],
            "quantity": 3000,
            "idempotency_key": "page-key-0001",
        }
        first = self.client.post("/api/v1/runs/manual", json=payload)
        second = self.client.post("/api/v1/runs/manual", json=payload)
        self.assertEqual(202, first.status_code)
        self.assertFalse(first.json()["already_submitted"])
        self.assertEqual(2000, first.json()["planned_count"])
        self.assertTrue(second.json()["already_submitted"])
        self.assertEqual(first.json()["id"], second.json()["id"])

        changed = dict(payload, quantity=20)
        conflict = self.client.post("/api/v1/runs/manual", json=changed)
        self.assertEqual(409, conflict.status_code)
        self.assertIn("幂等键", conflict.json()["message"])

        internal_payload = {k: v for k, v in payload.items() if k != "idempotency_key"}
        internal = self.client.post(
            "/internal/v1/runs/manual",
            headers={"Idempotency-Key": "internal-key-0001"},
            json=internal_payload,
        )
        self.assertEqual(202, internal.status_code)

    def test_api_datetimes_keep_their_utc_offset_after_sqlite_round_trip(self) -> None:
        created = self.client.post(
            "/api/v1/runs/manual",
            json={
                "platform_code": "dongchedi",
                "circle_urls": ["https://www.dongchedi.com/community/24729"],
                "quantity": 1,
                "idempotency_key": "utc-timezone-0001",
            },
        )
        self.assertEqual(202, created.status_code)

        response = self.client.get(f"/api/v1/runs/{created.json()['id']}")
        self.assertEqual(200, response.status_code)
        for field in ("created_at", "queued_at"):
            self.assertRegex(response.json()[field], r"(?:Z|\+00:00)$")

    def test_run_list_exposes_and_filters_circle_list_orders(self) -> None:
        with self.container.sessions.begin() as db:
            reply_run = ExtractionRun(
                number="20260818-130000-001",
                trigger_type="manual",
                input_mode="circle_discovery",
                idempotency_key="run-list-reply",
                request_hash="r" * 64,
            )
            publish_run = ExtractionRun(
                number="20260818-130000-002",
                trigger_type="manual",
                input_mode="circle_discovery",
                idempotency_key="run-list-publish",
                request_hash="p" * 64,
            )
            url_run = ExtractionRun(
                number="20260818-130000-003",
                trigger_type="manual",
                input_mode="url_list",
                idempotency_key="run-list-url",
                request_hash="u" * 64,
            )
            db.add_all([reply_run, publish_run, url_run])
            db.flush()
            db.add_all(
                [
                    CircleTask(
                        run_id=reply_run.id,
                        platform_code="dongchedi",
                        external_id="24729",
                        circle_url="https://www.dongchedi.com/community/24729",
                        list_order="latest_reply",
                        queue_sequence=1,
                        target_count=1,
                    ),
                    CircleTask(
                        run_id=publish_run.id,
                        platform_code="dongchedi",
                        external_id="24729",
                        circle_url=("https://www.dongchedi.com/community/24729/dongtai-release"),
                        list_order="latest_publish",
                        queue_sequence=2,
                        target_count=1,
                    ),
                    CircleTask(
                        run_id=url_run.id,
                        platform_code="dongchedi",
                        external_id="known-url-list",
                        circle_url="",
                        list_order="latest_reply",
                        queue_sequence=3,
                        target_count=1,
                    ),
                ]
            )

        reply = self.client.get("/api/v1/runs?list_order=latest_reply").json()
        publish = self.client.get("/api/v1/runs?list_order=latest_publish").json()

        self.assertEqual(1, reply["total"])
        self.assertEqual([reply_run.id], [item["id"] for item in reply["items"]])
        self.assertEqual(["latest_reply"], reply["items"][0]["list_orders"])
        self.assertEqual(["最新回复"], reply["items"][0]["list_order_names"])
        self.assertEqual(1, publish["total"])
        self.assertEqual([publish_run.id], [item["id"] for item in publish["items"]])
        self.assertEqual(["latest_publish"], publish["items"][0]["list_orders"])
        self.assertEqual(["最新发布"], publish["items"][0]["list_order_names"])

    def test_session_is_encrypted_and_never_returned(self) -> None:
        state = {
            "cookies": [
                {
                    "name": "token",
                    "value": "secret-value",
                    "domain": ".dongchedi.com",
                    "path": "/",
                }
            ]
        }
        imported = self.client.post(
            "/internal/v1/platforms/dongchedi/session/import",
            json={"storage_state": state},
        )
        self.assertEqual(200, imported.status_code)
        status = self.client.get("/api/v1/platforms/dongchedi/session").json()
        self.assertEqual("available", status["status"])
        self.assertNotIn("secret-value", json.dumps(status, ensure_ascii=False))
        with self.container.sessions() as db:
            encrypted = db.get(PlatformSession, "dongchedi").encrypted_state
            self.assertNotIn(b"secret-value", encrypted)


class AuthComponentTests(AppCase):
    def test_zero_byte_auth_page_is_reported_as_failed(self) -> None:
        response = FakeAuthResponse(headers={"content-length": "0"})
        with self.assertRaises(AuthPageLoadError) as raised:
            asyncio.run(self.container.auth._require_interactive_page(FakeAuthPage(), response))
        self.assertEqual("AUTH_PAGE_EMPTY", raised.exception.code)
        self.assertEqual(200, raised.exception.http_status)

    def test_auth_page_requires_an_interactive_dom(self) -> None:
        response = FakeAuthResponse(headers={"content-type": "text/html"})
        asyncio.run(
            self.container.auth._require_interactive_page(
                FakeAuthPage(html="<html><body><button>登录</button></body></html>" * 8),
                response,
            )
        )

    def test_cdp_stream_sends_latest_frame_and_acknowledges_it(self) -> None:
        cdp = FakeCDPSession()
        task = AuthTask(
            id="auth-cdp-stream",
            platform_code="dongchedi",
            ticket="ticket",
            expires_at=datetime.now(timezone.utc) + timedelta(minutes=1),
            status="active",
            page_status="ready",
        )
        task.page = FakeAuthPage()  # type: ignore[assignment]
        task.context = FakeCDPContext(auth_state("session"), cdp)  # type: ignore[assignment]
        self.container.auth.tasks[task.id] = task
        socket = FakeStreamSocket()

        asyncio.run(self.container.auth.stream(task.id, task.ticket, socket))  # type: ignore[arg-type]

        self.assertTrue(socket.accepted)
        self.assertEqual("threadsnap-auth", socket.accepted_subprotocol)
        self.assertTrue(socket.closed)
        frame = next(message for message in socket.messages if message["type"] == "frame")
        self.assertEqual("jpeg-frame", frame["data"])
        start = next(params for method, params in cdp.calls if method == "Page.startScreencast")
        self.assertEqual(85, start["quality"])
        self.assertEqual(1280, start["maxWidth"])
        self.assertIn(("Page.screencastFrameAck", {"sessionId": 7}), cdp.calls)
        self.assertIn(("Page.stopScreencast", None), cdp.calls)
        self.assertTrue(cdp.detached)

    def test_cdp_pointer_input_preserves_drag_state_and_clamps_coordinates(self) -> None:
        cdp = FakeCDPSession()
        asyncio.run(
            self.container.auth._dispatch_pointer(
                cdp,  # type: ignore[arg-type]
                "mouseMoved",
                {
                    "x": 1400,
                    "y": -12,
                    "button": "left",
                    "buttons": 1,
                    "modifiers": 99,
                },
            )
        )

        self.assertEqual("Input.dispatchMouseEvent", cdp.calls[-1][0])
        payload = cdp.calls[-1][1]
        self.assertEqual(1280.0, payload["x"])
        self.assertEqual(0.0, payload["y"])
        self.assertEqual(1, payload["buttons"])
        self.assertEqual(15, payload["modifiers"])

    def test_profile_is_encrypted_at_rest_and_restored_per_task(self) -> None:
        profiles = self.container.auth.profiles
        source = profiles.prepare("dongchedi", "task-one")
        marker = source / "Default" / "profile-marker.txt"
        marker.parent.mkdir(parents=True)
        marker.write_text("sensitive-profile-state", encoding="utf-8")

        encrypted = profiles.promote("dongchedi", source, "task-one")
        self.assertFalse(source.exists())
        self.assertNotIn(b"sensitive-profile-state", encrypted.read_bytes())

        restored = profiles.prepare("dongchedi", "task-two")
        self.assertEqual(
            "sensitive-profile-state",
            (restored / "Default" / "profile-marker.txt").read_text(encoding="utf-8"),
        )
        fresh = profiles.prepare("dongchedi", "task-fresh", inherit_current=False)
        self.assertFalse((fresh / "Default" / "profile-marker.txt").exists())
        self.assertTrue(encrypted.is_file())
        profiles.discard(fresh)
        type(profiles)(profiles.root, self.container.session_store.fernet)
        self.assertFalse(restored.exists())
        self.assertTrue(encrypted.is_file())

    def test_fresh_auth_task_replaces_active_task(self) -> None:
        first = self.client.post("/api/v1/platforms/dongchedi/auth/tasks").json()
        reused = self.client.post("/api/v1/platforms/dongchedi/auth/tasks").json()
        fresh = self.client.post("/api/v1/platforms/dongchedi/auth/tasks?fresh=true").json()

        self.assertEqual(first["id"], reused["id"])
        self.assertNotEqual(first["id"], fresh["id"])
        self.assertFalse(first["fresh_profile"])
        self.assertTrue(fresh["fresh_profile"])
        self.assertEqual("cancelled", self.container.auth.tasks[first["id"]].status)
        self.assertEqual("created", self.container.auth.tasks[fresh["id"]].status)

    def test_closing_an_already_closed_auth_browser_is_idempotent(self) -> None:
        task = AuthTask(
            id="already-closed",
            platform_code="dongchedi",
            ticket="ticket",
            expires_at=datetime.now(timezone.utc),
            status="active",
        )
        task.context = AlreadyClosedAuthContext(auth_state("closed"))  # type: ignore[assignment]
        task.playwright = FakePlaywright()  # type: ignore[assignment]
        task.page = FakeAuthPage()  # type: ignore[assignment]

        asyncio.run(self.container.auth._close_browser(task))

        self.assertIsNone(task.context)
        self.assertIsNone(task.page)
        self.assertIsNone(task.playwright)

    def test_successful_validation_promotes_profile_and_resumes_platform(self) -> None:
        task = AuthTask(
            id="auth-success",
            platform_code="dongchedi",
            ticket="ticket",
            expires_at=datetime.now(timezone.utc),
            status="active",
            page_status="ready",
        )
        task.profile_dir = self.container.auth.profiles.prepare("dongchedi", task.id)
        (task.profile_dir / "Default").mkdir(parents=True)
        (task.profile_dir / "Default" / "marker.txt").write_text("new-profile", encoding="utf-8")
        task.context = FakeAuthContext(auth_state("new-session"))  # type: ignore[assignment]
        task.playwright = FakePlaywright()  # type: ignore[assignment]
        task.page = FakeAuthPage()  # type: ignore[assignment]
        socket = FakeAuthSocket()
        resumed: list[str] = []

        class ValidCollector:
            def __init__(self, *_args: object, **_kwargs: object):
                pass

            def validate_circle(self, _url: str) -> dict:
                return {"external_id": "24729"}

        with (
            patch("threadsnap.auth.DongchediCollector", ValidCollector),
            patch.object(
                self.container.worker,
                "resume_platform",
                side_effect=lambda platform: resumed.append(platform),
            ),
        ):
            asyncio.run(
                self.container.auth._command(task, {"type": "finish"}, socket)  # type: ignore[arg-type]
            )

        self.assertEqual("completed", task.status)
        self.assertEqual("completed", task.page_status)
        self.assertEqual(["dongchedi"], resumed)
        self.assertEqual(
            "new-session",
            self.container.session_store.get_state("dongchedi")["cookies"][0]["value"],
        )
        encrypted = self.container.auth.profiles.current("dongchedi")
        self.assertTrue(encrypted.is_file())
        self.assertNotIn(b"new-profile", encrypted.read_bytes())
        self.assertEqual("completed", socket.messages[-1]["type"])

    def test_failed_validation_keeps_previous_session_and_profile(self) -> None:
        self.container.session_store.import_state("dongchedi", auth_state("old-session"))
        profiles = self.container.auth.profiles
        old_profile = profiles.prepare("dongchedi", "old-profile")
        (old_profile / "Default").mkdir(parents=True)
        (old_profile / "Default" / "marker.txt").write_text("old", encoding="utf-8")
        current = profiles.promote("dongchedi", old_profile, "old-profile")
        previous_encrypted = current.read_bytes()

        task = AuthTask(
            id="auth-failed",
            platform_code="dongchedi",
            ticket="ticket",
            expires_at=datetime.now(timezone.utc),
            status="active",
            page_status="ready",
        )
        task.profile_dir = profiles.prepare("dongchedi", task.id)
        (task.profile_dir / "Default" / "marker.txt").write_text("new", encoding="utf-8")
        task.context = FakeAuthContext(auth_state("new-session"))  # type: ignore[assignment]
        task.playwright = FakePlaywright()  # type: ignore[assignment]
        task.page = FakeAuthPage()  # type: ignore[assignment]
        socket = FakeAuthSocket()

        class InvalidCollector:
            def __init__(self, *_args: object, **_kwargs: object):
                pass

            def validate_circle(self, _url: str) -> dict:
                raise AuthenticationRequired("仍需登录")

        with patch("threadsnap.auth.DongchediCollector", InvalidCollector):
            asyncio.run(
                self.container.auth._command(task, {"type": "finish"}, socket)  # type: ignore[arg-type]
            )

        self.assertEqual("active", task.status)
        self.assertEqual("ready", task.page_status)
        self.assertEqual("validation_failed", socket.messages[-1]["type"])
        self.assertEqual(previous_encrypted, current.read_bytes())
        self.assertEqual(
            "old-session",
            self.container.session_store.get_state("dongchedi")["cookies"][0]["value"],
        )
        asyncio.run(self.container.auth._close(task))


class QueueAndRetryTests(AppCase):
    def setUp(self) -> None:
        super().setUp()
        self.container.worker._collector = lambda _platform, _snapshot_concurrency=None: (
            FakeCollector()
        )  # type: ignore[method-assign]

    def test_platform_fifo_processes_one_run_then_next(self) -> None:
        circle = self.save_verified_circle()
        one = self.container.runs.create_manual(
            ManualRunCreate(platform_code="dongchedi", circle_ids=[circle.id], quantity=2),
            scope="api",
            header_key="fifo-key-0001",
        )
        two = self.container.runs.create_manual(
            ManualRunCreate(platform_code="dongchedi", circle_ids=[circle.id], quantity=1),
            scope="api",
            header_key="fifo-key-0002",
        )
        self.assertTrue(self.container.worker.process_once())
        self.assertEqual("success", self.container.runs.get_run(one["id"])["status"])
        self.assertEqual("queued", self.container.runs.get_run(two["id"])["status"])
        self.assertTrue(self.container.worker.process_once())
        self.assertEqual("success", self.container.runs.get_run(two["id"])["status"])

    def test_worker_persists_and_publishes_progress_before_task_finishes(self) -> None:
        """小批量任务应逐条形成权威进度，而不是结束时才一次跳满。"""

        observed_counts: list[int] = []
        circle = self.save_verified_circle()
        run = self.container.runs.create_manual(
            ManualRunCreate(platform_code="dongchedi", circle_ids=[circle.id], quantity=3),
            scope="api",
            header_key="live-progress-0001",
        )
        container = self.container

        class ProgressCollector(FakeCollector):
            def collect_circle(
                self,
                _url: str,
                target: int,
                skip_post_ids: set[str] | None = None,
                on_progress=None,
            ) -> dict:
                records = []
                for index in range(target):
                    record = sample_record(str(3001 + index))
                    records.append(record)
                    assert on_progress is not None
                    on_progress(record, None)
                    observed_counts.append(container.runs.get_run(run["id"])["completed_count"])
                return {
                    "records": records,
                    "failures": [],
                    "stop_reason": "已经取得配置数量的有效帖子。",
                }

        self.container.worker._collector = lambda *_args: ProgressCollector()  # type: ignore[method-assign]

        self.assertTrue(self.container.worker.process_once())

        self.assertEqual([1, 2, 3], observed_counts)
        self.assertEqual("success", self.container.runs.get_run(run["id"])["status"])
        events = [
            event
            for event in self.container.events.wait_after(0, timeout=0)
            if event["type"] == "run.changed" and event["resource_id"] == run["id"]
        ]
        self.assertEqual("running", events[0]["summary"]["status"])
        self.assertEqual([1, 2, 3], [event["summary"]["completed_count"] for event in events[1:4]])
        self.assertEqual("success", events[-1]["summary"]["status"])

    def test_auth_resume_publishes_queued_progress_event(self) -> None:
        """认证恢复提交为排队状态后应立即通知列表，不只依赖 Session 事件。"""

        circle = self.save_verified_circle()
        run = self.container.runs.create_manual(
            ManualRunCreate(platform_code="dongchedi", circle_ids=[circle.id], quantity=2),
            scope="api",
            header_key="resume-progress-0001",
        )
        with self.container.sessions.begin() as db:
            stored_run = db.get(ExtractionRun, run["id"])
            task = db.scalar(select(CircleTask).where(CircleTask.run_id == run["id"]))
            assert stored_run is not None and task is not None
            stored_run.status = "waiting_for_auth"
            stored_run.completed_count = 1
            task.status = "waiting_for_auth"
            task.completed_count = 1
            db.add(
                PostSnapshot(
                    run_id=stored_run.id,
                    circle_task_id=task.id,
                    platform_post_id="4001",
                    url="https://www.dongchedi.com/ugc/article/4001",
                    order_index=0,
                )
            )

        self.container.worker.resume_platform("dongchedi")

        resumed = self.container.runs.get_run(run["id"])
        self.assertEqual("queued", resumed["status"])
        event = self.container.events.wait_after(0, timeout=0)[-1]
        self.assertEqual("run.changed", event["type"])
        self.assertEqual(run["id"], event["resource_id"])
        self.assertEqual("queued", event["summary"]["status"])
        self.assertEqual(1, event["summary"]["completed_count"])

    def test_plain_dynamic_body_mapping_is_persisted_to_snapshot(self) -> None:
        """普通动态的 motor_title 正文必须经 Worker 原样进入不可变快照。"""

        class PlainDynamicCollector(DongchediCollector):
            def _json_api(self, _endpoint: str, **_params: object) -> tuple[dict, int]:
                return (
                    {
                        "status": 0,
                        "data": {
                            "group_id_str": "7674878578118377534",
                            "thread_title": "",
                            "motor_title": "第一句话。第二句话！\n第三段正文。",
                            "content": "",
                            "image_urls": [{"url": "https://example.test/a.jpg"}],
                            "motor_profile_info": {"name": "作者甲"},
                            "motor_car_info": {"source_desc": "风云A9车友圈"},
                            "operation_status": 0,
                            "comment_count": 0,
                        },
                    },
                    200,
                )

            def collect_circle(
                self,
                _url: str,
                _target: int,
                skip_post_ids: set[str] | None = None,
                on_progress=None,
            ) -> dict:
                record = self.fetch_post(
                    "https://www.dongchedi.com/ugc/article/7674878578118377534"
                )
                assert record is not None
                if on_progress:
                    on_progress(record, None)
                return {
                    "records": [record],
                    "failures": [],
                    "stop_reason": "已经取得配置数量的有效帖子。",
                }

        circle = self.save_verified_circle()
        self.container.worker._collector = lambda _platform, _concurrency=None: (  # type: ignore[method-assign]
            PlainDynamicCollector(None)
        )
        run = self.container.runs.create_manual(
            ManualRunCreate(platform_code="dongchedi", circle_ids=[circle.id], quantity=1),
            scope="api",
            header_key="plain-dynamic-body-0001",
        )

        self.assertTrue(self.container.worker.process_once())
        items = self.container.runs.posts(run["id"])["items"]
        self.assertEqual(1, len(items))
        detail = self.container.runs.post_detail(run["id"], items[0]["id"])
        self.assertEqual("第一句话。", detail["title"])
        self.assertEqual("第一句话。第二句话！\n第三段正文。", detail["content"])

    def test_retry_contains_only_failed_urls_and_results_are_merged(self) -> None:
        circle = self.save_verified_circle()
        original = self.container.runs.create_manual(
            ManualRunCreate(platform_code="dongchedi", circle_ids=[circle.id], quantity=2),
            scope="api",
            header_key="retry-original-0001",
        )
        with self.container.sessions.begin() as db:
            run = db.get(ExtractionRun, original["id"])
            task = db.scalar(select(CircleTask).where(CircleTask.run_id == original["id"]))
            task.status = "partial_success"
            task.completed_count = 1
            task.failed_count = 1
            task.checkpoint = {
                "failed_urls": [
                    {
                        "url": "https://www.dongchedi.com/ugc/article/2002",
                        "code": "POST_NOT_FOUND",
                    }
                ]
            }
            db.add(
                PostSnapshot(
                    run_id=run.id,
                    circle_task_id=task.id,
                    platform_post_id="2001",
                    url="https://www.dongchedi.com/ugc/article/2001",
                    title="原结果",
                    order_index=0,
                )
            )
            run.status = "partial_success"
            run.completed_count = 1
            run.failed_count = 1
        retry = self.container.runs.retry(original["id"], "retry-key-0001", "api")
        with self.container.sessions() as db:
            retry_task = db.scalar(select(CircleTask).where(CircleTask.run_id == retry["id"]))
            self.assertEqual(
                ["https://www.dongchedi.com/ugc/article/2002"],
                retry_task.config_snapshot["known_post_urls"],
            )
            self.assertEqual(1, retry_task.target_count)
        self.container.worker.process_once()
        merged = self.container.runs.posts(retry["id"])
        self.assertEqual(2, merged["total"])
        self.assertEqual(["2001", "2002"], [item["platform_post_id"] for item in merged["items"]])
        first_id, second_id = [item["id"] for item in merged["items"]]
        first_navigation = self.container.runs.post_navigation(retry["id"], first_id)
        self.assertIsNone(first_navigation["previous_id"])
        self.assertEqual(second_id, first_navigation["next_id"])
        self.assertEqual(1, first_navigation["position"])
        self.assertEqual(2, first_navigation["total"])
        reverse_navigation = self.container.runs.post_navigation(
            retry["id"], first_id, sort_by="source", sort_direction="desc"
        )
        self.assertEqual(second_id, reverse_navigation["previous_id"])
        self.assertIsNone(reverse_navigation["next_id"])

    def test_auth_wait_blocks_only_its_platform_queue(self) -> None:
        with self.container.sessions.begin() as db:
            yiche = db.get(PlatformConfig, "yiche")
            yiche.adapter_status = "available"
            yiche.enabled = True
            blocked_run = ExtractionRun(
                number="20260814-110000-001",
                trigger_type="manual",
                status="waiting_for_auth",
                idempotency_scope="test",
                idempotency_key="blocked",
                request_hash="b" * 64,
            )
            free_run = ExtractionRun(
                number="20260814-110000-002",
                trigger_type="manual",
                status="queued",
                idempotency_scope="test",
                idempotency_key="free",
                request_hash="f" * 64,
                planned_count=1,
            )
            db.add_all([blocked_run, free_run])
            db.flush()
            db.add_all(
                [
                    CircleTask(
                        run_id=blocked_run.id,
                        platform_code="dongchedi",
                        external_id="24729",
                        circle_url="https://www.dongchedi.com/community/24729",
                        status="waiting_for_auth",
                        queue_sequence=1,
                        target_count=1,
                        error_code="AUTH_REQUIRED",
                    ),
                    CircleTask(
                        run_id=free_run.id,
                        platform_code="yiche",
                        external_id="88",
                        circle_url="https://example.test/circle/88",
                        status="queued",
                        queue_sequence=2,
                        target_count=1,
                    ),
                ]
            )
        self.assertTrue(self.container.worker.process_once())
        self.assertEqual("waiting_for_auth", self.container.runs.get_run(blocked_run.id)["status"])
        self.assertEqual("success", self.container.runs.get_run(free_run.id)["status"])

    def test_weekly_scheduler_deduplicates_same_second(self) -> None:
        circle = self.save_verified_circle()
        other_circle = self.save_verified_circle(name="B9", external_id="24730")
        with self.container.sessions.begin() as db:
            stored = db.get(Circle, circle.id)
            stored.auto_enabled = True
            other = db.get(Circle, other_circle.id)
            other.auto_enabled = True
        updated = self.client.put(
            "/api/v1/extraction-plan",
            json={
                "revision": 1,
                "rules": [
                    {
                        "id": "rule-0001",
                        "name": "定时规则",
                        "platform_quantities": {"dongchedi": 1},
                        "circle_ids": [circle.id],
                    }
                ],
                "nodes": [
                    {
                        "id": "node-0001",
                        "weekdays": [4],
                        "time": "10:00:00",
                        "enabled": True,
                        "rule_ids": ["rule-0001"],
                    }
                ],
            },
        ).json()
        now = datetime(2026, 8, 14, 2, 0, tzinfo=timezone.utc)
        first = self.container.scheduler.tick(now)
        second = self.container.scheduler.tick(now)
        self.assertIsNotNone(first)
        self.assertIsNone(second)
        with self.container.sessions() as db:
            events = list(db.scalars(select(ScheduleEvent)))
            self.assertEqual(1, len(events))
            self.assertEqual(updated["revision"], events[0].schedule_revision)
            self.assertEqual("node-0001", events[0].schedule_node_id)
            tasks = list(db.scalars(select(CircleTask)))
            self.assertEqual([circle.id], [task.circle_id for task in tasks])

    def test_automatic_auth_refresh_retries_inside_same_run(self) -> None:
        circle = self.save_verified_circle()
        run = self.container.runs.create_manual(
            ManualRunCreate(platform_code="dongchedi", circle_ids=[circle.id], quantity=1),
            scope="api",
            header_key="auto-refresh-0001",
        )
        collectors = iter([FakeCollector(auth=True), FakeCollector()])
        self.container.worker._collector = lambda _platform, _concurrency=None: next(collectors)  # type: ignore[method-assign]
        self.container.worker._refresh_after_auth = lambda *_args: True  # type: ignore[method-assign]
        self.assertTrue(self.container.worker.process_once())
        self.assertEqual("success", self.container.runs.get_run(run["id"])["status"])
        self.assertEqual(1, self.container.runs.posts(run["id"])["total"])

    def test_restart_recovers_running_tasks_to_fifo(self) -> None:
        circle = self.save_verified_circle()
        run = self.container.runs.create_manual(
            ManualRunCreate(platform_code="dongchedi", circle_ids=[circle.id], quantity=1),
            scope="api",
            header_key="recovery-key-0001",
        )
        with self.container.sessions.begin() as db:
            stored_run = db.get(ExtractionRun, run["id"])
            task = db.scalar(select(CircleTask).where(CircleTask.run_id == run["id"]))
            stored_run.status = "running"
            stored_run.started_at = datetime.now(timezone.utc)
            task.status = "running"
            task.started_at = datetime.now(timezone.utc)
        self.container.worker.recover_interrupted()
        recovered = self.container.runs.get_run(run["id"])
        self.assertEqual("queued", recovered["status"])
        self.assertEqual("queued", recovered["tasks"][0]["status"])

    def test_end_auth_wait_closes_platform_auth_entry(self) -> None:
        auth_task = self.client.post("/api/v1/platforms/dongchedi/auth/tasks").json()
        with self.container.sessions.begin() as db:
            run = ExtractionRun(
                number="20260814-120000-001",
                trigger_type="manual",
                status="waiting_for_auth",
                idempotency_scope="test",
                idempotency_key="end-auth-wait",
                request_hash="e" * 64,
            )
            db.add(run)
            db.flush()
            db.add(
                CircleTask(
                    run_id=run.id,
                    platform_code="dongchedi",
                    external_id="24729",
                    circle_url="https://www.dongchedi.com/community/24729",
                    status="waiting_for_auth",
                    queue_sequence=1,
                    target_count=1,
                )
            )
        ended = self.client.post(f"/api/v1/runs/{run.id}/end-auth-wait")
        self.assertEqual(200, ended.status_code)
        self.assertEqual("failed", ended.json()["status"])
        status = self.client.get(f"/api/v1/auth/tasks/{auth_task['id']}").json()
        self.assertEqual("cancelled", status["status"])
        self.assertEqual("已结束", status["status_name"])

    def test_transient_manual_circle_is_saved_only_to_manual_history(self) -> None:
        run = self.container.runs.create_manual(
            ManualRunCreate(
                platform_code="dongchedi",
                circle_urls=["https://www.dongchedi.com/community/24729"],
                quantity=1,
            ),
            scope="api",
            header_key="manual-history-0001",
        )
        self.container.worker.process_once()
        self.assertEqual("success", self.container.runs.get_run(run["id"])["status"])
        history = self.container.config.list_manual_history()
        self.assertEqual(1, len(history))
        self.assertEqual("manual_history", history[0]["source_kind"])
        self.assertFalse(history[0]["auto_enabled"])
        self.assertEqual([], self.container.config.list_vehicles())


class CollectorTests(unittest.TestCase):
    def test_circle_url_parses_two_independent_list_orders(self) -> None:
        latest_reply = parse_circle_url("https://www.dongchedi.com/community/24729/2?x=1")
        latest_publish = parse_circle_url(
            "https://www.dongchedi.com/community/24729/dongtai-release/2"
        )

        self.assertEqual("latest_reply", latest_reply.list_order)
        self.assertEqual("https://www.dongchedi.com/community/24729", latest_reply.url)
        self.assertEqual("latest_publish", latest_publish.list_order)
        self.assertEqual(
            "https://www.dongchedi.com/community/24729/dongtai-release",
            latest_publish.url,
        )

    def test_video_id_resolves_highest_bitrate_url_over_two_http_requests(self) -> None:
        """视频 ID 通过授权与播放信息接口解析，不加载帖子网页或媒体正文。"""

        video_id = "video-id-1"
        signed_query = "Action=GetPlayInfo&Version=2019-03-15&signature=test"
        encoded_token = base64.b64encode(
            json.dumps({"GetPlayInfoToken": signed_query, "Version": "v1"}).encode("utf-8")
        ).decode("ascii")
        responses = iter(
            [
                SimpleNamespace(
                    status_code=200,
                    content=json.dumps(
                        {
                            "status": 0,
                            "data": {"play_auth_token": encoded_token},
                        }
                    ).encode("utf-8"),
                    url="https://www.dongchedi.com/motor/pc/common/token",
                    headers={"content-type": "application/json"},
                ),
                SimpleNamespace(
                    status_code=200,
                    content=json.dumps(
                        {
                            "Result": {
                                "Data": {
                                    "VideoID": video_id,
                                    "PlayInfoList": [
                                        {
                                            "Bitrate": 500,
                                            "MainPlayUrl": "https://media.test/low.mp4",
                                        },
                                        {
                                            "Bitrate": 1000,
                                            "MainPlayUrl": "https://media.test/high.mp4",
                                            "BackupPlayUrl": "https://backup.test/high.mp4",
                                        },
                                    ],
                                }
                            }
                        }
                    ).encode("utf-8"),
                    url="https://vod.bytedanceapi.com/",
                    headers={"content-type": "application/json"},
                ),
            ]
        )
        requested: list[str] = []
        collector = DongchediCollector(None)

        def fake_get(url: str) -> SimpleNamespace:
            requested.append(url)
            return next(responses)

        collector._get = fake_get  # type: ignore[method-assign]

        urls = collector.resolve_video_urls(video_id)

        self.assertEqual(["https://backup.test/high.mp4"], urls)
        self.assertEqual(2, len(requested))
        self.assertIn("/motor/pc/common/token?", requested[0])
        self.assertIn(f"video_id={video_id}", requested[0])
        self.assertTrue(requested[1].startswith("https://vod.bytedanceapi.com/?"))
        self.assertIn("signature=test", requested[1])

    def test_plain_dynamic_uses_motor_title_as_body_when_content_is_empty(self) -> None:
        collector = DongchediCollector(None)
        collector._json_api = lambda _endpoint, **_params: (  # type: ignore[method-assign]
            {
                "status": 0,
                "data": {
                    "group_id_str": "7674878578118377534",
                    "thread_title": "",
                    "motor_title": "第一句话。第二句话！\n第三段正文。",
                    "content": "",
                    "image_urls": [{"url": "https://example.test/a.jpg"}],
                    "operation_status": 0,
                    "comment_count": 0,
                },
            },
            200,
        )

        record = collector.fetch_post("https://www.dongchedi.com/ugc/article/7674878578118377534")

        self.assertIsNotNone(record)
        assert record is not None
        self.assertEqual("第一句话。", record["title"])
        self.assertEqual("第一句话。第二句话！\n第三段正文。", record["content"])

    def test_plain_dynamic_preserves_explicit_thread_title(self) -> None:
        collector = DongchediCollector(None)
        collector._json_api = lambda _endpoint, **_params: (  # type: ignore[method-assign]
            {
                "status": 0,
                "data": {
                    "group_id_str": "1002",
                    "thread_title": "平台明确标题",
                    "motor_title": "第一句话。第二句话。",
                    "content": "",
                    "operation_status": 0,
                    "comment_count": 0,
                },
            },
            200,
        )

        record = collector.fetch_post("https://www.dongchedi.com/ugc/article/1002")

        self.assertIsNotNone(record)
        assert record is not None
        self.assertEqual("平台明确标题", record["title"])
        self.assertEqual("第一句话。第二句话。", record["content"])

    def test_pure_media_post_keeps_title_and_content_empty(self) -> None:
        collector = DongchediCollector(None)
        collector._json_api = lambda _endpoint, **_params: (  # type: ignore[method-assign]
            {
                "status": 0,
                "data": {
                    "group_id_str": "1003",
                    "thread_title": "",
                    "motor_title": "",
                    "content": "",
                    "image_urls": [{"url": "https://example.test/a.jpg"}],
                    "operation_status": 0,
                    "comment_count": 0,
                },
            },
            200,
        )

        record = collector.fetch_post("https://www.dongchedi.com/ugc/article/1003")

        self.assertIsNotNone(record)
        assert record is not None
        self.assertIsNone(record["title"])
        self.assertIsNone(record["content"])
        self.assertEqual(["https://example.test/a.jpg"], record["image_urls"])

    def test_rich_text_post_uses_motor_title_and_plain_content(self) -> None:
        collector = DongchediCollector(None)
        collector._json_api = lambda _endpoint, **_params: (  # type: ignore[method-assign]
            {
                "status": 0,
                "data": {
                    "group_id_str": "7674619924202979865",
                    "thread_title": "",
                    "motor_title": "我和qq3的故事～",
                    "content": (
                        "<p>这个夏天最惊喜的双重浪漫✨</p>"
                        "<p>8月15日，既是我的生日，也是提车的日子。</p>"
                        '<div class="syl-image-wrapper"><img src="https://example.test/a.jpg"></div>'
                    ),
                    "image_urls": [{"url": "https://example.test/a.jpg"}],
                    "motor_profile_info": {"name": "作者甲"},
                    "motor_car_info": {"source_desc": "QQ3 EV车友圈"},
                    "operation_status": 0,
                    "comment_count": 0,
                },
            },
            200,
        )

        record = collector.fetch_post("https://www.dongchedi.com/ugc/article/7674619924202979865")

        self.assertIsNotNone(record)
        assert record is not None
        self.assertEqual("我和qq3的故事～", record["title"])
        self.assertEqual(
            "这个夏天最惊喜的双重浪漫✨\n8月15日，既是我的生日，也是提车的日子。",
            record["content"],
        )
        self.assertNotIn("<p>", record["content"])
        self.assertEqual(["https://example.test/a.jpg"], record["image_urls"])

    def test_rich_text_post_without_platform_title_uses_plain_first_sentence(self) -> None:
        collector = DongchediCollector(None)
        collector._json_api = lambda _endpoint, **_params: (  # type: ignore[method-assign]
            {
                "status": 0,
                "data": {
                    "group_id_str": "1001",
                    "content": "<p>第一句话。</p><p>第二段内容。</p>",
                    "operation_status": 0,
                    "comment_count": 0,
                },
            },
            200,
        )

        record = collector.fetch_post("https://www.dongchedi.com/ugc/article/1001")

        self.assertIsNotNone(record)
        assert record is not None
        self.assertEqual("第一句话。", record["title"])
        self.assertEqual("第一句话。\n第二段内容。", record["content"])

    def test_collection_continues_to_later_pages_until_valid_target(self) -> None:
        collector = DongchediCollector(None)
        requested_sources: list[str] = []
        pages = {
            1: {
                "total_count": 60,
                "page_count": 2,
                "rows": [{"post_id": str(x), "url": f"u{x}"} for x in range(1, 4)],
            },
            2: {
                "total_count": 60,
                "page_count": 2,
                "rows": [{"post_id": str(x), "url": f"u{x}"} for x in range(4, 7)],
            },
        }

        def fetch_circle_page(source_url: str, page: int, _expected: int | None) -> dict:
            requested_sources.append(source_url)
            return pages[page]

        collector._fetch_circle_page = fetch_circle_page  # type: ignore[method-assign]
        collector.fetch_post = lambda url: (
            None if url in {"u1", "u2", "u3"} else sample_record(url[1:])
        )  # type: ignore[method-assign]
        result = collector.collect_circle("https://www.dongchedi.com/community/24729", 3)
        self.assertEqual(["4", "5", "6"], [item["platform_post_id"] for item in result["records"]])
        self.assertEqual(3, len(result["failures"]))
        self.assertEqual(
            ["https://www.dongchedi.com/community/24729"] * 2,
            requested_sources,
        )

    def test_collection_resume_skips_completed_page_and_continues(self) -> None:
        collector = DongchediCollector(None)
        requested_pages: list[int] = []
        pages = {
            1: {
                "total_count": 60,
                "page_count": 2,
                "rows": [{"post_id": str(x), "url": f"u{x}"} for x in range(1, 4)],
            },
            2: {
                "total_count": 60,
                "page_count": 2,
                "rows": [{"post_id": str(x), "url": f"u{x}"} for x in range(4, 7)],
            },
        }

        def fetch_circle_page(_source_url: str, page: int, _expected: int | None) -> dict:
            requested_pages.append(page)
            return pages[page]

        collector._fetch_circle_page = fetch_circle_page  # type: ignore[method-assign]
        collector.fetch_post = lambda url: sample_record(url[1:])  # type: ignore[method-assign]

        result = collector.collect_circle(
            "https://www.dongchedi.com/community/24729",
            3,
            skip_post_ids={"1", "2", "3"},
        )

        self.assertEqual([1, 2], requested_pages)
        self.assertEqual(["4", "5", "6"], [item["platform_post_id"] for item in result["records"]])
        self.assertEqual("已经取得配置数量的有效帖子。", result["stop_reason"])

    def test_collection_keeps_latest_publish_source_url(self) -> None:
        collector = DongchediCollector(None)
        requested_sources: list[str] = []

        def fetch_circle_page(source_url: str, _page: int, _expected: int | None) -> dict:
            requested_sources.append(source_url)
            return {
                "total_count": 1,
                "page_count": 1,
                "rows": [{"post_id": "1001", "url": "u1001"}],
            }

        collector._fetch_circle_page = fetch_circle_page  # type: ignore[method-assign]
        collector.fetch_post = lambda _url: sample_record("1001")  # type: ignore[method-assign]

        result = collector.collect_circle(
            "https://www.dongchedi.com/community/24729/dongtai-release",
            1,
        )

        self.assertEqual(["1001"], [item["platform_post_id"] for item in result["records"]])
        self.assertEqual(
            ["https://www.dongchedi.com/community/24729/dongtai-release"],
            requested_sources,
        )

    def test_first_page_underfill_triggers_browser_fallback(self) -> None:
        collector = DongchediCollector(None)
        cards = "".join(
            f'<section class="community-card"><a href="/ugc/article/{index}">{index}</a></section>'
            for index in range(10)
        )
        body = f'<html><head><meta charset="utf-8"><title>测试车友圈</title></head><body><div>共 60 条内容</div>{cards}</body></html>'.encode()
        collector._get = lambda _url: SimpleNamespace(
            url="https://www.dongchedi.com/community/24729",
            content=body,
            status_code=200,
            headers={"content-type": "text/html; charset=utf-8"},
        )  # type: ignore[method-assign]
        browser_rows = [
            {"post_id": str(x), "url": f"u{x}", "sort_label": None, "order_index": x}
            for x in range(30)
        ]
        collector._browser_page_rows = lambda _url: browser_rows  # type: ignore[method-assign]
        page = collector._fetch_circle_page("https://www.dongchedi.com/community/24729", 1, None)
        self.assertEqual(30, len(page["rows"]))
        self.assertEqual(60, page["total_count"])

    def test_json_post_text_containing_login_word_is_not_auth_page(self) -> None:
        response = SimpleNamespace(
            url="https://www.dongchedi.com/motor/pc/ugc/detail/common",
            content='{"status":0,"data":{"content":"这里只是讨论登录功能"}}'.encode(),
            headers={"content-type": "application/json; charset=utf-8"},
        )
        DongchediCollector._detect_auth(response)


class TemplateTests(AppCase):
    def _seed_completed_run(self) -> tuple[ExtractionRun, Circle]:
        circle = self.save_verified_circle()
        with self.container.sessions.begin() as db:
            run = ExtractionRun(
                number="20260814-100000-001",
                trigger_type="manual",
                status="success",
                idempotency_scope="test",
                idempotency_key="template-run",
                request_hash="x" * 64,
                planned_count=2,
                completed_count=2,
            )
            db.add(run)
            db.flush()
            task = CircleTask(
                run_id=run.id,
                circle_id=circle.id,
                platform_code="dongchedi",
                external_id=circle.external_id,
                circle_name=circle.name,
                circle_url=circle.url,
                status="success",
                queue_sequence=1,
                target_count=2,
                completed_count=2,
                config_snapshot={"vehicle_name": "风云A9最新回复", "source_name": "风云A9最新回复"},
            )
            db.add(task)
            db.flush()
            for order, post_id in enumerate(("3001", "3002")):
                record = sample_record(post_id)
                post = PostSnapshot(
                    run_id=run.id,
                    circle_task_id=task.id,
                    platform_post_id=post_id,
                    url=record["url"],
                    title=record["title"],
                    author=record["author"],
                    published_at=record["published_at"],
                    content=f"{record['content']}\n第二段正文",
                    reply_count=1,
                    like_count=3,
                    visibility="visible",
                    order_index=order,
                )
                db.add(post)
                db.flush()
                db.add(
                    CommentSnapshot(
                        post_id=post.id,
                        author="评论者乙",
                        content="一级评论",
                        published_at=record["comments"][0]["published_at"],
                        like_count=5,
                        order_index=0,
                    )
                )
            return run, circle

    def test_template_validation_and_one_post_per_row_export(self) -> None:
        run, circle = self._seed_completed_run()
        source_prefix = f"s.{circle.export_key}"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "A9"
        sheet["A1"] = "帖子标题"
        sheet["B1"] = "评论"
        sheet["C1"] = "正文"
        sheet["A2"] = f"{source_prefix}.post.title"
        sheet["B2"] = f"{source_prefix}.comments.content_with_likes"
        sheet["C2"] = f"{source_prefix}.post.content"
        sheet.column_dimensions["A"].width = 4
        sheet.column_dimensions["B"].width = 4
        sheet.column_dimensions["C"].width = 4
        sheet.row_dimensions[2].height = 42
        sheet.row_dimensions[3].height = 42
        source = Path(self.temp.name) / "source.xlsx"
        workbook.save(source)
        version = self.container.templates.upload("客户模板", source.name, source.read_bytes())
        exported = self.container.templates.create_export(run.id, version["version_id"])
        target = self.container.templates.export_path(exported["id"])
        result = load_workbook(target)
        output = result["A9"]
        self.assertEqual("标题3001", output["A2"].value)
        self.assertEqual("标题3002", output["A3"].value)
        self.assertIn("1. 作者：评论者乙", output["B2"].value)
        self.assertIn("时间：2026-08-14 10:35:00", output["B2"].value)
        self.assertIn("点赞：5赞", output["B2"].value)
        self.assertEqual("正文3001\n第二段正文", output["C2"].value)
        self.assertTrue(output["A2"].alignment.wrap_text)
        self.assertTrue(output["B2"].alignment.wrap_text)
        self.assertTrue(output["C2"].alignment.wrap_text)
        self.assertIsNone(output.row_dimensions[2].height)
        self.assertIsNone(output.row_dimensions[3].height)
        self.assertGreater(output.column_dimensions["A"].width, 4)
        self.assertGreater(output.column_dimensions["B"].width, 4)
        self.assertGreater(output.column_dimensions["C"].width, 4)
        self.assertTrue(output.column_dimensions["A"].bestFit)

    def test_short_source_key_template_tag_targets_one_circle_feed_source(self) -> None:
        run, circle = self._seed_completed_run()
        tags = self.container.templates.field_tags(circle.id)
        title_tag = next(item["tag"] for item in tags if item["field"] == "post.title")
        source_prefix = f"s.{circle.export_key}"
        self.assertEqual(10, len(circle.export_key))
        self.assertRegex(circle.export_key, r"^[23456789abcdefghjkmnpqrstuvwxyz]{10}$")
        self.assertEqual(
            f"{source_prefix}.post.title",
            title_tag,
        )
        self.assertNotIn(circle.id, title_tag)
        self.assertLess(len(title_tag), len("source.AaANnBC7ft2mv5vpUUolPQ.post.title"))
        fields = {item["field"] for item in tags}
        self.assertTrue(
            {"source.id", "source.name", "source.list_order", "source.list_order_name"} <= fields
        )
        self.assertNotIn("vehicle.name", fields)

        workbook = Workbook()
        workbook.active["A1"] = title_tag
        workbook.active["B1"] = f"{source_prefix}.name"
        workbook.active["C1"] = f"{source_prefix}.list_order_name"
        source = Path(self.temp.name) / "source-id.xlsx"
        workbook.save(source)

        version = self.container.templates.upload("来源模板", source.name, source.read_bytes())
        exported = self.container.templates.create_export(run.id, version["version_id"])
        output = load_workbook(self.container.templates.export_path(exported["id"])).active
        self.assertEqual("标题3001", output["A1"].value)
        self.assertEqual("标题3002", output["A2"].value)
        self.assertEqual("风云A9", output["B1"].value)
        self.assertEqual("最新回复", output["C1"].value)

    def test_template_field_rules_are_identical_across_platforms(self) -> None:
        dongchedi = self.save_verified_circle()
        with self.container.sessions.begin() as db:
            autohome = Circle(
                platform_code="autohome",
                external_id="series-1001",
                name="汽车之家测试来源",
                url="https://example.test/autohome/series-1001",
                source_kind="configured",
                validation_status="verified",
            )
            db.add(autohome)
            db.flush()

        dongchedi_tags = self.container.templates.field_tags(dongchedi.id)
        autohome_tags = self.container.templates.field_tags(autohome.id)

        self.assertNotEqual(dongchedi.export_key, autohome.export_key)
        self.assertEqual(
            [item["field"] for item in dongchedi_tags],
            [item["field"] for item in autohome_tags],
        )
        self.assertTrue(
            all(item["tag"].startswith(f"s.{dongchedi.export_key}.") for item in dongchedi_tags)
        )
        self.assertTrue(
            all(item["tag"].startswith(f"s.{autohome.export_key}.") for item in autohome_tags)
        )
        self.assertTrue(all("dongchedi" not in item["tag"] for item in dongchedi_tags))
        self.assertTrue(all("autohome" not in item["tag"] for item in autohome_tags))

    def test_template_source_file_can_be_downloaded(self) -> None:
        circle = self.save_verified_circle()
        workbook = Workbook()
        workbook.active["A1"] = f"s.{circle.export_key}.post.title"
        source = Path(self.temp.name) / "download-source.xlsx"
        workbook.save(source)
        version = self.container.templates.upload("客户/模板", source.name, source.read_bytes())

        response = self.client.get(
            f"/api/v1/templates/{version['template_id']}/versions/{version['version_id']}/download"
        )

        self.assertEqual(200, response.status_code)
        self.assertEqual(source.read_bytes(), response.content)
        self.assertEqual(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers["content-type"],
        )
        self.assertIn("v1.xlsx", response.headers["content-disposition"])

        missing = self.client.get(
            f"/api/v1/templates/{version['template_id']}/versions/missing/download"
        )
        self.assertEqual(404, missing.status_code)
        self.assertEqual("TEMPLATE_VERSION_NOT_FOUND", missing.json()["code"])

    def test_invalid_template_returns_cell_and_field(self) -> None:
        _, circle = self._seed_completed_run()
        workbook = Workbook()
        workbook.active["A1"] = f"s.{circle.export_key}.post.unknown"
        source = Path(self.temp.name) / "invalid.xlsx"
        workbook.save(source)
        with self.assertRaises(DomainError) as raised:
            self.container.templates.upload("无效模板", source.name, source.read_bytes())
        self.assertEqual("A1", raised.exception.details[0]["cell"])
        self.assertEqual("post.unknown", raised.exception.details[0]["field"])
        self.assertEqual("字段未注册", raised.exception.details[0]["reason"])

    def test_legacy_long_template_tag_is_not_accepted(self) -> None:
        _, circle = self._seed_completed_run()
        workbook = Workbook()
        workbook.active["A1"] = "source.AaANnBC7ft2mv5vpUUolPQ.post.title"
        workbook.active["B1"] = f"platform.dongchedi.source.{circle.id}.post.title"
        source = Path(self.temp.name) / "legacy.xlsx"
        workbook.save(source)

        with self.assertRaises(DomainError) as raised:
            self.container.templates.upload("旧标签模板", source.name, source.read_bytes())

        self.assertEqual("A1", raised.exception.details[0]["cell"])
        self.assertEqual("标签格式无效", raised.exception.details[0]["reason"])
        self.assertEqual("B1", raised.exception.details[1]["cell"])
        self.assertEqual("标签格式无效", raised.exception.details[1]["reason"])


if __name__ == "__main__":
    unittest.main()
