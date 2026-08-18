"""稳定标签 XLSX 模板校验、版本化和导出。"""

from __future__ import annotations

import hashlib
import re
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.cell import Cell
from sqlalchemy import func, select
from sqlalchemy.orm import Session, sessionmaker

from .config import Settings
from .errors import DomainError
from .ids import uuid7
from .models import (
    Circle,
    CircleTask,
    CommentSnapshot,
    ExportRecord,
    ExtractionRun,
    PostSnapshot,
    Template,
    TemplateVersion,
    utc_now,
)
from .services import related_run_ids

TAG_RE = re.compile(
    r"^s\.(?P<source_key>[23456789abcdefghjkmnpqrstuvwxyz]{10})\."
    r"(?P<field>[a-z0-9_.]+)$"
)
FIELD_REGISTRY: dict[str, dict[str, str]] = {
    "source.id": {"type": "text", "description": "稳定的圈子来源 ID"},
    "source.name": {"type": "text", "description": "用户填写的来源名称"},
    "source.list_order": {
        "type": "text",
        "description": "来源列表类型：latest_reply 或 latest_publish",
    },
    "source.list_order_name": {"type": "text", "description": "来源列表类型中文名称"},
    "circle.id": {"type": "text", "description": "平台圈子 ID"},
    "circle.name": {"type": "text", "description": "批次快照中的圈子名称"},
    "circle.url": {"type": "text", "description": "批次快照中的规范化圈子链接"},
    "post.platform_post_id": {"type": "text", "description": "平台帖子 ID"},
    "post.url": {"type": "text", "description": "帖子规范化链接"},
    "post.title": {"type": "text", "description": "帖子标题"},
    "post.author": {"type": "text", "description": "帖子作者"},
    "post.published_at": {"type": "datetime", "description": "帖子发布时间"},
    "post.content": {"type": "text", "description": "帖子正文"},
    "post.image_urls": {"type": "collection", "description": "编号后的图片 URL"},
    "post.video_urls": {"type": "collection", "description": "编号后的视频 URL"},
    "post.reply_count": {"type": "number", "description": "提取时回复数"},
    "post.like_count": {"type": "number", "description": "提取时点赞数"},
    "post.section": {"type": "text", "description": "来源版块"},
    "post.visibility": {"type": "text", "description": "标准化可见状态"},
    "post.raw_status": {"type": "text", "description": "平台原始状态 JSON"},
    "comments.content": {"type": "collection", "description": "主评论作者、时间和正文"},
    "comments.content_with_likes": {
        "type": "collection",
        "description": "主评论作者、时间、正文和可用点赞数",
    },
}
PUBLIC_FIELD_PRIORITY = ("source.name", "source.list_order_name")
PUBLIC_FIELDS = PUBLIC_FIELD_PRIORITY + tuple(
    field for field in FIELD_REGISTRY if field not in PUBLIC_FIELD_PRIORITY
)
SHORT_FIELD_NAMES = {
    "source.id": "id",
    "source.name": "name",
    "source.list_order": "list_order",
    "source.list_order_name": "list_order_name",
}
SHORT_FIELD_ALIASES = {short: field for field, short in SHORT_FIELD_NAMES.items()}


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


class TemplateService:
    def __init__(self, factory: sessionmaker[Session], settings: Settings):
        self.factory = factory
        self.settings = settings
        self.zone = ZoneInfo(settings.timezone)

    def field_tags(self, source_id: str | None = None) -> list[dict[str, str]]:
        if not source_id:
            return [{"field": field, **FIELD_REGISTRY[field]} for field in PUBLIC_FIELDS]
        with self.factory() as db:
            source = db.get(Circle, source_id)
        if not source:
            raise DomainError("SOURCE_NOT_FOUND", "指定来源不存在。", status_code=404)
        prefix = f"s.{source.export_key}"
        return [
            {
                "tag": f"{prefix}.{SHORT_FIELD_NAMES.get(field, field)}",
                "field": field,
                **FIELD_REGISTRY[field],
            }
            for field in PUBLIC_FIELDS
        ]

    def validate(self, data: bytes) -> list[dict[str, Any]]:
        try:
            from io import BytesIO

            workbook = load_workbook(BytesIO(data), data_only=False)
        except Exception as exc:
            raise DomainError("TEMPLATE_INVALID", f"XLSX 模板读取失败：{exc}") from exc
        errors: list[dict[str, Any]] = []
        bindings: list[dict[str, Any]] = []
        with self.factory() as db:
            known_sources = {source.export_key: source.id for source in db.scalars(select(Circle))}
        for sheet in workbook.worksheets:
            merged = list(sheet.merged_cells.ranges)
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str) or not value.startswith(
                        ("s.", "source.", "platform.")
                    ):
                        continue
                    tag = value.strip()
                    match = TAG_RE.match(tag)
                    if not match:
                        errors.append(
                            {
                                "sheet": sheet.title,
                                "cell": cell.coordinate,
                                "field": value,
                                "reason": "标签格式无效",
                            }
                        )
                        continue
                    raw_field = match.group("field")
                    field = SHORT_FIELD_ALIASES.get(raw_field, raw_field)
                    if field not in FIELD_REGISTRY:
                        errors.append(
                            {
                                "sheet": sheet.title,
                                "cell": cell.coordinate,
                                "field": field,
                                "reason": "字段未注册",
                            }
                        )
                    source_id = known_sources.get(match.group("source_key"))
                    if source_id is None:
                        errors.append(
                            {
                                "sheet": sheet.title,
                                "cell": cell.coordinate,
                                "field": value,
                                "reason": "标签引用的来源尚未保存",
                            }
                        )
                    merged_range = next(
                        (str(area) for area in merged if cell.coordinate in area), None
                    )
                    if merged_range:
                        errors.append(
                            {
                                "sheet": sheet.title,
                                "cell": cell.coordinate,
                                "field": value,
                                "reason": f"标签不能位于合并单元格 {merged_range}",
                            }
                        )
                    bindings.append(
                        {
                            "sheet": sheet.title,
                            "cell": cell.coordinate,
                            "source_id": source_id,
                            "field": field,
                            "tag": tag,
                        }
                    )
        if not bindings and not errors:
            errors.append(
                {
                    "sheet": None,
                    "cell": None,
                    "field": None,
                    "reason": "模板中没有找到任何稳定字段标签",
                }
            )
        if errors:
            raise DomainError("TEMPLATE_INVALID", "XLSX 模板校验失败。", details=errors)
        return bindings

    def upload(self, name: str, filename: str, data: bytes) -> dict[str, Any]:
        if not filename.lower().endswith(".xlsx"):
            raise DomainError("TEMPLATE_FILE_TYPE_INVALID", "模板文件必须使用 .xlsx 格式。")
        clean_name = name.strip()
        if not clean_name:
            raise DomainError("TEMPLATE_NAME_REQUIRED", "模板名称不能为空。")
        bindings = self.validate(data)
        digest = sha256_bytes(data)
        with self.factory.begin() as db:
            template = db.scalar(select(Template).where(Template.name == clean_name))
            if not template:
                template = Template(name=clean_name)
                db.add(template)
                db.flush()
            version = (
                db.scalar(
                    select(func.max(TemplateVersion.version)).where(
                        TemplateVersion.template_id == template.id
                    )
                )
                or 0
            ) + 1
            version_id = uuid7()
            path = self.settings.template_dir / template.id / f"{version_id}.xlsx"
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(data)
            item = TemplateVersion(
                id=version_id,
                template_id=template.id,
                version=version,
                file_path=str(path.resolve()),
                file_sha256=digest,
                bindings=bindings,
            )
            template.hidden = False
            db.add(item)
            db.flush()
            return self.version_dict(template, item)

    def list_templates(self) -> list[dict[str, Any]]:
        with self.factory() as db:
            templates = list(
                db.scalars(
                    select(Template).where(Template.hidden.is_(False)).order_by(Template.name)
                )
            )
            output = []
            for template in templates:
                versions = list(
                    db.scalars(
                        select(TemplateVersion)
                        .where(TemplateVersion.template_id == template.id)
                        .order_by(TemplateVersion.version.desc())
                    )
                )
                output.append(
                    {
                        "id": template.id,
                        "name": template.name,
                        "versions": [self.version_dict(template, item) for item in versions],
                    }
                )
            return output

    @staticmethod
    def version_dict(template: Template, item: TemplateVersion) -> dict[str, Any]:
        return {
            "template_id": template.id,
            "template_name": template.name,
            "version_id": item.id,
            "version": item.version,
            "file_sha256": item.file_sha256,
            "bindings": item.bindings,
            "created_at": item.created_at,
        }

    def hide_template(self, template_id: str) -> None:
        with self.factory.begin() as db:
            template = db.get(Template, template_id)
            if not template:
                raise DomainError("TEMPLATE_NOT_FOUND", "指定模板不存在。", status_code=404)
            template.hidden = True

    def template_path(self, template_id: str, version_id: str) -> tuple[Path, str]:
        """返回指定模板版本的原始文件路径和安全下载文件名。"""

        with self.factory() as db:
            template = db.get(Template, template_id)
            version = db.get(TemplateVersion, version_id)
            if not template or not version or version.template_id != template_id:
                raise DomainError(
                    "TEMPLATE_VERSION_NOT_FOUND",
                    "指定模板版本不存在。",
                    status_code=404,
                )
            path = Path(version.file_path)
            clean_name = re.sub(r'[\\/:*?"<>|\x00-\x1f]', "_", template.name).strip(". ")
            filename = f"{clean_name or 'template'}-v{version.version}.xlsx"
        if not path.is_file():
            raise DomainError(
                "TEMPLATE_FILE_MISSING",
                "模板版本记录存在，但原始文件已经丢失。",
                status_code=500,
            )
        return path, filename

    def create_export(self, run_id: str, template_version_id: str) -> dict[str, Any]:
        with self.factory() as db:
            run = db.get(ExtractionRun, run_id)
            version = db.get(TemplateVersion, template_version_id)
            if not run:
                raise DomainError("RUN_NOT_FOUND", "指定提取批次不存在。", status_code=404)
            if run.status not in {"success", "partial_success", "failed"} or (
                run.status == "failed" and run.completed_count == 0
            ):
                raise DomainError(
                    "EXPORT_NOT_ALLOWED",
                    "当前批次状态没有可导出的持久化结果。",
                    status_code=409,
                )
            if not version:
                raise DomainError(
                    "TEMPLATE_VERSION_NOT_FOUND",
                    "指定模板版本不存在。",
                    status_code=404,
                )
            existing = db.scalar(
                select(ExportRecord).where(
                    ExportRecord.run_id == run_id,
                    ExportRecord.summary_version == run.summary_version,
                    ExportRecord.template_version_id == template_version_id,
                    ExportRecord.status == "success",
                )
            )
            if existing and existing.file_path and Path(existing.file_path).is_file():
                return self.export_dict(existing)
            summary_version = run.summary_version
        export_id = uuid7()
        output_path = self.settings.export_dir / run_id / f"{export_id}.xlsx"
        output_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            self._render(run_id, version, output_path)
            digest = hashlib.sha256(output_path.read_bytes()).hexdigest()
        except Exception:
            output_path.unlink(missing_ok=True)
            raise
        with self.factory.begin() as db:
            record = ExportRecord(
                id=export_id,
                run_id=run_id,
                summary_version=summary_version,
                template_version_id=template_version_id,
                status="success",
                file_path=str(output_path.resolve()),
                file_sha256=digest,
                finished_at=utc_now(),
            )
            db.add(record)
            db.flush()
            return self.export_dict(record)

    def _related_run_ids(self, db: Session, root_id: str) -> list[str]:
        return related_run_ids(db, root_id)

    def _render(self, run_id: str, version: TemplateVersion, output_path: Path) -> None:
        workbook = load_workbook(version.file_path, data_only=False)
        with self.factory() as db:
            run_ids = self._related_run_ids(db, run_id)
            tasks = list(
                db.scalars(
                    select(CircleTask)
                    .where(CircleTask.run_id.in_(run_ids))
                    .order_by(CircleTask.created_at, CircleTask.queue_sequence)
                )
            )
            selected_sources: dict[str, list[CircleTask]] = {}
            for task in tasks:
                if task.completed_count and task.circle_id:
                    selected_sources.setdefault(task.circle_id, []).append(task)

            def collect(
                task_group: list[CircleTask],
            ) -> list[tuple[PostSnapshot, list[CommentSnapshot], CircleTask]]:
                values = []
                seen_post_ids: set[str] = set()
                for task in task_group:
                    posts = list(
                        db.scalars(
                            select(PostSnapshot)
                            .where(PostSnapshot.circle_task_id == task.id)
                            .order_by(PostSnapshot.order_index)
                        )
                    )
                    for post in posts:
                        if post.platform_post_id in seen_post_ids:
                            continue
                        seen_post_ids.add(post.platform_post_id)
                        comments = list(
                            db.scalars(
                                select(CommentSnapshot)
                                .where(CommentSnapshot.post_id == post.id)
                                .order_by(CommentSnapshot.order_index)
                            )
                        )
                        values.append((post, comments, task))
                return values

            source_post_map = {key: collect(group) for key, group in selected_sources.items()}
        errors: list[dict[str, Any]] = []
        for binding in version.bindings:
            sheet = workbook[binding["sheet"]]
            origin = sheet[binding["cell"]]
            posts = source_post_map.get(binding["source_id"], [])
            for offset, item in enumerate(posts):
                target = sheet.cell(row=origin.row + offset, column=origin.column)
                if offset and target.value not in (None, ""):
                    errors.append(
                        {
                            "sheet": sheet.title,
                            "cell": target.coordinate,
                            "field": binding["tag"],
                            "reason": "预计写入范围已有文字、公式或其他内容",
                        }
                    )
                    continue
                if offset:
                    self._copy_style(origin, target)
                target.value = self._field_value(binding["field"], *item)
                if (
                    FIELD_REGISTRY[binding["field"]]["type"] == "datetime"
                    and target.value is not None
                ):
                    target.number_format = "yyyy-mm-dd hh:mm:ss"
                if FIELD_REGISTRY[binding["field"]]["type"] == "collection":
                    alignment = copy(target.alignment)
                    alignment.wrap_text = True
                    target.alignment = alignment
            if not posts:
                origin.value = None
        if errors:
            raise DomainError("EXPORT_RANGE_CONFLICT", "模板预计写入范围存在冲突。", details=errors)
        workbook.save(output_path)

    @staticmethod
    def _copy_style(origin: Cell, target: Cell) -> None:
        target.font = copy(origin.font)
        target.fill = copy(origin.fill)
        target.border = copy(origin.border)
        target.alignment = copy(origin.alignment)
        target.number_format = origin.number_format
        target.protection = copy(origin.protection)

    def _field_value(
        self,
        field: str,
        post: PostSnapshot,
        comments: list[CommentSnapshot],
        task: CircleTask,
    ) -> Any:
        if field == "source.id":
            return task.circle_id
        if field == "source.name":
            return task.config_snapshot.get("source_name")
        if field == "source.list_order":
            return task.list_order
        if field == "source.list_order_name":
            return "最新发布" if task.list_order == "latest_publish" else "最新回复"
        if field == "circle.id":
            return task.external_id
        if field == "circle.name":
            return task.circle_name
        if field == "circle.url":
            return task.circle_url
        values = {
            "post.platform_post_id": post.platform_post_id,
            "post.url": post.url,
            "post.title": post.title,
            "post.author": post.author,
            "post.published_at": self._local_naive(post.published_at),
            "post.content": post.content,
            "post.image_urls": self._numbered(post.image_urls),
            "post.video_urls": self._numbered(post.video_urls),
            "post.reply_count": post.reply_count,
            "post.like_count": post.like_count,
            "post.section": post.section,
            "post.visibility": post.visibility,
            "post.raw_status": json_text(post.raw_status),
            "comments.content": self._comments(comments, False),
            "comments.content_with_likes": self._comments(comments, True),
        }
        return values[field]

    def _local_naive(self, value: datetime | None) -> datetime | None:
        if value is None:
            return None
        if value.tzinfo is None:
            value = value.replace(tzinfo=ZoneInfo("UTC"))
        return value.astimezone(self.zone).replace(tzinfo=None)

    @staticmethod
    def _numbered(values: list[str]) -> str:
        return "\n".join(
            f"{index}. {value}{'；' if index < len(values) else ''}"
            for index, value in enumerate(values, 1)
        )

    def _comments(self, comments: list[CommentSnapshot], with_likes: bool) -> str:
        blocks = []
        for index, item in enumerate(comments, 1):
            local = self._local_naive(item.published_at)
            lines = [
                f"{index}. 作者：{item.author or ''}",
                f"   时间：{local:%Y-%m-%d %H:%M:%S}" if local else "   时间：",
                f"   评论：{item.content or ''}{'；' if index < len(comments) else ''}",
            ]
            if with_likes and item.like_count is not None:
                lines.append(f"   点赞：{item.like_count}赞")
            blocks.append("\n".join(lines))
        return "\n\n".join(blocks)

    @staticmethod
    def export_dict(item: ExportRecord) -> dict[str, Any]:
        return {
            "id": item.id,
            "run_id": item.run_id,
            "summary_version": item.summary_version,
            "template_version_id": item.template_version_id,
            "status": item.status,
            "file_sha256": item.file_sha256,
            "created_at": item.created_at,
            "finished_at": item.finished_at,
        }

    def export_path(self, export_id: str) -> Path:
        with self.factory() as db:
            item = db.get(ExportRecord, export_id)
            if not item or item.status != "success" or not item.file_path:
                raise DomainError("EXPORT_NOT_FOUND", "指定导出文件不存在。", status_code=404)
            path = Path(item.file_path)
        if not path.is_file():
            raise DomainError(
                "EXPORT_FILE_MISSING", "导出记录存在，但文件已经丢失。", status_code=500
            )
        return path


def json_text(value: Any) -> str | None:
    if value is None:
        return None
    import json

    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
