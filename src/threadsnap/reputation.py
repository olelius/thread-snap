"""垂媒口碑巡检领域服务与隔离合成验收运行。"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import shutil
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from time import monotonic
from typing import Any, Literal
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorksheetImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
from pydantic import BaseModel
from sqlalchemy import delete, func, or_, select
from sqlalchemy.orm import Session, sessionmaker

from .config import Settings
from .errors import DomainError
from .ids import uuid7
from .models import (
    Circle,
    CircleTask,
    PlatformConfig,
    ReputationDeleteJob,
    ReputationEvidence,
    ReputationMappingValidationAttempt,
    ReputationMappingValidationRun,
    ReputationResult,
    ReputationRun,
    ReputationScheduleEvent,
    ReputationSchedulerState,
    ReputationScopeDraft,
    ReputationScopeVersion,
    ReputationTombstone,
    ValidationJob,
)
from .reputation_adapter import (
    ReputationAdapterError,
    ReputationMappingTarget,
    ReputationPageResult,
)
from .reputation_registry import REPUTATION_PLATFORMS, ReputationPlatformSpec
from .session_store import SessionStore

FIXTURE_VERSION = "reputation-synthetic-v2-all-evidence"
PLATFORM_CODE = "dongchedi"
PLATFORM_NAME = "懂车帝"
DEFAULT_PROJECT_GROUP = "奇瑞项目组"
INSPECTION_TIME = time(10, 0)
INSPECTION_TIME_TEXT = INSPECTION_TIME.isoformat()
# 正式口碑巡检及其失败项补跑使用独立、固定的页面并发，避免帖子提取配置
# 调整后在计划时点突然打开更多浏览器窗口。
REPUTATION_RUN_CONCURRENCY = 2
SCENARIOS: dict[str, dict[str, str]] = {
    "baseline_initialization": {
        "name": "基线初始化",
        "description": "27款车型首次建档，全量证据，不计算涨跌。",
    },
    "daily_mixed_changes": {
        "name": "日常混合变化",
        "description": "覆盖上涨、下降、分化、无变化、异常和27项全量证据。",
    },
    "month_end_mixed_changes": {
        "name": "月末混合变化",
        "description": "沿用前日变化、月末标识和27项全量页面证据。",
    },
}

GREEN_FILL = PatternFill("solid", fgColor="E2F0D9")
RED_FILL = PatternFill("solid", fgColor="F4CCCC")
NEUTRAL_FILL = PatternFill("solid", fgColor="F8FAFC")
HEADER_FILL = PatternFill("solid", fgColor="E8EEF8")


class SyntheticRunCreate(BaseModel):
    scenario_id: Literal[
        "baseline_initialization", "daily_mixed_changes", "month_end_mixed_changes"
    ]


class MappingPasteRow(BaseModel):
    vehicle_id: str
    platform_vehicle_id: str
    platform_url: str
    platform_display_name: str


class MappingPasteRequest(BaseModel):
    revision: int
    platform_code: str
    rows: list[MappingPasteRow]


class ScopePublishRequest(BaseModel):
    revision: int
    initial_review_acknowledged: bool = False


class MappingValidationRequest(BaseModel):
    revision: int
    platform_code: str = PLATFORM_CODE
    vehicle_ids: list[str] | None = None


class ScopeVehicleCreateRequest(BaseModel):
    revision: int
    series_name: str
    vehicle_name: str
    project_group: str
    role: Literal["focus", "competitor"]
    platform_code: str = PLATFORM_CODE
    platform_vehicle_id: str
    platform_url: str
    platform_display_name: str


class ScopeVehicleUpdateRequest(ScopeVehicleCreateRequest):
    """修改车型草稿时复用新增车型的完整字段合同。"""


class ScopeVehicleRevisionRequest(BaseModel):
    revision: int


@dataclass(frozen=True)
class SyntheticVehicle:
    vehicle_id: str
    series_name: str
    vehicle_name: str
    role: str
    role_position: int
    vehicle_position: int


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _text_hash(value: Any) -> str:
    payload = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _mapping_hash(
    vehicle_id: str, mapping: dict[str, Any], platform_code: str = PLATFORM_CODE
) -> str:
    """绑定内部车型与全部身份关键映射字段。"""

    return _text_hash(
        {
            "vehicle_id": vehicle_id,
            "platform_code": platform_code,
            "platform_vehicle_id": mapping.get("platform_vehicle_id"),
            "platform_url": mapping.get("platform_url"),
            "platform_display_name": mapping.get("platform_display_name"),
        }
    )


def _metric(
    current: Decimal | None,
    baseline: Decimal | None,
    *,
    inverse: bool = False,
    state: str = "available",
    raw: str | None = None,
    baseline_raw: str | None = None,
) -> dict[str, Any]:
    """按页面可见十进制值生成一个指标的比较投影。"""

    if state != "available":
        return {
            "raw": raw,
            "value": None,
            "baseline_raw": baseline_raw,
            "baseline_value": None,
            "delta": None,
            "direction": "none",
            "tone": "neutral",
            "comparison_status": state,
        }
    if current is None:
        raise ValueError("available 指标必须提供当前值")
    current_text = raw or format(current, "f")
    if baseline is None:
        return {
            "raw": current_text,
            "value": format(current, "f"),
            "baseline_raw": None,
            "baseline_value": None,
            "delta": None,
            "direction": "none",
            "tone": "neutral",
            "comparison_status": "no_baseline",
        }
    delta = current - baseline
    if delta == 0:
        direction = "same"
        tone = "neutral"
    else:
        direction = "up" if delta > 0 else "down"
        positive = delta < 0 if inverse else delta > 0
        tone = "positive" if positive else "negative"
    return {
        "raw": current_text,
        "value": format(current, "f"),
        "baseline_raw": baseline_raw or format(baseline, "f"),
        "baseline_value": format(baseline, "f"),
        "delta": format(delta, "+f"),
        "direction": direction,
        "tone": tone,
        "comparison_status": "comparable",
    }


def _vehicles() -> list[SyntheticVehicle]:
    values: list[SyntheticVehicle] = []
    for index in range(27):
        focus = index < 14
        role_position = index + 1 if focus else index - 13
        values.append(
            SyntheticVehicle(
                vehicle_id=f"synthetic-{index + 1:02d}",
                series_name=f"{'重点' if focus else '竞品'}车系{(index // 5) + 1}",
                vehicle_name=f"{'重点' if focus else '竞品'}车型{role_position:02d}",
                role="focus" if focus else "competitor",
                role_position=0 if focus else 1,
                vehicle_position=role_position,
            )
        )
    return values


def _scenario_rows(scenario_id: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    baseline_mode = scenario_id == "baseline_initialization"
    for index, vehicle in enumerate(_vehicles()):
        score_base = Decimal("3.80") + Decimal(index) / Decimal(100)
        rank_base = Decimal(5 + (index % 5))
        volume_base = Decimal(500 + index * 37)
        review_article_count_base = Decimal(5000 + index * 113)
        negative_rate_base = Decimal(32 + (index % 8))
        pattern = index % 9
        status = "success"
        error_code = None
        error_message = None
        if baseline_mode:
            score = _metric(score_base, None)
            rank = _metric(rank_base, None, inverse=True)
            volume = _metric(volume_base, None)
            review_article_count = _metric(review_article_count_base, None)
            negative_rate = _metric(negative_rate_base, None, inverse=True)
        elif pattern == 0:
            score = _metric(score_base + Decimal("0.12"), score_base)
            rank = _metric(rank_base - 1, rank_base, inverse=True)
            volume = _metric(volume_base + 120, volume_base)
            review_article_count = _metric(review_article_count_base + 45, review_article_count_base)
            negative_rate = _metric(negative_rate_base - 3, negative_rate_base, inverse=True)
        elif pattern == 1:
            score = _metric(score_base - Decimal("0.08"), score_base)
            rank = _metric(rank_base + 2, rank_base, inverse=True)
            volume = _metric(volume_base - 80, volume_base)
            review_article_count = _metric(review_article_count_base - 20, review_article_count_base)
            negative_rate = _metric(negative_rate_base + 4, negative_rate_base, inverse=True)
        elif pattern == 2:
            score = _metric(score_base + Decimal("0.05"), score_base)
            rank = _metric(rank_base + 1, rank_base, inverse=True)
            volume = _metric(volume_base, volume_base)
            review_article_count = _metric(review_article_count_base + 12, review_article_count_base)
            negative_rate = _metric(negative_rate_base + 1, negative_rate_base, inverse=True)
        elif pattern == 3:
            score = _metric(score_base, score_base)
            rank = _metric(rank_base, rank_base, inverse=True)
            volume = _metric(volume_base, volume_base)
            review_article_count = _metric(review_article_count_base, review_article_count_base)
            negative_rate = _metric(negative_rate_base, negative_rate_base, inverse=True)
        elif pattern == 4:
            score = _metric(score_base, score_base)
            rank = _metric(rank_base, rank_base, inverse=True)
            volume = _metric(volume_base + 300, volume_base)
            review_article_count = _metric(review_article_count_base + 180, review_article_count_base)
            negative_rate = _metric(negative_rate_base, negative_rate_base, inverse=True)
        elif pattern == 5:
            score = _metric(score_base, None)
            rank = _metric(rank_base, None, inverse=True)
            volume = _metric(volume_base, None)
            review_article_count = _metric(review_article_count_base, None)
            negative_rate = _metric(negative_rate_base, None, inverse=True)
        elif pattern == 6:
            score = _metric(None, None, state="not_available", raw="暂无评分")
            rank = _metric(None, None, state="not_available", raw="暂无排名")
            volume = _metric(volume_base, volume_base)
            review_article_count = _metric(review_article_count_base, review_article_count_base)
            negative_rate = _metric(
                None, None, state="not_available", raw="暂无差评率"
            )
        elif pattern == 7:
            score = _metric(None, None, state="unknown")
            rank = _metric(None, None, state="unknown")
            volume = _metric(None, None, state="unknown")
            review_article_count = _metric(None, None, state="unknown")
            negative_rate = _metric(None, None, state="unknown")
            status = "failed"
            error_code = "SYNTHETIC_UNKNOWN"
            error_message = "合成场景：页面结构无法可靠解析。"
        else:
            score = _metric(None, None, state="auth_required")
            rank = _metric(None, None, state="auth_required")
            volume = _metric(None, None, state="auth_required")
            review_article_count = _metric(None, None, state="auth_required")
            negative_rate = _metric(None, None, state="auth_required")
            status = "failed"
            error_code = "AUTH_REQUIRED"
            error_message = "合成场景：共享平台会话需要更新。"
        rows.append(
            {
                "vehicle": vehicle,
                "status": status,
                "error_code": error_code,
                "error_message": error_message,
                "metrics": {
                    "score": score,
                    "rank": rank,
                    "volume": volume,
                    "review_article_count": review_article_count,
                    "negative_rate": negative_rate,
                },
                "evidence_required": True,
            }
        )
    return rows


class ReputationService:
    """口碑巡检查询、范围管理和隔离合成运行服务。"""

    def __init__(
        self,
        sessions: sessionmaker[Session],
        settings: Settings,
        *,
        session_store: SessionStore | None = None,
        event_publisher=None,
        adapter_factory=None,
    ) -> None:
        self.sessions = sessions
        self.settings = settings
        self.session_store = session_store
        self.event_publisher = event_publisher
        # adapter_factory 是既有测试注入口，仅覆盖懂车帝；其他平台始终按注册表解析。
        self.adapter_factories = {
            code: spec.adapter_factory for code, spec in REPUTATION_PLATFORMS.items()
        }
        self.adapter_factory = adapter_factory or self.adapter_factories[PLATFORM_CODE]
        self.adapter_factories[PLATFORM_CODE] = self.adapter_factory

    @staticmethod
    def _platform_spec(platform_code: str) -> ReputationPlatformSpec:
        """把外部平台代码收敛为已接入口碑平台注册项。"""

        try:
            return REPUTATION_PLATFORMS[platform_code.strip().lower()]
        except KeyError as error:
            raise DomainError("REPUTATION_SCOPE_PLATFORM", "口碑平台尚未接入。") from error

    @property
    def synthetic_enabled(self) -> bool:
        return (
            self.settings.runtime_mode == "test"
            and self.settings.enable_reputation_synthetic_runs
            and self.settings.reputation_test_database
        )

    def capabilities(self) -> dict[str, Any]:
        return {
            "reputation_synthetic_runs": self.synthetic_enabled,
            "real_adapter_status": "available",
            "real_adapter_message": (
                "懂车帝、汽车之家和易车真实页面适配器已接入；"
                "各平台映射验证均读取实时页面并只保存车型口碑指标区域截图。"
            ),
            "reputation_platforms": [
                {
                    "code": spec.code,
                    "display_name": spec.display_name,
                    "adapter_version": spec.adapter_version,
                    "validation_contract_version": spec.validation_contract_version,
                }
                for spec in REPUTATION_PLATFORMS.values()
            ],
            "scenarios": [
                {"id": key, **value} for key, value in SCENARIOS.items()
            ]
            if self.synthetic_enabled
            else [],
        }

    def _require_synthetic(self) -> None:
        if not self.synthetic_enabled:
            raise DomainError(
                "REPUTATION_SYNTHETIC_DISABLED",
                "手动运行测试只在显式测试模式和隔离测试数据库中开放。",
                status_code=404,
            )

    def list_runs(self, offset: int = 0, limit: int = 50) -> dict[str, Any]:
        with self.sessions() as db:
            root_filter = ReputationRun.source_type != "retry"
            total = (
                db.scalar(select(func.count()).select_from(ReputationRun).where(root_filter)) or 0
            )
            runs = db.scalars(
                select(ReputationRun)
                .where(root_filter)
                .order_by(ReputationRun.created_at.desc(), ReputationRun.id.desc())
                .offset(offset)
                .limit(limit)
            ).all()
            items = []
            for run in runs:
                value = self._run_dict(run)
                value["retry_runs"] = [
                    self._run_dict(item)
                    for item in db.scalars(
                        select(ReputationRun)
                        .where(
                            ReputationRun.source_type == "retry",
                            ReputationRun.root_run_id == run.id,
                        )
                        .order_by(ReputationRun.created_at, ReputationRun.id)
                    ).all()
                ]
                if run.source_type == "scheduled":
                    value.update(self._chain_summary(db, run))
                items.append(value)
            return {
                "items": items,
                "total": int(total),
                "offset": offset,
                "limit": limit,
            }

    def get_run(self, run_id: str, prefix: str = "/api/v1") -> dict[str, Any]:
        with self.sessions() as db:
            run = db.get(ReputationRun, run_id)
            if not run:
                raise DomainError(
                    "REPUTATION_RUN_NOT_FOUND", "口碑巡检运行不存在。", status_code=404
                )
            results = (
                self._selected_chain_results(db, run)
                if run.source_type == "scheduled"
                else db.scalars(
                    select(ReputationResult)
                    .where(ReputationResult.run_id == run_id)
                    .order_by(
                        ReputationResult.role_position,
                        ReputationResult.vehicle_position,
                        ReputationResult.id,
                    )
                ).all()
            )
            results = self._results_in_scope_order(db, run, list(results))
            evidence_by_result = {
                item.result_id: item
                for item in db.scalars(
                    select(ReputationEvidence).where(
                        ReputationEvidence.result_id.in_([row.id for row in results])
                    )
                ).all()
            }
            payload = self._run_dict(run)
            payload["results"] = [
                self._result_dict(row, evidence_by_result.get(row.id), prefix) for row in results
            ]
            payload["downloads"] = {
                **(
                    {"txt": f"{prefix}/reputation/runs/{run.id}/report.txt"}
                    if run.report_path
                    else {}
                ),
                **(
                    {"xlsx": f"{prefix}/reputation/runs/{run.id}/export.xlsx"}
                    if run.xlsx_path
                    else {}
                ),
                **(
                    {"evidence_zip": f"{prefix}/reputation/runs/{run.id}/evidence.zip"}
                    if run.required_evidence_count or run.evidence_zip_path
                    else {}
                ),
            }
            root_id = run.root_run_id or run.id
            payload["retry_runs"] = [
                self._run_dict(item)
                for item in db.scalars(
                    select(ReputationRun)
                    .where(
                        ReputationRun.source_type == "retry",
                        ReputationRun.root_run_id == root_id,
                    )
                    .order_by(ReputationRun.created_at, ReputationRun.id)
                ).all()
            ]
            root = db.get(ReputationRun, root_id)
            if root and root.source_type == "scheduled":
                payload.update(self._chain_summary(db, root))
            return payload

    @staticmethod
    def _selected_chain_results(
        db: Session, root: ReputationRun
    ) -> list[ReputationResult]:
        """为正式根批次逐车型选择关联链中当前最完整、同级最新的结果。"""

        chain = db.scalars(
            select(ReputationRun)
            .where(or_(ReputationRun.id == root.id, ReputationRun.root_run_id == root.id))
            .order_by(ReputationRun.created_at, ReputationRun.id)
        ).all()
        if not chain:
            return []
        chain_order = {item.id: index for index, item in enumerate(chain)}
        candidates = db.scalars(
            select(ReputationResult).where(
                ReputationResult.run_id.in_([item.id for item in chain])
            )
        ).all()
        status_priority = {"failed": 0, "partial_success": 1, "success": 2}
        selected: dict[str, ReputationResult] = {}
        for result in candidates:
            key = f"{result.vehicle_id}|{result.platform_code}"
            current = selected.get(key)
            priority = (
                status_priority.get(result.status, -1),
                result.collected_at,
                chain_order.get(result.run_id, -1),
                result.id,
            )
            current_priority = (
                status_priority.get(current.status, -1),
                current.collected_at,
                chain_order.get(current.run_id, -1),
                current.id,
            ) if current else None
            if current_priority is None or priority > current_priority:
                selected[key] = result
        return list(selected.values())

    @staticmethod
    def _results_in_scope_order(
        db: Session,
        run: ReputationRun,
        results: list[ReputationResult],
    ) -> list[ReputationResult]:
        """按批次冻结的车型映射顺序返回结果，旧批次保留原有排序兜底。"""

        if not run.scope_version_id:
            return results
        version = db.get(ReputationScopeVersion, run.scope_version_id)
        if not version:
            return results
        vehicles = (version.snapshot or {}).get("vehicles", [])
        vehicle_order = {
            str(vehicle.get("id")): index
            for index, vehicle in enumerate(vehicles)
            if vehicle.get("id")
        }
        if not vehicle_order:
            return results
        platform_order = {
            str(platform_code): index
            for index, platform_code in enumerate(run.platform_codes or [])
        }
        return sorted(
            results,
            key=lambda result: (
                vehicle_order.get(result.vehicle_id, len(vehicle_order)),
                platform_order.get(result.platform_code, len(platform_order)),
                result.role_position,
                result.vehicle_position,
                result.id,
            ),
        )

    @staticmethod
    def _chain_summary(db: Session, root: ReputationRun) -> dict[str, Any]:
        chain = db.scalars(
            select(ReputationRun)
            .where(or_(ReputationRun.id == root.id, ReputationRun.root_run_id == root.id))
            .order_by(ReputationRun.created_at, ReputationRun.id)
        ).all()
        successful: set[str] = set()
        for item in chain:
            results = db.scalars(
                select(ReputationResult).where(ReputationResult.run_id == item.id)
            ).all()
            evidence_ids = {
                value.result_id
                for value in db.scalars(
                    select(ReputationEvidence).where(
                        ReputationEvidence.result_id.in_([row.id for row in results])
                    )
                ).all()
            }
            for result in results:
                if result.status == "success" and (
                    not result.evidence_required or result.id in evidence_ids
                ):
                    successful.add(f"{result.vehicle_id}|{result.platform_code}")
        resolved = len(set(root.target_keys) & successful)
        unresolved = max(0, len(root.target_keys) - resolved)
        selected_results = ReputationService._selected_chain_results(db, root)
        selected_evidence_count = int(
            db.scalar(
                select(func.count())
                .select_from(ReputationEvidence)
                .where(
                    ReputationEvidence.result_id.in_(
                        [result.id for result in selected_results]
                    )
                )
            )
            or 0
        ) if selected_results else 0
        return {
            "resolved_count": resolved,
            "unresolved_count": unresolved,
            "linked_status": "success" if unresolved == 0 else "partial_success" if resolved else "failed",
            "linked_complete_evidence_count": selected_evidence_count,
        }

    def create_synthetic(self, scenario_id: str) -> dict[str, Any]:
        self._require_synthetic()
        if scenario_id not in SCENARIOS:
            raise DomainError("REPUTATION_SCENARIO_UNKNOWN", "未知的口碑合成场景。")
        now = datetime.now(timezone.utc)
        rows = _scenario_rows(scenario_id)
        input_hash = _text_hash(
            {
                "fixture_version": FIXTURE_VERSION,
                "scenario_id": scenario_id,
                "rows": [
                    {
                        "vehicle_id": item["vehicle"].vehicle_id,
                        "metrics": item["metrics"],
                        "status": item["status"],
                    }
                    for item in rows
                ],
            }
        )
        run_id = uuid7()
        run_number = f"RP-T-{now:%Y%m%d-%H%M%S}-{run_id[-4:].upper()}"
        run_type = {
            "baseline_initialization": "baseline_initialization",
            "daily_mixed_changes": "daily",
            "month_end_mixed_changes": "month_end",
        }[scenario_id]
        run_dir = self.settings.reputation_dir / run_id
        run_dir.mkdir(parents=True, exist_ok=False)
        failures = sum(1 for row in rows if row["status"] != "success")
        required = sum(1 for row in rows if row["evidence_required"])
        run = ReputationRun(
            id=run_id,
            number=run_number,
            source_type="synthetic",
            scenario_id=scenario_id,
            fixture_version=FIXTURE_VERSION,
            input_hash=input_hash,
            run_type=run_type,
            planned_date=now.astimezone(ZoneInfo(self.settings.timezone)).date().isoformat(),
            status="partial_success" if failures else "success",
            platform_codes=[PLATFORM_CODE],
            planned_count=len(rows),
            completed_count=len(rows) - failures,
            failed_count=failures,
            required_evidence_count=required,
            complete_evidence_count=required,
            report_status="success",
            created_at=now,
            started_at=now,
            finished_at=now,
        )
        try:
            with self.sessions.begin() as db:
                db.add(run)
                db.flush()
                for item in rows:
                    vehicle: SyntheticVehicle = item["vehicle"]
                    result = ReputationResult(
                        run_id=run_id,
                        vehicle_id=vehicle.vehicle_id,
                        series_name=vehicle.series_name,
                        vehicle_name=vehicle.vehicle_name,
                        role=vehicle.role,
                        role_position=vehicle.role_position,
                        vehicle_position=vehicle.vehicle_position,
                        platform_code=PLATFORM_CODE,
                        platform_name=PLATFORM_NAME,
                        status=item["status"],
                        metrics=item["metrics"],
                        evidence_required=item["evidence_required"],
                        error_code=item["error_code"],
                        error_message=item["error_message"],
                        collected_at=now,
                    )
                    db.add(result)
                    db.flush()
                    if item["evidence_required"]:
                        evidence = self._create_evidence(run_dir, result, item["metrics"])
                        db.add(evidence)
                db.flush()
                stored_results = db.scalars(
                    select(ReputationResult)
                    .where(ReputationResult.run_id == run_id)
                    .order_by(ReputationResult.role_position, ReputationResult.vehicle_position)
                ).all()
                evidence_by_result = {
                    item.result_id: item
                    for item in db.scalars(
                        select(ReputationEvidence).where(
                            ReputationEvidence.result_id.in_([row.id for row in stored_results])
                        )
                    ).all()
                }
                report = self._render_report(run, stored_results)
                report_path = run_dir / f"{run_number}.txt"
                report_path.write_text(report, encoding="utf-8")
                xlsx_path = run_dir / f"{run_number}.xlsx"
                self._create_xlsx(run, stored_results, evidence_by_result, xlsx_path)
                run.report_text = report
                run.report_path = str(report_path)
                run.xlsx_path = str(xlsx_path)
        except Exception:
            shutil.rmtree(run_dir, ignore_errors=True)
            raise
        if self.event_publisher:
            self.event_publisher("reputation.run.changed", run_id, status=run.status)
        return self.get_run(run_id)

    def create_real_acceptance(self, validation_run_ids: list[str]) -> dict[str, Any]:
        """把已完成的真实映射验证结果冻结为一次可见的基线验收批次。"""

        run_ids = list(dict.fromkeys(item.strip() for item in validation_run_ids if item.strip()))
        if not run_ids:
            raise DomainError("REPUTATION_VALIDATION_RUNS_REQUIRED", "请提供真实映射验证运行编号。")
        with self.sessions() as db:
            draft = db.get(ReputationScopeDraft, "current")
            version = db.get(ReputationScopeVersion, draft.published_version_id) if draft else None
            if not version:
                raise DomainError("REPUTATION_SCOPE_NOT_PUBLISHED", "请先发布口碑巡检范围。")
            validation_runs = [db.get(ReputationMappingValidationRun, item) for item in run_ids]
            if any(item is None for item in validation_runs):
                raise DomainError("REPUTATION_VALIDATION_RUN_NOT_FOUND", "存在无效的验证运行编号。")
            attempts = db.scalars(
                select(ReputationMappingValidationAttempt)
                .where(
                    ReputationMappingValidationAttempt.run_id.in_(run_ids),
                    ReputationMappingValidationAttempt.status == "success",
                )
                .order_by(ReputationMappingValidationAttempt.finished_at)
            ).all()
            snapshot = json.loads(json.dumps(version.snapshot, ensure_ascii=False))
            vehicles = [
                item for item in snapshot.get("vehicles", []) if item.get("enabled", True)
            ]
            run_platforms = {
                item.id: item.platform_code for item in validation_runs if item is not None
            }
            by_target = {
                (run_platforms[item.run_id], item.vehicle_id): item for item in attempts
            }
            platform_codes = [
                code
                for code in REPUTATION_PLATFORMS
                if any(vehicle.get("mappings", {}).get(code) for vehicle in vehicles)
            ]
            targets = [
                (vehicle, code) for vehicle in vehicles for code in platform_codes
            ]
            missing = [
                {"vehicle_id": vehicle["id"], "platform_code": code}
                for vehicle, code in targets
                if (code, vehicle["id"]) not in by_target
            ]
            if missing:
                raise DomainError(
                    "REPUTATION_ACCEPTANCE_INCOMPLETE",
                    "真实验证结果尚未覆盖全部已发布车型。",
                    details=[{**item, "reason": "缺少成功结果"} for item in missing],
                )
            for vehicle, code in targets:
                mapping = vehicle.get("mappings", {}).get(code) or {}
                attempt = by_target[(code, vehicle["id"])]
                if attempt.mapping_hash != _mapping_hash(vehicle["id"], mapping, code):
                    raise DomainError(
                        "REPUTATION_ACCEPTANCE_MAPPING_CHANGED",
                        f"车型{vehicle['id']}的成功结果不属于当前已发布映射。",
                    )
                raw = attempt.metric_region_path
                digest = attempt.metric_region_sha256
                if not raw or not digest or not Path(raw).is_file() or _sha256(Path(raw)) != digest:
                    raise DomainError(
                        "REPUTATION_ACCEPTANCE_EVIDENCE_INVALID",
                        f"车型{vehicle['id']}的真实页面证据缺失或校验失败。",
                    )
            input_hash = _text_hash(
                {
                    "scope_version_id": version.id,
                    "validation_run_ids": run_ids,
                    "attempt_ids": [by_target[(code, vehicle["id"])].id for vehicle, code in targets],
                }
            )
            existing = db.scalar(
                select(ReputationRun).where(
                    ReputationRun.source_type == "real_acceptance",
                    ReputationRun.input_hash == input_hash,
                )
            )
            if existing:
                return self.get_run(existing.id)
            started_at = min(item.started_at for item in validation_runs if item is not None)
            finished_at = max(item.finished_at for item in validation_runs if item and item.finished_at)

        now = datetime.now(timezone.utc)
        run_id = uuid7()
        run_number = f"RP-A-{now:%Y%m%d-%H%M%S}-{run_id[-4:].upper()}"
        run_dir = self.settings.reputation_dir / run_id
        run_dir.mkdir(parents=True, exist_ok=False)
        run = ReputationRun(
            id=run_id,
            number=run_number,
            source_type="real_acceptance",
            scenario_id="real_validation_acceptance",
            input_hash=input_hash,
            run_type="baseline_initialization",
            scope_version_id=version.id,
            planned_date=now.astimezone(ZoneInfo(self.settings.timezone)).date().isoformat(),
            status="success",
            platform_codes=platform_codes,
            planned_count=len(targets),
            completed_count=len(targets),
            failed_count=0,
            required_evidence_count=len(targets),
            complete_evidence_count=len(targets),
            report_status="success",
            created_at=now,
            started_at=started_at,
            finished_at=finished_at,
        )

        def metric(raw: str | None, *, inverse: bool = False) -> dict[str, Any]:
            return (
                _metric(Decimal(raw), None, inverse=inverse, raw=raw)
                if raw is not None
                else _metric(None, None, inverse=inverse, state="not_available")
            )

        try:
            with self.sessions.begin() as db:
                db.add(run)
                db.flush()
                for vehicle, platform_code in targets:
                    attempt = by_target[(platform_code, vehicle["id"])]
                    spec = self._platform_spec(platform_code)
                    result = ReputationResult(
                        run_id=run_id,
                        vehicle_id=vehicle["id"],
                        series_name=vehicle["series_name"],
                        vehicle_name=vehicle["vehicle_name"],
                        role=vehicle["role"],
                        role_position=0 if vehicle["role"] == "focus" else 1,
                        vehicle_position=int(vehicle["role_order"]),
                        platform_code=platform_code,
                        platform_name=spec.display_name,
                        status="success",
                        metrics={
                            "score": metric(attempt.metrics.get("score")),
                            "rank": metric(attempt.metrics.get("rank"), inverse=True),
                            "volume": metric(attempt.metrics.get("volume")),
                            "review_article_count": metric(
                                attempt.metrics.get("review_article_count")
                            ),
                            "negative_rate": metric(
                                attempt.metrics.get("negative_rate"), inverse=True
                            ),
                        },
                        evidence_required=True,
                        collected_at=attempt.finished_at,
                    )
                    db.add(result)
                    db.flush()
                    evidence_dir = run_dir / "evidence" / platform_code / vehicle["id"]
                    evidence_dir.mkdir(parents=True, exist_ok=True)
                    metric_path = evidence_dir / "region.png"
                    shutil.copy2(attempt.metric_region_path, metric_path)
                    with Image.open(metric_path) as source:
                        width, height = source.size
                    db.add(
                        ReputationEvidence(
                            result_id=result.id,
                            full_page_path=str(metric_path),
                            metric_region_path=str(metric_path),
                            full_page_sha256=attempt.metric_region_sha256,
                            metric_region_sha256=attempt.metric_region_sha256,
                            width=width,
                            height=height,
                        )
                    )
                db.flush()
                stored_results = db.scalars(
                    select(ReputationResult)
                    .where(ReputationResult.run_id == run_id)
                    .order_by(ReputationResult.role_position, ReputationResult.vehicle_position)
                ).all()
                evidence_by_result = {
                    item.result_id: item
                    for item in db.scalars(
                        select(ReputationEvidence).where(
                            ReputationEvidence.result_id.in_([row.id for row in stored_results])
                        )
                    ).all()
                }
                report = self._render_report(run, stored_results)
                report_path = run_dir / f"{run_number}.txt"
                report_path.write_text(report, encoding="utf-8")
                xlsx_path = run_dir / f"{run_number}.xlsx"
                self._create_xlsx(run, stored_results, evidence_by_result, xlsx_path)
                run.report_text = report
                run.report_path = str(report_path)
                run.xlsx_path = str(xlsx_path)
        except Exception:
            shutil.rmtree(run_dir, ignore_errors=True)
            raise
        if self.event_publisher:
            self.event_publisher("reputation.run.changed", run_id, status=run.status)
        return self.get_run(run_id)

    def recover_interrupted(self) -> int:
        """进程启动时把未提交终态的正式运行恢复为排队，避免重复创建批次。"""

        with self.sessions.begin() as db:
            rows = db.scalars(
                select(ReputationRun).where(
                    ReputationRun.source_type.in_(["scheduled", "retry"]),
                    ReputationRun.status == "running",
                )
            ).all()
            for row in rows:
                row.status = "queued"
                row.error_message = "服务重启后从不可变批次身份恢复执行。"
            reports = db.scalars(
                select(ReputationRun).where(ReputationRun.report_status == "generating")
            ).all()
            for row in reports:
                row.report_status = "failed"
                row.error_message = "服务重启中断派生产物生成，可按既定重试预算继续。"
            return len(rows) + len(reports)

    def check_schedule(self, now: datetime | None = None) -> dict[str, Any]:
        """对账固定10:00巡检、跨日漏触发水位和终态汇报。"""

        current = (now or datetime.now(timezone.utc)).astimezone(timezone.utc)
        zone = ZoneInfo(self.settings.timezone)
        local_now = current.astimezone(zone)
        missed_dates: list[str] = []
        with self.sessions.begin() as db:
            state = db.get(ReputationSchedulerState, 1)
            if state:
                previous_day = state.last_checked_at.astimezone(zone).date()
                cursor = previous_day + timedelta(days=1)
                while cursor < local_now.date():
                    schedule_type = self._schedule_type(cursor)
                    existing = db.scalar(
                        select(ReputationScheduleEvent).where(
                            ReputationScheduleEvent.planned_date == cursor.isoformat(),
                            ReputationScheduleEvent.run_type == schedule_type,
                        )
                    )
                    tombstone = db.scalar(
                        select(ReputationTombstone).where(
                            ReputationTombstone.planned_date == cursor.isoformat(),
                            ReputationTombstone.run_type == schedule_type,
                        )
                    )
                    if not existing and not tombstone:
                        planned = datetime.combine(cursor, INSPECTION_TIME, tzinfo=zone).astimezone(
                            timezone.utc
                        )
                        db.add(
                            ReputationScheduleEvent(
                                planned_date=cursor.isoformat(),
                                run_type=schedule_type,
                                planned_at=planned,
                                status="missed",
                                message="服务跨自然日恢复，按合同只记录漏触发，不补采历史页面。",
                            )
                        )
                        missed_dates.append(cursor.isoformat())
                    cursor += timedelta(days=1)
                state.last_checked_at = current
            else:
                db.add(ReputationSchedulerState(id=1, last_checked_at=current))

        planned_at = datetime.combine(
            local_now.date(), INSPECTION_TIME, tzinfo=zone
        ).astimezone(timezone.utc)
        created_run_id = None
        if current >= planned_at:
            created_run_id = self._ensure_official_run(planned_at, current)

        with self.sessions() as db:
            queued = list(
                db.scalars(
                    select(ReputationRun.id)
                    .where(
                        ReputationRun.source_type.in_(["scheduled", "retry"]),
                        ReputationRun.status == "queued",
                    )
                    .order_by(ReputationRun.created_at, ReputationRun.id)
                ).all()
            )
            reports = list(
                db.scalars(
                    select(ReputationRun.id)
                    .where(
                        ReputationRun.source_type == "scheduled",
                        ReputationRun.status.in_(["success", "partial_success", "failed"]),
                        or_(
                            ReputationRun.report_status.in_(["pending", "waiting"]),
                            (
                                (ReputationRun.report_status == "failed")
                                & (ReputationRun.report_attempt_count < 2)
                            ),
                        ),
                    )
                    .order_by(ReputationRun.finished_at, ReputationRun.id)
                ).all()
            )
        return {
            "created_run_id": created_run_id,
            "queued_run_ids": queued,
            "report_run_ids": reports,
            "missed_dates": missed_dates,
        }

    @staticmethod
    def _schedule_type(day: date) -> str:
        return "month_end" if (day + timedelta(days=1)).month != day.month else "daily"

    def _ensure_official_run(self, planned_at: datetime, current: datetime) -> str | None:
        zone = ZoneInfo(self.settings.timezone)
        day = planned_at.astimezone(zone).date()
        planned_date = day.isoformat()
        schedule_type = self._schedule_type(day)
        idempotency_key = f"reputation:{planned_date}:{schedule_type}"
        with self.sessions.begin() as db:
            tombstone = db.scalar(
                select(ReputationTombstone).where(
                    ReputationTombstone.planned_date == planned_date,
                    ReputationTombstone.run_type == schedule_type,
                )
            )
            if tombstone:
                return None
            event = db.scalar(
                select(ReputationScheduleEvent).where(
                    ReputationScheduleEvent.planned_date == planned_date,
                    ReputationScheduleEvent.run_type == schedule_type,
                )
            )
            if event:
                return event.run_id
            existing = db.scalar(
                select(ReputationRun).where(ReputationRun.idempotency_key == idempotency_key)
            )
            if existing:
                return existing.id
            version = db.scalar(
                select(ReputationScopeVersion)
                .where(ReputationScopeVersion.published_at <= planned_at)
                .order_by(ReputationScopeVersion.published_at.desc(), ReputationScopeVersion.version.desc())
                .limit(1)
            )
            if not version:
                db.add(
                    ReputationScheduleEvent(
                        planned_date=planned_date,
                        run_type=schedule_type,
                        planned_at=planned_at,
                        status="configuration_error",
                        message="计划时刻没有已发布的口碑巡检范围，未创建空批次。",
                    )
                )
                return None
            snapshot = json.loads(json.dumps(version.snapshot, ensure_ascii=False))
            vehicles = [
                vehicle for vehicle in snapshot.get("vehicles", []) if vehicle.get("enabled", True)
            ]
            platform_codes = [
                code
                for code in REPUTATION_PLATFORMS
                if any(vehicle.get("mappings", {}).get(code) for vehicle in vehicles)
            ]
            target_keys = [
                f"{vehicle['id']}|{code}"
                for vehicle in vehicles
                for code in platform_codes
                if vehicle.get("mappings", {}).get(code)
            ]
            expected_count = len(vehicles) * len(platform_codes)
            platforms = {code: db.get(PlatformConfig, code) for code in platform_codes}
            if (
                not target_keys
                or len(target_keys) != expected_count
                or any(not platforms[code] or not platforms[code].enabled for code in platform_codes)
            ):
                db.add(
                    ReputationScheduleEvent(
                        planned_date=planned_date,
                        run_type=schedule_type,
                        planned_at=planned_at,
                        status="configuration_error",
                        message="计划范围没有当前具备运行资格的车型平台项，未创建空批次。",
                        scope_version_id=version.id,
                    )
                )
                return None
            previous_real_runs = db.scalar(
                select(func.count()).select_from(ReputationRun).where(
                    ReputationRun.source_type.in_(["scheduled", "real_acceptance"]),
                    ReputationRun.planned_date < planned_date,
                )
            ) or 0
            deleted_official_runs = (
                db.scalar(
                    select(func.count()).select_from(ReputationTombstone).where(
                        ReputationTombstone.planned_date < planned_date
                    )
                )
                or 0
            )
            baseline_date = (day - timedelta(days=1)).isoformat()
            baseline_source, baseline_snapshot = self._freeze_baseline(db, baseline_date)
            run_id = uuid7()
            # 真实初始化验收批次承载的就是首日基线。只要前一自然日已经有
            # 完整的真实基线快照，下一次正式巡检就应直接进入日常比较；不能
            # 因其不是调度器创建而再次初始化。
            run_type = (
                "baseline_initialization"
                if previous_real_runs == 0
                and deleted_official_runs == 0
                and not baseline_snapshot
                else schedule_type
            )
            run_number = f"RP-S-{day:%Y%m%d}-{run_id[-4:].upper()}"
            run = ReputationRun(
                id=run_id,
                number=run_number,
                source_type="scheduled",
                run_type=run_type,
                schedule_type=schedule_type,
                planned_date=planned_date,
                target_keys=target_keys,
                root_run_id=run_id,
                scope_version_id=version.id,
                idempotency_key=idempotency_key,
                planned_at=planned_at,
                report_planned_at=None,
                delayed=current > planned_at + timedelta(minutes=1),
                concurrency=REPUTATION_RUN_CONCURRENCY,
                baseline_date=baseline_date,
                baseline_frozen_at=current,
                baseline_source_run_id=baseline_source,
                baseline_snapshot=baseline_snapshot,
                status="queued",
                platform_codes=platform_codes,
                planned_count=len(target_keys),
                report_status="waiting",
                created_at=current,
            )
            db.add(run)
            db.add(
                ReputationScheduleEvent(
                    planned_date=planned_date,
                    run_type=schedule_type,
                    planned_at=planned_at,
                    status="queued",
                    message=(
                        "服务同日恢复，已创建延迟正式口碑巡检批次。"
                        if run.delayed
                        else "已按10:00计划创建正式口碑巡检批次。"
                    ),
                    scope_version_id=version.id,
                    run_id=run_id,
                    created_at=current,
                )
            )
            return run_id

    @staticmethod
    def _freeze_baseline(
        db: Session, baseline_date: str
    ) -> tuple[str | None, dict[str, Any]]:
        roots = db.scalars(
            select(ReputationRun)
            .where(
                ReputationRun.source_type.in_(["scheduled", "real_acceptance"]),
                ReputationRun.planned_date == baseline_date,
            )
            .order_by(ReputationRun.created_at.desc())
        ).all()
        # 同一天若已经存在正式调度批次，它优先于初始化验收批次；验收批次
        # 只负责在正式日检尚未出现时提供首日真实基线。
        root = next((item for item in roots if item.source_type == "scheduled"), None)
        if root is None:
            root = next((item for item in roots if item.source_type == "real_acceptance"), None)
        if not root:
            return None, {}
        chain = db.scalars(
            select(ReputationRun)
            .where(or_(ReputationRun.id == root.id, ReputationRun.root_run_id == root.id))
            .order_by(ReputationRun.created_at, ReputationRun.id)
        ).all()
        snapshot: dict[str, Any] = {}
        for chain_run in chain:
            for result in db.scalars(
                select(ReputationResult).where(ReputationResult.run_id == chain_run.id)
            ).all():
                if result.status not in {"success", "partial_success"}:
                    continue
                key = f"{result.vehicle_id}|{result.platform_code}"
                snapshot[key] = {
                    "metrics": result.metrics,
                    "source_run_id": chain_run.id,
                    "collected_at": result.collected_at.isoformat(),
                }
        return root.id, snapshot

    @staticmethod
    def _decimal_value(raw: str | None) -> Decimal | None:
        if raw is None:
            return None
        try:
            return Decimal(str(raw).replace(",", "").replace("%", "").strip())
        except Exception:
            return None

    @classmethod
    def _official_metric(
        cls,
        raw: str | None,
        baseline: dict[str, Any] | None,
        *,
        inverse: bool = False,
        scope: str | None = None,
        missing_state: Literal["not_available", "unknown"] = "not_available",
    ) -> dict[str, Any]:
        current = cls._decimal_value(raw)
        if current is None:
            value = _metric(None, None, state=missing_state, raw=raw)
        else:
            baseline_metric = baseline or {}
            baseline_value = cls._decimal_value(baseline_metric.get("value"))
            if scope and baseline_metric.get("scope") and baseline_metric.get("scope") != scope:
                value = {
                    "raw": raw or format(current, "f"),
                    "value": format(current, "f"),
                    "baseline_raw": baseline_metric.get("raw"),
                    "baseline_value": baseline_metric.get("value"),
                    "delta": None,
                    "direction": "none",
                    "tone": "neutral",
                    "comparison_status": "not_comparable",
                }
            else:
                value = _metric(
                    current,
                    baseline_value,
                    inverse=inverse,
                    raw=raw,
                    baseline_raw=baseline_metric.get("raw"),
                )
        if scope:
            value["scope"] = scope
        return value

    @classmethod
    def _official_metrics(
        cls, page: ReputationPageResult, baseline_row: dict[str, Any] | None
    ) -> dict[str, Any]:
        baseline = (baseline_row or {}).get("metrics", {})
        missing_state: Literal["not_available", "unknown"] = (
            "not_available" if page.reputation_not_available else "unknown"
        )
        review_article_count = cls._official_metric(
            page.review_article_count_raw,
            baseline.get("review_article_count"),
            missing_state=missing_state,
        )
        review_article_count["source_url"] = page.review_article_count_url
        negative_rate = cls._official_metric(
            page.negative_rate_raw,
            baseline.get("negative_rate"),
            inverse=True,
            missing_state=missing_state,
        )
        negative_rate.update(
            {
                "source_url": page.negative_rate_url,
                "positive_count": page.negative_rate_positive_count,
                "negative_count": page.negative_rate_negative_count,
            }
        )
        return {
            "score": cls._official_metric(
                page.score_raw, baseline.get("score"), missing_state=missing_state
            ),
            "rank": cls._official_metric(
                page.rank_raw,
                baseline.get("rank"),
                inverse=True,
                scope=page.rank_scope,
                missing_state=missing_state,
            ),
            "volume": cls._official_metric(
                page.volume_raw, baseline.get("volume"), missing_state=missing_state
            ),
            "review_article_count": review_article_count,
            "negative_rate": negative_rate,
        }

    @staticmethod
    def _needs_evidence(
        run_type: str,
        schedule_type: str | None,
        role: str,
        metrics: dict[str, Any],
    ) -> bool:
        """所有新巡检执行项都要求保留同源指标区域证据。"""

        del run_type, schedule_type, role, metrics
        return True

    def _persist_official_result(
        self,
        run_id: str,
        platform_code: str,
        target: ReputationMappingTarget,
        vehicle: dict[str, Any],
        baseline: dict[str, Any] | None,
        result: ReputationPageResult | Exception,
        attempt_count: int,
        required_count: int,
    ) -> dict[str, int]:
        """幂等提交一个车型终态，并在事务完成后发布权威进度。"""

        collected_at = datetime.now(timezone.utc)
        with self.sessions.begin() as db:
            run = db.get(ReputationRun, run_id)
            if not run:
                raise RuntimeError("口碑巡检批次在执行期间丢失")
            row = db.scalar(
                select(ReputationResult).where(
                    ReputationResult.run_id == run_id,
                    ReputationResult.vehicle_id == target.vehicle_id,
                    ReputationResult.platform_code == platform_code,
                )
            )
            if row:
                db.execute(
                    delete(ReputationEvidence).where(ReputationEvidence.result_id == row.id)
                )

            if isinstance(result, ReputationPageResult):
                metrics = self._official_metrics(result, baseline)
                evidence_required = self._needs_evidence(
                    run.run_type, run.schedule_type, vehicle["role"], metrics
                )
                has_evidence = bool(
                    result.metric_region_path
                    and result.metric_region_sha256
                    and result.metric_region_path.is_file()
                    and _sha256(result.metric_region_path) == result.metric_region_sha256
                )
                row_status = "success"
                error_code = None
                error_message = None
                unknown_names = [
                    name
                    for name, metric in metrics.items()
                    if metric.get("comparison_status") == "unknown"
                ]
                if unknown_names:
                    row_status = "partial_success"
                    error_code = "REPUTATION_METRIC_UNKNOWN"
                    error_message = f"指标来源尚未取得可靠值：{','.join(unknown_names)}。"
                if evidence_required and not has_evidence:
                    row_status = "partial_success"
                    error_code = "REPUTATION_EVIDENCE_MISSING"
                    error_message = (
                        f"{error_message or ''}本项必需页面证据缺失。"
                    )
                duration_ms = result.duration_ms
            else:
                metrics = {
                    name: _metric(None, None, state="unknown")
                    for name in (
                        "score",
                        "rank",
                        "volume",
                        "review_article_count",
                        "negative_rate",
                    )
                }
                evidence_required = True
                has_evidence = False
                row_status = "failed"
                parsed = self._validation_error(result)
                error_code = parsed["code"]
                error_message = parsed["message"]
                duration_ms = None

            values = {
                "run_id": run_id,
                "vehicle_id": target.vehicle_id,
                "series_name": vehicle["series_name"],
                "vehicle_name": vehicle["vehicle_name"],
                "role": vehicle["role"],
                "role_position": 0 if vehicle["role"] == "focus" else 1,
                "vehicle_position": int(vehicle["role_order"]),
                "platform_code": platform_code,
                "platform_name": REPUTATION_PLATFORMS[platform_code].display_name,
                "status": row_status,
                "metrics": metrics,
                "evidence_required": evidence_required,
                "mapping_snapshot": vehicle.get("mappings", {}).get(platform_code) or {},
                "attempt_count": attempt_count,
                "duration_ms": duration_ms,
                "error_code": error_code,
                "error_message": error_message,
                "collected_at": collected_at,
            }
            if row is None:
                row = ReputationResult(**values)
                db.add(row)
            else:
                for name, value in values.items():
                    setattr(row, name, value)
            db.flush()
            if isinstance(result, ReputationPageResult) and has_evidence and evidence_required:
                db.add(
                    ReputationEvidence(
                        result_id=row.id,
                        full_page_path=str(result.metric_region_path),
                        metric_region_path=str(result.metric_region_path),
                        full_page_sha256=str(result.metric_region_sha256),
                        metric_region_sha256=str(result.metric_region_sha256),
                        width=result.width,
                        height=result.height,
                    )
                )
                db.flush()

            completed = int(
                db.scalar(
                    select(func.count())
                    .select_from(ReputationResult)
                    .where(
                        ReputationResult.run_id == run_id,
                        ReputationResult.status.in_(["success", "partial_success"]),
                    )
                )
                or 0
            )
            failed = int(
                db.scalar(
                    select(func.count())
                    .select_from(ReputationResult)
                    .where(
                        ReputationResult.run_id == run_id,
                        ReputationResult.status == "failed",
                    )
                )
                or 0
            )
            complete_evidence = int(
                db.scalar(
                    select(func.count())
                    .select_from(ReputationEvidence)
                    .join(ReputationResult, ReputationEvidence.result_id == ReputationResult.id)
                    .where(ReputationResult.run_id == run_id)
                )
                or 0
            )
            run.completed_count = completed
            run.failed_count = failed
            run.required_evidence_count = required_count
            run.complete_evidence_count = complete_evidence
            status = run.status

        summary = {
            "completed_count": completed,
            "failed_count": failed,
            "required_evidence_count": required_count,
            "complete_evidence_count": complete_evidence,
        }
        if self.event_publisher:
            self.event_publisher(
                "reputation.run.changed",
                run_id,
                status=status,
                **summary,
            )
        return summary

    def execute_run(self, run_id: str) -> dict[str, Any]:
        """执行一个已持久化的正式或失败项补跑批次。"""

        started = datetime.now(timezone.utc)
        with self.sessions.begin() as db:
            run = db.get(ReputationRun, run_id)
            if not run:
                raise DomainError(
                    "REPUTATION_RUN_NOT_FOUND", "口碑巡检运行不存在。", status_code=404
                )
            if run.status in {"success", "partial_success", "failed"}:
                return self.get_run(run_id)
            if run.source_type not in {"scheduled", "retry"}:
                raise DomainError("REPUTATION_RUN_NOT_EXECUTABLE", "该运行不属于正式执行队列。")
            version = db.get(ReputationScopeVersion, run.scope_version_id)
            if not version:
                run.status = "failed"
                run.error_message = "批次冻结的范围版本不存在。"
                run.finished_at = started
                return self.get_run(run_id)
            run.status = "running"
            run.started_at = started
            run.error_message = None
            # 防御旧的排队数据或异常写入；终态历史批次会在上方直接返回，不会被改写。
            run.concurrency = REPUTATION_RUN_CONCURRENCY
            snapshot = json.loads(json.dumps(version.snapshot, ensure_ascii=False))
            target_keys = set(run.target_keys)
            baseline_snapshot = json.loads(json.dumps(run.baseline_snapshot, ensure_ascii=False))
            concurrency = REPUTATION_RUN_CONCURRENCY
            run_type = run.run_type
            schedule_type = run.schedule_type

        vehicles = {
            item["id"]: item
            for item in snapshot.get("vehicles", [])
            if item.get("enabled", True)
        }
        targets: list[ReputationMappingTarget] = []
        target_platform_codes: list[str] = []
        for vehicle in vehicles.values():
            for platform_code in run.platform_codes:
                key = f"{vehicle['id']}|{platform_code}"
                if target_keys and key not in target_keys:
                    continue
                mapping = vehicle.get("mappings", {}).get(platform_code)
                if not mapping:
                    continue
                spec = self._platform_spec(platform_code)
                targets.append(
                    ReputationMappingTarget(
                        vehicle_id=vehicle["id"],
                        platform_vehicle_id=str(mapping["platform_vehicle_id"]),
                        platform_url=spec.normalize_url(
                            str(mapping["platform_url"]), str(mapping["platform_vehicle_id"])
                        ),
                        platform_display_name=str(mapping["platform_display_name"]),
                        mapping_hash=_mapping_hash(vehicle["id"], mapping, platform_code),
                    )
                )
                target_platform_codes.append(platform_code)

        required_count = len(targets)
        with self.sessions.begin() as db:
            run = db.get(ReputationRun, run_id)
            if not run:
                raise RuntimeError("口碑巡检批次在执行期间丢失")
            db.execute(
                delete(ReputationEvidence).where(
                    ReputationEvidence.result_id.in_(
                        select(ReputationResult.id).where(ReputationResult.run_id == run_id)
                    )
                )
            )
            db.execute(delete(ReputationResult).where(ReputationResult.run_id == run_id))
            run.completed_count = 0
            run.failed_count = 0
            run.required_evidence_count = required_count
            run.complete_evidence_count = 0
        if self.event_publisher:
            self.event_publisher(
                "reputation.run.changed",
                run_id,
                status="running",
                completed_count=0,
                failed_count=0,
                required_evidence_count=required_count,
                complete_evidence_count=0,
            )

        root = self.settings.reputation_dir / run_id / "collection"
        root.mkdir(parents=True, exist_ok=True)
        persisted_indexes: set[int] = set()

        def persist(index: int, result: ReputationPageResult | Exception, attempt: int) -> None:
            target = targets[index]
            platform_code = target_platform_codes[index]
            key = f"{target.vehicle_id}|{platform_code}"
            self._persist_official_result(
                run_id,
                platform_code,
                target,
                vehicles[target.vehicle_id],
                baseline_snapshot.get(key),
                result,
                attempt,
                required_count,
            )
            persisted_indexes.add(index)

        deadline = monotonic() + 45 * 60
        final_results: list[ReputationPageResult | Exception] = [
            ReputationAdapterError("REPUTATION_NOT_STARTED", "口碑平台项尚未执行。")
            for _ in targets
        ]
        attempt_counts = [1 for _ in targets]
        for platform_code in run.platform_codes:
            indexes = [
                index
                for index, code in enumerate(target_platform_codes)
                if code == platform_code
            ]
            if not indexes:
                continue
            spec = self._platform_spec(platform_code)
            storage_state = (
                self.session_store.get_state(platform_code) if self.session_store else None
            )
            if not storage_state:
                for index in indexes:
                    final_results[index] = ReputationAdapterError(
                        "AUTH_REQUIRED", f"{spec.display_name}共享Session需要更新。"
                    )
                continue
            remaining = int(deadline - monotonic())
            if remaining <= 0:
                for index in indexes:
                    final_results[index] = ReputationAdapterError(
                        "REPUTATION_BATCH_TIMEOUT", "口碑巡检达到45分钟批次上限。"
                    )
                continue

            def evidence_policy(target: ReputationMappingTarget, measurement: dict[str, Any]) -> bool:
                vehicle = vehicles[target.vehicle_id]
                baseline = baseline_snapshot.get(f"{target.vehicle_id}|{platform_code}")
                metrics = {
                    "score": self._official_metric(
                        measurement.get("score_raw") or measurement.get("score"),
                        (baseline or {}).get("metrics", {}).get("score"),
                    ),
                    "rank": self._official_metric(
                        measurement.get("rank_raw") or measurement.get("rank"),
                        (baseline or {}).get("metrics", {}).get("rank"),
                        inverse=True,
                        scope=str(measurement.get("rank_scope") or "同级车评分"),
                    ),
                }
                return self._needs_evidence(run_type, schedule_type, vehicle["role"], metrics)

            factory = (
                self.adapter_factory
                if platform_code == PLATFORM_CODE
                else self.adapter_factories[platform_code]
            )
            adapter = factory(
                storage_state,
                concurrency=concurrency,
                headless=self.settings.auth_browser_headless,
                timeout_seconds=90,
                batch_timeout_seconds=remaining,
                evidence_policy=evidence_policy,
                prefer_http_first=False,
                include_review_article_count=True,
                include_negative_rate=True,
            )
            group_targets = [targets[index] for index in indexes]
            try:
                first = adapter.validate_sync(
                    group_targets,
                    root / platform_code / f"attempt-1-{uuid7()}",
                    on_result=lambda index, _target, result: (
                        None
                        if isinstance(result, ReputationAdapterError) and result.retryable
                        else persist(indexes[index], result, 1)
                    ),
                )
                retry_indexes = [
                    index
                    for index, result in enumerate(first)
                    if isinstance(result, ReputationAdapterError) and result.retryable
                ]
                remaining = int(deadline - monotonic())
                if retry_indexes and remaining > 0:
                    adapter.batch_timeout_seconds = remaining
                    retried = adapter.validate_sync(
                        [group_targets[index] for index in retry_indexes],
                        root / platform_code / f"attempt-2-{uuid7()}",
                        on_result=lambda index, _target, result: persist(
                            indexes[retry_indexes[index]], result, 2
                        ),
                    )
                elif retry_indexes:
                    retried = [
                        ReputationAdapterError(
                            "REPUTATION_BATCH_TIMEOUT",
                            "口碑巡检达到45分钟批次上限，未完成项已停止。",
                        )
                        for _ in retry_indexes
                    ]
                else:
                    retried = []
                retry_results = dict(zip(retry_indexes, retried, strict=True))
                group_results = [
                    retry_results.get(index, result) for index, result in enumerate(first)
                ]
                for local_index, global_index in enumerate(indexes):
                    final_results[global_index] = group_results[local_index]
                    attempt_counts[global_index] = 2 if local_index in retry_results else 1
            except Exception as error:
                for index in indexes:
                    final_results[index] = error
            finally:
                close = getattr(adapter, "close", None)
                if callable(close):
                    close()

        for index, (result, attempt_count) in enumerate(
            zip(final_results, attempt_counts, strict=True)
        ):
            if index not in persisted_indexes:
                persist(index, result, attempt_count)

        finished = datetime.now(timezone.utc)
        with self.sessions.begin() as db:
            run = db.get(ReputationRun, run_id)
            if not run:
                raise RuntimeError("口碑巡检批次在执行期间丢失")
            completed = run.completed_count
            failed = run.failed_count
            required = run.required_evidence_count
            complete_evidence = run.complete_evidence_count
            partial = int(
                db.scalar(
                    select(func.count())
                    .select_from(ReputationResult)
                    .where(
                        ReputationResult.run_id == run_id,
                        ReputationResult.status == "partial_success",
                    )
                )
                or 0
            )
            run.finished_at = finished
            if completed == 0:
                run.status = "failed"
            elif failed or partial or complete_evidence < required:
                run.status = "partial_success"
            else:
                run.status = "success"
            run.report_status = "waiting" if run.source_type == "scheduled" else "not_applicable"
            final_status = run.status
            event = db.scalar(
                select(ReputationScheduleEvent).where(ReputationScheduleEvent.run_id == run.id)
            )
            if event:
                event.status = run.status
                event.message = "正式口碑巡检已到达终态，正在生成汇报。"
        if self.event_publisher:
            self.event_publisher(
                "reputation.run.changed",
                run_id,
                status=final_status,
                completed_count=completed,
                failed_count=failed,
                required_evidence_count=required,
                complete_evidence_count=complete_evidence,
            )
        return self.get_run(run_id)

    def can_execute_official(self, run_id: str) -> bool:
        """只有同平台已经开始的普通任务释放后，正式口碑批次才取得容量。"""

        with self.sessions() as db:
            run = db.get(ReputationRun, run_id)
            if not run or run.status != "queued":
                return False
            for platform_code in run.platform_codes:
                running_tasks = db.scalar(
                    select(func.count()).select_from(CircleTask).where(
                        CircleTask.platform_code == platform_code,
                        CircleTask.status.in_(
                            ["queued", "running"] if run.source_type == "retry" else ["running"]
                        ),
                    )
                ) or 0
                running_validations = db.scalar(
                    select(func.count())
                    .select_from(ValidationJob)
                    .join(Circle, ValidationJob.circle_id == Circle.id)
                    .where(
                        Circle.platform_code == platform_code,
                        ValidationJob.status.in_(
                            ["queued", "running"] if run.source_type == "retry" else ["running"]
                        ),
                    )
                ) or 0
                if running_tasks or running_validations:
                    return False
            return True

    def generate_report(self, run_id: str, now: datetime | None = None) -> dict[str, Any]:
        """从终态冻结结果生成正文、TXT与固定版式XLSX；同一输入幂等复用。"""

        current = (now or datetime.now(timezone.utc)).astimezone(timezone.utc)
        with self.sessions() as db:
            current_run = db.get(ReputationRun, run_id)
            if not current_run:
                raise DomainError(
                    "REPUTATION_RUN_NOT_FOUND", "口碑巡检运行不存在。", status_code=404
                )
            if current_run.source_type != "scheduled":
                raise DomainError("REPUTATION_REPORT_NOT_APPLICABLE", "该运行不生成正式定时汇报。")
            if current_run.status not in {"success", "partial_success", "failed"}:
                raise DomainError(
                    "REPUTATION_REPORT_NOT_READY",
                    "巡检尚未终态，汇报继续等待。",
                    status_code=409,
                )
            failed_only = current_run.status == "failed"
            already_success = bool(
                current_run.report_status == "success"
                and current_run.report_path
                and current_run.xlsx_path
                and Path(current_run.report_path).is_file()
                and Path(current_run.xlsx_path).is_file()
            )
        if failed_only:
            with self.sessions.begin() as db:
                run = db.get(ReputationRun, run_id)
                run.report_status = "not_generated"
                run.report_generated_at = current
                event = db.scalar(
                    select(ReputationScheduleEvent).where(
                        ReputationScheduleEvent.run_id == run_id
                    )
                )
                if event:
                    event.message = "正式口碑巡检全部失败，未生成普通排名汇报。"
            return self.get_run(run_id)
        if already_success:
            return self.get_run(run_id)
        with self.sessions.begin() as db:
            run = db.get(ReputationRun, run_id)
            # 旧版本可能给尚未生成的批次保留固定汇报时点；终态触发合同
            # 生效后，首次生成时清除该等待门槛，只记录实际生成时间。
            run.report_planned_at = None
            run.report_status = "generating"
            run.report_attempt_count += 1
        with self.sessions() as db:
            run = db.get(ReputationRun, run_id)
            results = db.scalars(
                select(ReputationResult)
                .where(ReputationResult.run_id == run_id)
                .order_by(ReputationResult.role_position, ReputationResult.vehicle_position)
            ).all()
            evidence_by_result = {
                item.result_id: item
                for item in db.scalars(
                    select(ReputationEvidence).where(
                        ReputationEvidence.result_id.in_([row.id for row in results])
                    )
                ).all()
            }
            run_dir = self.settings.reputation_dir / run.id
            run_dir.mkdir(parents=True, exist_ok=True)
            report = self._render_report(run, results, set(evidence_by_result))
            report_path = run_dir / f"{run.number}.txt"
            xlsx_path = run_dir / f"{run.number}.xlsx"
            report_temp = run_dir / f".{run.number}.{uuid7()}.tmp.txt"
            xlsx_temp = run_dir / f".{run.number}.{uuid7()}.tmp.xlsx"
            try:
                report_temp.write_text(report, encoding="utf-8")
                self._create_xlsx(run, results, evidence_by_result, xlsx_temp)
                report_temp.replace(report_path)
                xlsx_temp.replace(xlsx_path)
            except Exception as error:
                report_temp.unlink(missing_ok=True)
                xlsx_temp.unlink(missing_ok=True)
                with self.sessions.begin() as update_db:
                    failed = update_db.get(ReputationRun, run_id)
                    failed.report_status = "failed"
                    failed.error_message = f"口碑派生产物生成失败：{type(error).__name__}"
                raise DomainError(
                    "REPUTATION_ARTIFACT_GENERATION_FAILED",
                    "口碑汇报或XLSX生成失败，可按冻结输入重试。",
                    status_code=503,
                ) from error
        with self.sessions.begin() as db:
            run = db.get(ReputationRun, run_id)
            run.report_text = report
            run.report_path = str(report_path)
            run.xlsx_path = str(xlsx_path)
            run.report_generated_at = current
            run.report_status = "success"
            run.error_message = None
            event = db.scalar(
                select(ReputationScheduleEvent).where(ReputationScheduleEvent.run_id == run_id)
            )
            if event:
                event.message = "正式口碑巡检已到达终态，汇报已生成。"
        if self.event_publisher:
            self.event_publisher("reputation.run.changed", run_id, status="report_success")
        return self.get_run(run_id)

    def retry_failed(self, run_id: str) -> dict[str, Any]:
        """只为关联链中仍未取得完整成功的车型平台项创建补跑批次。"""

        now = datetime.now(timezone.utc)
        with self.sessions.begin() as db:
            selected = db.get(ReputationRun, run_id)
            if not selected:
                raise DomainError(
                    "REPUTATION_RUN_NOT_FOUND", "口碑巡检运行不存在。", status_code=404
                )
            root_id = selected.root_run_id or selected.id
            root = db.get(ReputationRun, root_id)
            if not root or root.source_type != "scheduled":
                raise DomainError("REPUTATION_RETRY_FORBIDDEN", "只有正式口碑巡检关联链可以补跑。")
            if root.status not in {"partial_success", "failed"}:
                raise DomainError(
                    "REPUTATION_RETRY_NOT_NEEDED",
                    "原批次没有需要补跑的失败项。",
                    status_code=409,
                )
            chain = db.scalars(
                select(ReputationRun)
                .where(or_(ReputationRun.id == root.id, ReputationRun.root_run_id == root.id))
                .order_by(ReputationRun.created_at, ReputationRun.id)
            ).all()
            successful: set[str] = set()
            for item in chain:
                evidence_ids = {
                    value.result_id
                    for value in db.scalars(
                        select(ReputationEvidence).join(
                            ReputationResult, ReputationEvidence.result_id == ReputationResult.id
                        ).where(ReputationResult.run_id == item.id)
                    ).all()
                }
                for result in db.scalars(
                    select(ReputationResult).where(ReputationResult.run_id == item.id)
                ).all():
                    if result.status == "success" and (
                        not result.evidence_required or result.id in evidence_ids
                    ):
                        successful.add(f"{result.vehicle_id}|{result.platform_code}")
            pending = sorted(set(root.target_keys) - successful)
            if not pending:
                raise DomainError(
                    "REPUTATION_RETRY_NOT_NEEDED",
                    "关联链失败项已经全部补跑成功。",
                    status_code=409,
                )
            input_hash = _text_hash({"root": root.id, "targets": pending})
            existing = db.scalar(
                select(ReputationRun).where(
                    ReputationRun.source_type == "retry",
                    ReputationRun.root_run_id == root.id,
                    ReputationRun.input_hash == input_hash,
                    ReputationRun.status.in_(["queued", "running"]),
                )
            )
            if existing:
                return self.get_run(existing.id)
            retry_id = uuid7()
            retry = ReputationRun(
                id=retry_id,
                number=f"RP-R-{now:%Y%m%d-%H%M%S}-{retry_id[-4:].upper()}",
                source_type="retry",
                run_type=root.run_type,
                schedule_type=root.schedule_type,
                planned_date=root.planned_date,
                target_keys=pending,
                root_run_id=root.id,
                parent_run_id=selected.id,
                scope_version_id=root.scope_version_id,
                input_hash=input_hash,
                planned_at=root.planned_at,
                report_planned_at=root.report_planned_at,
                delayed=root.delayed,
                concurrency=REPUTATION_RUN_CONCURRENCY,
                baseline_date=root.baseline_date,
                baseline_frozen_at=root.baseline_frozen_at,
                baseline_source_run_id=root.baseline_source_run_id,
                baseline_snapshot=root.baseline_snapshot,
                status="queued",
                platform_codes=root.platform_codes,
                planned_count=len(pending),
                report_status="not_applicable",
                created_at=now,
            )
            db.add(retry)
        if self.event_publisher:
            self.event_publisher("reputation.run.changed", retry_id, status="queued")
        return self.get_run(retry_id)

    def schedule_status(self) -> dict[str, Any]:
        """返回固定日程及最近事件，供前端只读展示。"""

        with self.sessions() as db:
            event = db.scalar(
                select(ReputationScheduleEvent)
                .order_by(ReputationScheduleEvent.planned_at.desc())
                .limit(1)
            )
            return {
                "timezone": self.settings.timezone,
                "inspection_time": INSPECTION_TIME_TEXT,
                "report_time": None,
                "last_event": {
                    "planned_date": event.planned_date,
                    "run_type": event.run_type,
                    "status": event.status,
                    "message": event.message,
                    "run_id": event.run_id,
                    "planned_at": event.planned_at.isoformat(),
                }
                if event
                else None,
            }

    def delete_official(self, run_id: str) -> dict[str, Any]:
        """按关联链执行同盘隔离、数据库提交和墓碑写入。"""

        now = datetime.now(timezone.utc)
        with self.sessions() as db:
            selected = db.get(ReputationRun, run_id)
            if not selected:
                raise DomainError(
                    "REPUTATION_RUN_NOT_FOUND", "口碑巡检运行不存在。", status_code=404
                )
            root_id = selected.root_run_id or selected.id
            root_run = db.get(ReputationRun, root_id)
            if not root_run or root_run.source_type != "scheduled":
                raise DomainError(
                    "REPUTATION_DELETE_FORBIDDEN", "只有终态每日正式巡检批次可以整体删除。"
                )
            if root_run.status not in {"success", "partial_success", "failed"}:
                raise DomainError(
                    "REPUTATION_DELETE_NOT_TERMINAL", "巡检尚未终态，暂不能删除。", status_code=409
                )
            job_key = f"delete:{root_id}"
            existing = db.scalar(
                select(ReputationDeleteJob).where(ReputationDeleteJob.idempotency_key == job_key)
            )
            if existing:
                return self._delete_job_dict(existing)
            chain = db.scalars(
                select(ReputationRun).where(
                    or_(ReputationRun.id == root_id, ReputationRun.root_run_id == root_id)
                )
            ).all()
            result_rows = db.scalars(
                select(ReputationResult).where(
                    ReputationResult.run_id.in_([item.id for item in chain])
                )
            ).all()
            result_hash = _text_hash(
                [
                    {
                        "id": item.id,
                        "run_id": item.run_id,
                        "vehicle_id": item.vehicle_id,
                        "platform_code": item.platform_code,
                        "status": item.status,
                        "metrics": item.metrics,
                    }
                    for item in sorted(result_rows, key=lambda value: value.id)
                ]
            )
            chain_ids = [item.id for item in chain]
            schedule_type = root_run.schedule_type or self._schedule_type(
                date.fromisoformat(root_run.planned_date)
            )
            planned_date = root_run.planned_date
            idempotency_key = root_run.idempotency_key or (
                f"reputation:{planned_date}:{schedule_type}"
            )

        storage_root = self.settings.reputation_dir.resolve()
        files: list[Path] = []
        for chain_id in chain_ids:
            directory = (self.settings.reputation_dir / chain_id).resolve()
            if directory.is_relative_to(storage_root) and directory.is_dir():
                files.extend(path for path in directory.rglob("*") if path.is_file())
        manifest = [
            {
                "path": str(path),
                "relative_path": path.relative_to(storage_root).as_posix(),
                "sha256": _sha256(path),
                "size": path.stat().st_size,
            }
            for path in sorted(set(files))
        ]
        job_id = uuid7()
        quarantine = storage_root / ".quarantine" / job_id
        with self.sessions.begin() as db:
            db.add(
                ReputationDeleteJob(
                    id=job_id,
                    root_run_id=root_id,
                    idempotency_key=job_key,
                    status="deleting",
                    manifest=manifest,
                    quarantine_path=str(quarantine),
                    created_at=now,
                    updated_at=now,
                )
            )

        moved: list[tuple[Path, Path]] = []
        try:
            for item in manifest:
                source = Path(item["path"]).resolve()
                if not source.is_relative_to(storage_root) or _sha256(source) != item["sha256"]:
                    raise RuntimeError(f"删除清单校验失败：{item['relative_path']}")
                destination = quarantine / item["relative_path"]
                destination.parent.mkdir(parents=True, exist_ok=True)
                source.replace(destination)
                moved.append((source, destination))
        except Exception as error:
            for source, destination in reversed(moved):
                source.parent.mkdir(parents=True, exist_ok=True)
                if destination.exists() and not source.exists():
                    destination.replace(source)
            with self.sessions.begin() as db:
                job = db.get(ReputationDeleteJob, job_id)
                if job:
                    job.status = "delete_failed"
                    job.error_message = str(error)
                    job.updated_at = datetime.now(timezone.utc)
            return self.get_delete_job(job_id)

        try:
            with self.sessions.begin() as db:
                db.execute(delete(ReputationRun).where(ReputationRun.id.in_(chain_ids)))
                db.add(
                    ReputationTombstone(
                        planned_date=planned_date,
                        run_type=schedule_type,
                        idempotency_key=idempotency_key,
                        original_run_id=root_id,
                        result_hash=result_hash,
                        deleted_at=datetime.now(timezone.utc),
                    )
                )
                event = db.scalar(
                    select(ReputationScheduleEvent).where(
                        ReputationScheduleEvent.planned_date == planned_date,
                        ReputationScheduleEvent.run_type == schedule_type,
                    )
                )
                if event:
                    event.status = "deleted"
                    event.message = "正式口碑巡检关联链已删除，日期幂等身份由墓碑保留。"
                    event.run_id = None
                job = db.get(ReputationDeleteJob, job_id)
                if job:
                    job.status = "storage_cleanup_pending"
                    job.updated_at = datetime.now(timezone.utc)
        except Exception as error:
            for source, destination in reversed(moved):
                source.parent.mkdir(parents=True, exist_ok=True)
                if destination.exists() and not source.exists():
                    destination.replace(source)
            with self.sessions.begin() as db:
                job = db.get(ReputationDeleteJob, job_id)
                if job:
                    job.status = "delete_failed"
                    job.error_message = str(error)
                    job.updated_at = datetime.now(timezone.utc)
            return self.get_delete_job(job_id)

        cleanup_error = None
        try:
            if quarantine.exists():
                shutil.rmtree(quarantine)
            for chain_id in chain_ids:
                shutil.rmtree(storage_root / chain_id, ignore_errors=True)
        except Exception as error:
            cleanup_error = str(error)
        with self.sessions.begin() as db:
            job = db.get(ReputationDeleteJob, job_id)
            if job:
                job.status = "storage_cleanup_pending" if cleanup_error else "success"
                job.error_message = cleanup_error
                job.updated_at = datetime.now(timezone.utc)
                job.completed_at = None if cleanup_error else datetime.now(timezone.utc)
        return self.get_delete_job(job_id)

    def get_delete_job(self, job_id: str) -> dict[str, Any]:
        with self.sessions() as db:
            job = db.get(ReputationDeleteJob, job_id)
            if not job:
                raise DomainError(
                    "REPUTATION_DELETE_JOB_NOT_FOUND", "删除作业不存在。", status_code=404
                )
            return self._delete_job_dict(job)

    def retry_delete_cleanup(self, job_id: str) -> dict[str, Any]:
        """重试数据库提交后尚未完成的隔离区清理。"""

        with self.sessions() as db:
            job = db.get(ReputationDeleteJob, job_id)
            if not job:
                raise DomainError(
                    "REPUTATION_DELETE_JOB_NOT_FOUND", "删除作业不存在。", status_code=404
                )
            if job.status == "success":
                return self._delete_job_dict(job)
            if job.status != "storage_cleanup_pending":
                raise DomainError(
                    "REPUTATION_DELETE_RETRY_NOT_READY",
                    "该删除作业不处于存储清理待重试状态。",
                    status_code=409,
                )
            quarantine = Path(job.quarantine_path)
        try:
            if quarantine.exists():
                shutil.rmtree(quarantine)
        except Exception as error:
            with self.sessions.begin() as db:
                job = db.get(ReputationDeleteJob, job_id)
                job.error_message = str(error)
                job.updated_at = datetime.now(timezone.utc)
            return self.get_delete_job(job_id)
        with self.sessions.begin() as db:
            job = db.get(ReputationDeleteJob, job_id)
            job.status = "success"
            job.error_message = None
            job.updated_at = datetime.now(timezone.utc)
            job.completed_at = datetime.now(timezone.utc)
        return self.get_delete_job(job_id)

    @staticmethod
    def _delete_job_dict(job: ReputationDeleteJob) -> dict[str, Any]:
        return {
            "id": job.id,
            "root_run_id": job.root_run_id,
            "status": job.status,
            "file_count": len(job.manifest),
            "error_message": job.error_message,
            "created_at": job.created_at.isoformat(),
            "updated_at": job.updated_at.isoformat(),
            "completed_at": job.completed_at.isoformat() if job.completed_at else None,
        }

    def delete_synthetic(self, run_id: str) -> dict[str, Any]:
        self._require_synthetic()
        with self.sessions.begin() as db:
            run = db.get(ReputationRun, run_id)
            if not run:
                raise DomainError(
                    "REPUTATION_RUN_NOT_FOUND", "口碑巡检运行不存在。", status_code=404
                )
            if run.source_type != "synthetic":
                raise DomainError(
                    "REPUTATION_DELETE_FORBIDDEN",
                    "这里只能删除隔离合成运行。",
                    status_code=403,
                )
            paths = [run.report_path, run.xlsx_path, run.evidence_zip_path]
            db.delete(run)
        run_dir = self.settings.reputation_dir / run_id
        shutil.rmtree(run_dir, ignore_errors=True)
        for raw in paths:
            if raw:
                Path(raw).unlink(missing_ok=True)
        return {"deleted": True}

    def compact_region_evidence(self) -> dict[str, int]:
        """把历史双截图证据收敛为单张指标区域图，并移除未再引用的长截图。"""

        obsolete: set[Path] = set()
        zip_paths: set[Path] = set()
        attempt_count = 0
        evidence_count = 0
        with self.sessions.begin() as db:
            attempts = db.scalars(
                select(ReputationMappingValidationAttempt).where(
                    ReputationMappingValidationAttempt.status == "success"
                )
            ).all()
            for item in attempts:
                region = Path(item.metric_region_path) if item.metric_region_path else None
                if (
                    not region
                    or not item.metric_region_sha256
                    or not region.is_file()
                    or _sha256(region) != item.metric_region_sha256
                ):
                    continue
                old = Path(item.full_page_path) if item.full_page_path else None
                if old and old != region:
                    obsolete.add(old)
                item.full_page_path = str(region)
                item.full_page_sha256 = item.metric_region_sha256
                attempt_count += 1
            evidence_rows = db.scalars(select(ReputationEvidence)).all()
            for item in evidence_rows:
                region = Path(item.metric_region_path)
                if not region.is_file() or _sha256(region) != item.metric_region_sha256:
                    continue
                old = Path(item.full_page_path)
                if old != region:
                    obsolete.add(old)
                item.full_page_path = str(region)
                item.full_page_sha256 = item.metric_region_sha256
                with Image.open(region) as source:
                    item.width, item.height = source.size
                evidence_count += 1
            for run in db.scalars(select(ReputationRun)).all():
                if run.evidence_zip_path:
                    zip_paths.add(Path(run.evidence_zip_path))
                    run.evidence_zip_path = None
        root = self.settings.reputation_dir.resolve()
        referenced = {
            Path(raw).resolve()
            for raw in [
                *[item.metric_region_path for item in attempts if item.metric_region_path],
                *[item.metric_region_path for item in evidence_rows],
            ]
        }
        removed = 0
        for path in obsolete | zip_paths:
            resolved = path.resolve()
            if resolved.is_relative_to(root) and resolved not in referenced and resolved.is_file():
                resolved.unlink()
                removed += 1
        return {
            "validation_attempts": attempt_count,
            "run_evidence": evidence_count,
            "removed_files": removed,
        }

    def get_file(self, run_id: str, kind: str) -> Path:
        with self.sessions() as db:
            run = db.get(ReputationRun, run_id)
            if not run:
                raise DomainError(
                    "REPUTATION_RUN_NOT_FOUND", "口碑巡检运行不存在。", status_code=404
                )
            raw = run.report_path if kind == "txt" else run.xlsx_path
            if not raw or not Path(raw).is_file():
                raise DomainError(
                    "REPUTATION_ARTIFACT_NOT_FOUND",
                    "口碑交付文件尚未生成。",
                    status_code=404,
                )
            return Path(raw)

    def get_evidence_file(self, evidence_id: str, kind: str) -> Path:
        with self.sessions() as db:
            evidence = db.get(ReputationEvidence, evidence_id)
            if not evidence:
                raise DomainError(
                    "REPUTATION_EVIDENCE_NOT_FOUND",
                    "口碑页面证据不存在。",
                    status_code=404,
                )
            raw = evidence.full_page_path if kind == "full" else evidence.metric_region_path
            path = Path(raw)
            if not path.is_file():
                raise DomainError(
                    "REPUTATION_EVIDENCE_FILE_MISSING",
                    "口碑页面证据文件缺失。",
                    status_code=404,
                )
            return path

    def evidence_zip(self, run_id: str) -> Path:
        with self.sessions.begin() as db:
            run = db.get(ReputationRun, run_id)
            if not run:
                raise DomainError(
                    "REPUTATION_RUN_NOT_FOUND", "口碑巡检运行不存在。", status_code=404
                )
            if run.evidence_zip_path and Path(run.evidence_zip_path).is_file():
                return Path(run.evidence_zip_path)
            results = db.scalars(
                select(ReputationResult)
                .where(ReputationResult.run_id == run_id)
                .order_by(
                    ReputationResult.role_position,
                    ReputationResult.vehicle_position,
                    ReputationResult.platform_code,
                )
            ).all()
            result_by_id = {row.id: row for row in results}
            evidence = db.scalars(
                select(ReputationEvidence).where(
                    ReputationEvidence.result_id.in_(list(result_by_id))
                )
            ).all()
            evidence_by_result = {item.result_id: item for item in evidence}
            run_dir = self.settings.reputation_dir / run_id
            zip_path = run_dir / f"{run.number}-evidence.zip"
            zip_temp = run_dir / f".{run.number}-{uuid7()}.tmp.zip"
            manifest: list[dict[str, Any]] = []
            checksums: list[str] = []
            try:
                with zipfile.ZipFile(zip_temp, "w", compression=zipfile.ZIP_STORED) as archive:
                    for result in [item for item in results if item.evidence_required]:
                        item = evidence_by_result.get(result.id)
                        name = f"{result.platform_code}/{result.vehicle_id}/region.png"
                        if item and Path(item.metric_region_path).is_file():
                            archive.write(item.metric_region_path, name)
                            checksums.append(f"{item.metric_region_sha256}  {name}")
                            manifest.append(
                                {
                                    "status": "complete",
                                    "evidence_id": item.id,
                                    "result_id": result.id,
                                    "vehicle_id": result.vehicle_id,
                                    "platform_code": result.platform_code,
                                    "region_path": name,
                                    "region_sha256": item.metric_region_sha256,
                                }
                            )
                        else:
                            manifest.append(
                                {
                                    "status": "missing",
                                    "result_id": result.id,
                                    "vehicle_id": result.vehicle_id,
                                    "platform_code": result.platform_code,
                                    "reason": result.error_message or "必需指标区域证据缺失。",
                                }
                            )
                    archive.writestr(
                        "manifest.json",
                        json.dumps(
                            {"schema_version": "reputation-evidence-region-v1", "items": manifest},
                            ensure_ascii=False,
                            indent=2,
                        ),
                    )
                    archive.writestr("SHA256SUMS", "\n".join(checksums) + "\n")
                zip_temp.replace(zip_path)
            except Exception:
                zip_temp.unlink(missing_ok=True)
                raise
            run.evidence_zip_path = str(zip_path)
            return zip_path

    def get_scope(self) -> dict[str, Any]:
        with self.sessions() as db:
            draft = db.get(ReputationScopeDraft, "current")
            if not draft:
                return {
                    "initialized": False,
                    "revision": 0,
                    "vehicles": [],
                    "published_version": None,
                    "message": "尚未通过UTF-8 CSV初始化27款车型。",
                }
            data = draft.data or {}
            referenced_ids = self._scope_reference_ids(db)
            vehicles = json.loads(json.dumps(data.get("vehicles", []), ensure_ascii=False))
            for vehicle in vehicles:
                vehicle["project_group"] = str(
                    vehicle.get("project_group") or DEFAULT_PROJECT_GROUP
                )
                vehicle["removal_mode"] = (
                    "disable" if vehicle.get("id") in referenced_ids else "delete"
                )
            return {
                "initialized": True,
                "revision": draft.revision,
                "vehicles": vehicles,
                "published_version": self._published_version_dict(db, draft),
                "source_sha256": draft.source_sha256,
                "updated_at": draft.updated_at.isoformat(),
            }

    @staticmethod
    def _scope_reference_ids(db: Session) -> set[str]:
        """返回已进入不可变版本或正式结果的车型身份。"""

        referenced: set[str] = set(
            db.scalars(select(ReputationResult.vehicle_id).distinct()).all()
        )
        for version in db.scalars(select(ReputationScopeVersion)).all():
            for vehicle in (version.snapshot or {}).get("vehicles", []):
                vehicle_id = str(vehicle.get("id") or "")
                if vehicle_id:
                    referenced.add(vehicle_id)
        return referenced

    @staticmethod
    def _normalize_role_orders(vehicles: list[dict[str, Any]], role: str) -> None:
        """按车型映射显示顺序重排角色序号，停用车型接在启用车型之后。"""

        grouped = [
            vehicle
            for enabled in (True, False)
            for vehicle in vehicles
            if vehicle.get("role") == role
            and bool(vehicle.get("enabled", True)) is enabled
        ]
        for position, vehicle in enumerate(grouped, start=1):
            vehicle["role_order"] = position

    @staticmethod
    def _require_scope_revision(
        draft: ReputationScopeDraft | None, revision: int
    ) -> ReputationScopeDraft:
        if not draft:
            raise DomainError("REPUTATION_SCOPE_UNINITIALIZED", "请先初始化口碑车型范围。")
        if draft.revision != revision:
            raise DomainError(
                "REPUTATION_SCOPE_CONFLICT",
                "范围草稿已经变化，请刷新后重试。",
                status_code=409,
            )
        return draft

    @classmethod
    def _scope_vehicle_fields(cls, value: ScopeVehicleCreateRequest) -> dict[str, str]:
        """校验并规范化新增与修改对话框共用的车型字段。"""

        spec = cls._platform_spec(value.platform_code)
        fields = {
            "series_name": value.series_name.strip(),
            "vehicle_name": value.vehicle_name.strip(),
            "project_group": value.project_group.strip(),
            "platform_vehicle_id": value.platform_vehicle_id.strip(),
            "platform_display_name": value.platform_display_name.strip(),
            "platform_code": spec.code,
        }
        if not all(fields.values()):
            raise DomainError(
                "REPUTATION_SCOPE_VEHICLE_REQUIRED",
                "车型名称、项目组归属和平台映射不能为空。",
            )
        if len(fields["project_group"]) > 80:
            raise DomainError(
                "REPUTATION_SCOPE_PROJECT_GROUP_TOO_LONG", "项目组归属不能超过80个字符。"
            )
        try:
            fields["platform_url"] = spec.normalize_url(
                value.platform_url, fields["platform_vehicle_id"]
            )
        except ReputationAdapterError as error:
            raise DomainError(error.code, error.message) from error
        return fields

    def create_scope_vehicle(self, value: ScopeVehicleCreateRequest) -> dict[str, Any]:
        """新增一条独立车型身份；内部 ID 只由服务端生成且永不复用。"""

        fields = self._scope_vehicle_fields(value)
        with self.sessions.begin() as db:
            draft = self._require_scope_revision(
                db.get(ReputationScopeDraft, "current"), value.revision
            )
            data = json.loads(json.dumps(draft.data, ensure_ascii=False))
            vehicles = data.setdefault("vehicles", [])
            if any(
                str(vehicle.get("mappings", {}).get(fields["platform_code"], {}).get("platform_vehicle_id"))
                == fields["platform_vehicle_id"]
                for vehicle in vehicles
            ):
                raise DomainError("REPUTATION_SCOPE_DUPLICATE", "平台车型 ID 已被现有车型使用。")
            role_order = 1 + max(
                (
                    int(vehicle.get("role_order") or 0)
                    for vehicle in vehicles
                    if vehicle.get("role") == value.role
                    and bool(vehicle.get("enabled", True))
                ),
                default=0,
            )
            vehicles.append(
                {
                    "id": f"rep-{uuid7()}",
                    "series_name": fields["series_name"],
                    "vehicle_name": fields["vehicle_name"],
                    "project_group": fields["project_group"],
                    "role": value.role,
                    "role_order": role_order,
                    "enabled": True,
                    "mappings": {
                        fields["platform_code"]: {
                            "platform_vehicle_id": fields["platform_vehicle_id"],
                            "platform_url": fields["platform_url"],
                            "platform_display_name": fields["platform_display_name"],
                            "validation_status": "unverified",
                        }
                    },
                }
            )
            self._normalize_role_orders(vehicles, value.role)
            draft.data = data
            draft.revision += 1
        return self.get_scope()

    def update_scope_vehicle(
        self, vehicle_id: str, value: ScopeVehicleUpdateRequest
    ) -> dict[str, Any]:
        """修改当前范围草稿中的车型信息，映射变化时撤销原验证绑定。"""

        fields = self._scope_vehicle_fields(value)
        mapping_changed = False
        changed = False
        with self.sessions.begin() as db:
            draft = self._require_scope_revision(
                db.get(ReputationScopeDraft, "current"), value.revision
            )
            data = json.loads(json.dumps(draft.data, ensure_ascii=False))
            vehicles = data.get("vehicles", [])
            vehicle = next((row for row in vehicles if row.get("id") == vehicle_id), None)
            if not vehicle:
                raise DomainError(
                    "REPUTATION_SCOPE_VEHICLE_NOT_FOUND", "车型不存在。", status_code=404
                )
            if any(
                row.get("id") != vehicle_id
                and str(
                    row.get("mappings", {})
                    .get(fields["platform_code"], {})
                    .get("platform_vehicle_id")
                )
                == fields["platform_vehicle_id"]
                for row in vehicles
            ):
                raise DomainError("REPUTATION_SCOPE_DUPLICATE", "平台车型 ID 已被现有车型使用。")

            mapping = vehicle.setdefault("mappings", {}).get(fields["platform_code"]) or {}
            mapping_changed = any(
                str(mapping.get(key) or "") != fields[key]
                for key in (
                    "platform_vehicle_id",
                    "platform_url",
                    "platform_display_name",
                )
            )
            old_role = str(vehicle.get("role") or "")
            changed = mapping_changed or old_role != value.role or any(
                str(vehicle.get(key) or "") != fields[key]
                for key in ("series_name", "vehicle_name", "project_group")
            )
            if changed:
                vehicle.update(
                    {
                        "series_name": fields["series_name"],
                        "vehicle_name": fields["vehicle_name"],
                        "project_group": fields["project_group"],
                        "role": value.role,
                    }
                )
                if mapping_changed:
                    vehicle["mappings"][fields["platform_code"]] = {
                        "platform_vehicle_id": fields["platform_vehicle_id"],
                        "platform_url": fields["platform_url"],
                        "platform_display_name": fields["platform_display_name"],
                        "validation_status": "unverified",
                    }
                if old_role != value.role:
                    self._normalize_role_orders(vehicles, old_role)
                    self._normalize_role_orders(vehicles, value.role)
                draft.data = data
                draft.revision += 1
        scope = self.get_scope()
        scope["last_vehicle_action"] = "updated" if changed else "unchanged"
        scope["last_vehicle_mapping_changed"] = mapping_changed
        return scope

    def remove_scope_vehicle(self, vehicle_id: str, revision: int) -> dict[str, Any]:
        """未引用车型永久删除；已有历史引用的车型只停用。"""

        action = "deleted"
        with self.sessions.begin() as db:
            draft = self._require_scope_revision(
                db.get(ReputationScopeDraft, "current"), revision
            )
            data = json.loads(json.dumps(draft.data, ensure_ascii=False))
            vehicles = data.get("vehicles", [])
            vehicle = next((row for row in vehicles if row.get("id") == vehicle_id), None)
            if not vehicle:
                raise DomainError(
                    "REPUTATION_SCOPE_VEHICLE_NOT_FOUND", "车型不存在。", status_code=404
                )
            role = str(vehicle.get("role") or "")
            if vehicle_id in self._scope_reference_ids(db):
                if not vehicle.get("enabled", True):
                    raise DomainError("REPUTATION_SCOPE_VEHICLE_DISABLED", "车型已经停用。")
                vehicle["enabled"] = False
                action = "disabled"
            else:
                vehicles.remove(vehicle)
            self._normalize_role_orders(vehicles, role)
            draft.data = data
            draft.revision += 1
        scope = self.get_scope()
        scope["last_vehicle_action"] = action
        return scope

    def restore_scope_vehicle(
        self, vehicle_id: str, value: ScopeVehicleRevisionRequest
    ) -> dict[str, Any]:
        """恢复历史车型，并放到对应角色当前启用列表末尾。"""

        with self.sessions.begin() as db:
            draft = self._require_scope_revision(
                db.get(ReputationScopeDraft, "current"), value.revision
            )
            data = json.loads(json.dumps(draft.data, ensure_ascii=False))
            vehicles = data.get("vehicles", [])
            vehicle = next((row for row in vehicles if row.get("id") == vehicle_id), None)
            if not vehicle:
                raise DomainError(
                    "REPUTATION_SCOPE_VEHICLE_NOT_FOUND", "车型不存在。", status_code=404
                )
            if vehicle.get("enabled", True):
                raise DomainError("REPUTATION_SCOPE_VEHICLE_ENABLED", "车型当前已经启用。")
            vehicle["enabled"] = True
            vehicle["role_order"] = 1 + max(
                (
                    int(row.get("role_order") or 0)
                    for row in vehicles
                    if row is not vehicle
                    and row.get("role") == vehicle.get("role")
                    and bool(row.get("enabled", True))
                ),
                default=0,
            )
            self._normalize_role_orders(vehicles, str(vehicle.get("role") or ""))
            draft.data = data
            draft.revision += 1
        return self.get_scope()

    def initialize_scope_csv(self, path: Path) -> dict[str, Any]:
        raw = path.read_bytes()
        source_hash = hashlib.sha256(raw).hexdigest()
        rows = list(csv.DictReader(io.StringIO(raw.decode("utf-8-sig"), newline="")))
        expected = [
            "schema_version",
            "seed_key",
            "series_name",
            "vehicle_name",
            "role",
            "role_order",
            "platform_code",
            "platform_vehicle_id",
            "platform_url",
            "platform_display_name",
        ]
        if not rows or list(rows[0]) != expected:
            raise DomainError("REPUTATION_SCOPE_CSV_HEADER", "初始化CSV表头与固定Schema不一致。")
        if len(rows) != 27:
            raise DomainError("REPUTATION_SCOPE_COUNT", "初始化CSV必须恰好包含27款车型。")
        focus = [row for row in rows if row["role"] == "focus"]
        competitor = [row for row in rows if row["role"] == "competitor"]
        if len(focus) != 14 or len(competitor) != 13:
            raise DomainError("REPUTATION_SCOPE_ROLE_COUNT", "初始化CSV必须为14款重点和13款竞品。")
        if len({row["seed_key"] for row in rows}) != 27:
            raise DomainError("REPUTATION_SCOPE_DUPLICATE", "初始化键必须全局唯一。")
        vehicles = []
        for row in rows:
            if row["platform_code"] != PLATFORM_CODE:
                raise DomainError("REPUTATION_SCOPE_PLATFORM", "当前初始化只接受已接入平台映射。")
            try:
                normalized_url = REPUTATION_PLATFORMS[PLATFORM_CODE].normalize_url(
                    row["platform_url"], row["platform_vehicle_id"]
                )
            except ReputationAdapterError as error:
                raise DomainError(error.code, error.message) from error
            try:
                role_order = int(row["role_order"])
            except ValueError as error:
                raise DomainError("REPUTATION_SCOPE_ORDER", "角色内顺序必须是整数。") from error
            vehicles.append(
                {
                    "id": row["seed_key"],
                    "series_name": row["series_name"],
                    "vehicle_name": row["vehicle_name"],
                    "project_group": DEFAULT_PROJECT_GROUP,
                    "role": row["role"],
                    "role_order": role_order,
                    "enabled": True,
                    "mappings": {
                        PLATFORM_CODE: {
                            "platform_vehicle_id": row["platform_vehicle_id"],
                            "platform_url": normalized_url,
                            "platform_display_name": row["platform_display_name"],
                            "validation_status": "unverified",
                        }
                    },
                }
            )
        for group, expected_count in ((focus, 14), (competitor, 13)):
            positions = sorted(int(row["role_order"]) for row in group)
            if positions != list(range(1, expected_count + 1)):
                raise DomainError("REPUTATION_SCOPE_ORDER", "角色内顺序必须从1连续编号。")
        with self.sessions.begin() as db:
            existing = db.get(ReputationScopeDraft, "current")
            if existing:
                if existing.source_sha256 == source_hash:
                    return self.get_scope()
                raise DomainError("REPUTATION_SCOPE_EXISTS", "口碑范围已经初始化，不能覆盖导入。")
            db.add(
                ReputationScopeDraft(
                    id="current",
                    revision=1,
                    data={"schema_version": rows[0]["schema_version"], "vehicles": vehicles},
                    source_sha256=source_hash,
                )
            )
        return self.get_scope()

    def preview_mappings(self, value: MappingPasteRequest) -> dict[str, Any]:
        scope = self.get_scope()
        if not scope["initialized"]:
            raise DomainError("REPUTATION_SCOPE_UNINITIALIZED", "请先初始化口碑车型范围。")
        if value.revision != scope["revision"]:
            raise DomainError(
                "REPUTATION_SCOPE_CONFLICT",
                "范围草稿已经变化，请刷新后重试。",
                status_code=409,
            )
        spec = self._platform_spec(getattr(value, "platform_code", PLATFORM_CODE))
        vehicle_ids = {row["id"] for row in scope["vehicles"]}
        errors: list[dict[str, str]] = []
        seen_vehicle: set[str] = set()
        seen_platform: set[str] = set()
        for index, row in enumerate(value.rows, start=1):
            if row.vehicle_id not in vehicle_ids:
                errors.append({"row": str(index), "reason": "内部车型ID不存在"})
            if row.vehicle_id in seen_vehicle:
                errors.append({"row": str(index), "reason": "内部车型ID重复"})
            if row.platform_vehicle_id in seen_platform:
                errors.append({"row": str(index), "reason": "平台车型ID重复"})
            try:
                spec.normalize_url(row.platform_url, row.platform_vehicle_id)
            except ReputationAdapterError as error:
                errors.append({"row": str(index), "reason": error.message})
            if not row.platform_vehicle_id.strip():
                errors.append({"row": str(index), "reason": "平台车型ID不能为空"})
            if not row.platform_display_name.strip():
                errors.append({"row": str(index), "reason": "平台展示名不能为空"})
            seen_vehicle.add(row.vehicle_id)
            seen_platform.add(row.platform_vehicle_id)
        payload = value.model_dump()
        return {
            "valid": not errors,
            "errors": errors,
            "changed_count": len(value.rows),
            "unchanged_count": len(vehicle_ids - seen_vehicle),
            "input_hash": _text_hash(payload),
            "revision": value.revision,
        }

    def save_mappings(self, value: MappingPasteRequest) -> dict[str, Any]:
        preview = self.preview_mappings(value)
        if not preview["valid"]:
            raise DomainError(
                "REPUTATION_MAPPING_INVALID",
                "批量映射存在错误，本次没有保存任何行。",
                details=preview["errors"],
            )
        with self.sessions.begin() as db:
            draft = db.get(ReputationScopeDraft, "current")
            if not draft or draft.revision != value.revision:
                raise DomainError(
                    "REPUTATION_SCOPE_CONFLICT",
                    "范围草稿已经变化，请刷新后重试。",
                    status_code=409,
                )
            data = json.loads(json.dumps(draft.data, ensure_ascii=False))
            by_id = {row["id"]: row for row in data["vehicles"]}
            for row in value.rows:
                by_id[row.vehicle_id].setdefault("mappings", {})[value.platform_code] = {
                    "platform_vehicle_id": row.platform_vehicle_id,
                    "platform_url": self._platform_spec(value.platform_code).normalize_url(
                        row.platform_url, row.platform_vehicle_id
                    ),
                    "platform_display_name": row.platform_display_name,
                    "validation_status": "unverified",
                }
            draft.data = data
            draft.revision += 1
        return self.get_scope()

    def validate_mappings(self, value: MappingValidationRequest) -> dict[str, Any]:
        """并发验证当前草稿映射，并把单次三门禁结果绑定回草稿。"""

        spec = self._platform_spec(getattr(value, "platform_code", PLATFORM_CODE))
        scope = self.get_scope()
        if not scope["initialized"]:
            raise DomainError("REPUTATION_SCOPE_UNINITIALIZED", "请先初始化口碑车型范围。")
        if value.revision != scope["revision"]:
            raise DomainError(
                "REPUTATION_SCOPE_CONFLICT",
                "范围草稿已经变化，请刷新后重试。",
                status_code=409,
            )
        requested_ids = value.vehicle_ids or [
            row["id"] for row in scope["vehicles"] if row.get("enabled", True)
        ]
        if not requested_ids or len(set(requested_ids)) != len(requested_ids):
            raise DomainError("REPUTATION_VALIDATION_TARGETS", "验证车型不能为空或重复。")
        by_id = {row["id"]: row for row in scope["vehicles"]}
        unknown = [item for item in requested_ids if item not in by_id]
        if unknown:
            raise DomainError(
                "REPUTATION_VALIDATION_TARGETS",
                "验证请求包含不存在的内部车型ID。",
                details=[{"vehicle_id": item, "reason": "内部车型ID不存在"} for item in unknown],
            )
        disabled = [item for item in requested_ids if not by_id[item].get("enabled", True)]
        if disabled:
            raise DomainError(
                "REPUTATION_VALIDATION_TARGETS",
                "停用车型不参与映射验证。",
                details=[{"vehicle_id": item, "reason": "车型已停用"} for item in disabled],
            )
        if not self.session_store:
            raise DomainError(
                "REPUTATION_SESSION_STORE_MISSING", "真实口碑适配器未连接共享Session。"
            )
        storage_state = self.session_store.get_state(spec.code)
        if not storage_state:
            raise DomainError(
                "AUTH_REQUIRED", f"请先在平台配置完成{spec.display_name}认证。", status_code=409
            )
        targets: list[ReputationMappingTarget] = []
        for vehicle_id in requested_ids:
            mapping = by_id[vehicle_id].get("mappings", {}).get(spec.code)
            if not mapping:
                raise DomainError(
                    "REPUTATION_MAPPING_MISSING",
                    f"车型{vehicle_id}缺少{spec.display_name}映射。",
                )
            try:
                normalized_url = spec.normalize_url(
                    str(mapping.get("platform_url") or ""),
                    str(mapping.get("platform_vehicle_id") or ""),
                )
            except ReputationAdapterError as error:
                raise DomainError(error.code, error.message) from error
            targets.append(
                ReputationMappingTarget(
                    vehicle_id=vehicle_id,
                    platform_vehicle_id=str(mapping["platform_vehicle_id"]),
                    platform_url=normalized_url,
                    platform_display_name=str(mapping["platform_display_name"]),
                    mapping_hash=_mapping_hash(vehicle_id, mapping, spec.code),
                )
            )
        with self.sessions() as db:
            platform = db.get(PlatformConfig, spec.code)
            concurrency = max(1, min(int(platform.internal_concurrency if platform else 2), 8))
        run_id = uuid7()
        input_hash = _text_hash(
            [{"vehicle_id": item.vehicle_id, "mapping_hash": item.mapping_hash} for item in targets]
        )
        now = datetime.now(timezone.utc)
        with self.sessions.begin() as db:
            db.add(
                ReputationMappingValidationRun(
                    id=run_id,
                    platform_code=spec.code,
                    status="running",
                    input_hash=input_hash,
                    concurrency=concurrency,
                    requested_count=len(targets),
                    started_at=now,
                )
            )
        factory = self.adapter_factory if spec.code == PLATFORM_CODE else self.adapter_factories[spec.code]
        adapter = factory(
            storage_state,
            concurrency=concurrency,
            headless=self.settings.auth_browser_headless,
            timeout_seconds=90,
        )
        root = self.settings.reputation_dir / "mapping-validations" / run_id
        try:
            first = adapter.validate_sync(targets, root / "attempt-1")
            retry_indexes = [
                index
                for index, result in enumerate(first)
                if isinstance(result, ReputationAdapterError) and result.retryable
            ]
            retry_results: dict[int, ReputationPageResult | Exception] = {}
            if retry_indexes:
                retried = adapter.validate_sync(
                    [targets[index] for index in retry_indexes], root / "attempt-2"
                )
                retry_results = dict(zip(retry_indexes, retried, strict=True))
        except Exception as error:
            finished = datetime.now(timezone.utc)
            with self.sessions.begin() as db:
                failed_run = db.get(ReputationMappingValidationRun, run_id)
                if failed_run:
                    failed_run.status = "failed"
                    failed_run.failed_count = len(targets)
                    failed_run.finished_at = finished
            raise DomainError(
                "REPUTATION_VALIDATION_RUNTIME_FAILED",
                f"真实页面验证运行失败：{type(error).__name__}",
                status_code=503,
            ) from error
        finally:
            close = getattr(adapter, "close", None)
            if callable(close):
                close()
        final_results = [retry_results.get(index, result) for index, result in enumerate(first)]
        finished = datetime.now(timezone.utc)
        succeeded = sum(isinstance(item, ReputationPageResult) for item in final_results)
        with self.sessions.begin() as db:
            run = db.get(ReputationMappingValidationRun, run_id)
            if not run:
                raise RuntimeError("映射验证运行记录丢失")
            draft = db.get(ReputationScopeDraft, "current")
            if not draft:
                raise RuntimeError("口碑范围草稿丢失")
            current_data = json.loads(json.dumps(draft.data, ensure_ascii=False))
            current_by_id = {row["id"]: row for row in current_data["vehicles"]}
            for index, (target, first_result, final_result) in enumerate(
                zip(targets, first, final_results, strict=True)
            ):
                attempts = [first_result]
                if index in retry_results:
                    attempts.append(retry_results[index])
                attempt_records: list[ReputationMappingValidationAttempt] = []
                for attempt_number, result in enumerate(attempts, start=1):
                    record = self._validation_attempt_record(
                        run_id, target, attempt_number, result, finished, spec
                    )
                    attempt_records.append(record)
                    db.add(record)
                mapping = current_by_id[target.vehicle_id]["mappings"][spec.code]
                if _mapping_hash(target.vehicle_id, mapping, spec.code) != target.mapping_hash:
                    continue
                if isinstance(final_result, ReputationPageResult):
                    final_attempt_number = 2 if index in retry_results else 1
                    mapping.update(
                        {
                            "validation_status": "verified",
                            "validation_run_id": run_id,
                            "validation_attempt_number": final_attempt_number,
                            "validation_attempt_id": attempt_records[-1].id,
                            "validation_contract_version": spec.validation_contract_version,
                            "validated_mapping_hash": target.mapping_hash,
                            "validated_at": finished.isoformat(),
                            "actual_name": final_result.actual_name,
                            "latest_metrics": {
                                "score": final_result.score_raw,
                                "rank": final_result.rank_raw,
                                "volume": final_result.volume_raw,
                                "review_article_count": final_result.review_article_count_raw,
                                "negative_rate": final_result.negative_rate_raw,
                                "rank_scope": final_result.rank_scope,
                            },
                            "validation_error": None,
                        }
                    )
                elif mapping.get("validation_status") != "verified":
                    error = self._validation_error(final_result)
                    mapping.update(
                        {
                            "validation_status": "failed",
                            "validation_run_id": run_id,
                            "validation_attempt_number": len(attempts),
                            "validation_error": error["message"],
                        }
                    )
            draft.data = current_data
            draft.revision += 1
            run.status = (
                "success"
                if succeeded == len(targets)
                else "partial_success"
                if succeeded
                else "failed"
            )
            run.succeeded_count = succeeded
            run.failed_count = len(targets) - succeeded
            run.finished_at = finished
        if self.event_publisher:
            self.event_publisher("reputation.scope.changed", "current", status="validated")
        return self.get_mapping_validation(run_id)

    @staticmethod
    def _validation_error(error: Exception) -> dict[str, str]:
        if isinstance(error, ReputationAdapterError):
            return {"code": error.code, "message": error.message}
        return {
            "code": "REPUTATION_VALIDATION_INTERNAL_ERROR",
            "message": f"真实页面验证异常：{type(error).__name__}",
        }

    def _validation_attempt_record(
        self,
        run_id: str,
        target: ReputationMappingTarget,
        attempt_number: int,
        result: ReputationPageResult | Exception,
        finished: datetime,
        spec: ReputationPlatformSpec,
    ) -> ReputationMappingValidationAttempt:
        if isinstance(result, ReputationPageResult):
            return ReputationMappingValidationAttempt(
                id=uuid7(),
                run_id=run_id,
                vehicle_id=target.vehicle_id,
                attempt_number=attempt_number,
                mapping_hash=target.mapping_hash,
                contract_version=spec.validation_contract_version,
                adapter_version=spec.adapter_version,
                status="success",
                actual_name=result.actual_name,
                final_url=result.final_url,
                metrics={
                    "score": result.score_raw,
                    "rank": result.rank_raw,
                    "volume": result.volume_raw,
                    "review_article_count": result.review_article_count_raw,
                    "negative_rate": result.negative_rate_raw,
                    "rank_scope": result.rank_scope,
                },
                gate_results={
                    "identity": "passed",
                    "metrics": "passed",
                    "evidence": "passed",
                    "measurements": result.measurements,
                    "metric_rect": result.metric_rect,
                    "viewport": spec.viewport,
                    "document": {
                        "width": result.measurements[-1].get("document_width"),
                        "height": result.measurements[-1].get("document_height"),
                    },
                    "region": {"width": result.width, "height": result.height},
                },
                full_page_path=str(result.full_page_path),
                metric_region_path=str(result.metric_region_path),
                full_page_sha256=result.full_page_sha256,
                metric_region_sha256=result.metric_region_sha256,
                duration_ms=result.duration_ms,
                finished_at=finished,
            )
        error = self._validation_error(result)
        return ReputationMappingValidationAttempt(
            id=uuid7(),
            run_id=run_id,
            vehicle_id=target.vehicle_id,
            attempt_number=attempt_number,
            mapping_hash=target.mapping_hash,
            contract_version=spec.validation_contract_version,
            adapter_version=spec.adapter_version,
            status="failed",
            metrics={},
            gate_results={},
            error_code=error["code"],
            error_message=error["message"],
            finished_at=finished,
        )

    def get_mapping_validation(self, run_id: str, prefix: str = "/api/v1") -> dict[str, Any]:
        with self.sessions() as db:
            run = db.get(ReputationMappingValidationRun, run_id)
            if not run:
                raise DomainError(
                    "REPUTATION_VALIDATION_NOT_FOUND", "映射验证运行不存在。", status_code=404
                )
            attempts = db.scalars(
                select(ReputationMappingValidationAttempt)
                .where(ReputationMappingValidationAttempt.run_id == run_id)
                .order_by(
                    ReputationMappingValidationAttempt.vehicle_id,
                    ReputationMappingValidationAttempt.attempt_number,
                )
            ).all()
            return {
                "id": run.id,
                "platform_code": run.platform_code,
                "status": run.status,
                "requested_count": run.requested_count,
                "succeeded_count": run.succeeded_count,
                "failed_count": run.failed_count,
                "concurrency": run.concurrency,
                "started_at": run.started_at.isoformat(),
                "finished_at": run.finished_at.isoformat() if run.finished_at else None,
                "attempts": [
                    {
                        "id": item.id,
                        "vehicle_id": item.vehicle_id,
                        "attempt_number": item.attempt_number,
                        "status": item.status,
                        "actual_name": item.actual_name,
                        "metrics": item.metrics,
                        "error_code": item.error_code,
                        "error_message": item.error_message,
                        "duration_ms": item.duration_ms,
                        "full_page_url": f"{prefix}/reputation/mapping-validations/attempts/{item.id}/full"
                        if item.full_page_path
                        else None,
                        "metric_region_url": f"{prefix}/reputation/mapping-validations/attempts/{item.id}/metric"
                        if item.metric_region_path
                        else None,
                    }
                    for item in attempts
                ],
                "scope": self.get_scope(),
            }

    def get_mapping_validation_evidence(self, attempt_id: str, kind: str) -> Path:
        with self.sessions() as db:
            attempt = db.get(ReputationMappingValidationAttempt, attempt_id)
            if not attempt:
                raise DomainError(
                    "REPUTATION_VALIDATION_ATTEMPT_NOT_FOUND",
                    "映射验证尝试不存在。",
                    status_code=404,
                )
            raw = attempt.full_page_path if kind == "full" else attempt.metric_region_path
            path = Path(raw) if raw else None
            expected = attempt.full_page_sha256 if kind == "full" else attempt.metric_region_sha256
        if not path or not path.is_file() or _sha256(path) != expected:
            raise DomainError(
                "REPUTATION_VALIDATION_EVIDENCE_MISSING",
                "映射验证证据不存在或校验失败。",
                status_code=404,
            )
        return path

    def publish_preview(self) -> dict[str, Any]:
        scope = self.get_scope()
        if not scope["initialized"]:
            raise DomainError("REPUTATION_SCOPE_UNINITIALIZED", "请先初始化口碑车型范围。")
        vehicles = [row for row in scope["vehicles"] if row.get("enabled", True)]
        platform_codes = [
            code
            for code in REPUTATION_PLATFORMS
            if any(vehicle.get("mappings", {}).get(code) for vehicle in vehicles)
        ]
        verified = sum(
            self._mapping_is_verified(vehicle, code)
            for vehicle in vehicles
            for code in platform_codes
        )
        with self.sessions() as db:
            draft = db.get(ReputationScopeDraft, "current")
            published = (
                db.get(ReputationScopeVersion, draft.published_version_id)
                if draft and draft.published_version_id
                else None
            )
            previous = [
                row
                for row in ((published.snapshot or {}).get("vehicles", []) if published else [])
                if row.get("enabled", True)
            ]
        current_by_id = {str(row["id"]): row for row in vehicles}
        previous_by_id = {str(row["id"]): row for row in previous}
        added_ids = current_by_id.keys() - previous_by_id.keys()
        disabled_ids = previous_by_id.keys() - current_by_id.keys()

        def business_mapping(row: dict[str, Any]) -> dict[str, Any]:
            return {
                code: {
                    key: row.get("mappings", {}).get(code, {}).get(key)
                    for key in ("platform_vehicle_id", "platform_url", "platform_display_name")
                }
                for code in platform_codes
            }

        shared_ids = current_by_id.keys() & previous_by_id.keys()
        role_changed = sum(
            current_by_id[item].get("role") != previous_by_id[item].get("role")
            or current_by_id[item].get("role_order") != previous_by_id[item].get("role_order")
            for item in shared_ids
        )
        mapping_changed = sum(
            business_mapping(current_by_id[item]) != business_mapping(previous_by_id[item])
            for item in shared_ids
        )
        identity_changed = sum(
            any(
                current_by_id[item].get(key) != previous_by_id[item].get(key)
                for key in ("series_name", "vehicle_name")
            )
            for item in shared_ids
        )
        has_changes = published is None or bool(
            added_ids
            or disabled_ids
            or role_changed
            or mapping_changed
            or identity_changed
        )
        expected_mapping_count = len(vehicles) * len(platform_codes)
        complete_mapping_count = sum(
            bool(vehicle.get("mappings", {}).get(code))
            for vehicle in vehicles
            for code in platform_codes
        )
        verified_all = bool(expected_mapping_count) and verified == expected_mapping_count
        can_publish = verified_all and has_changes
        return {
            "revision": scope["revision"],
            "initial_publish": scope["published_version"] is None,
            "vehicle_count": len(vehicles),
            "focus_count": sum(row["role"] == "focus" for row in vehicles),
            "competitor_count": sum(row["role"] == "competitor" for row in vehicles),
            "verified_mapping_count": verified,
            "expected_mapping_count": expected_mapping_count,
            "complete_mapping_count": complete_mapping_count,
            "platform_codes": platform_codes,
            "added_count": len(added_ids),
            "disabled_count": len(disabled_ids),
            "role_changed_count": role_changed,
            "mapping_changed_count": mapping_changed,
            "identity_changed_count": identity_changed,
            "has_changes": has_changes,
            "can_publish": can_publish,
            "warning": (
                None
                if can_publish
                else "当前范围没有待发布变更。"
                if verified_all
                else "真实页面验证尚未全部完成，当前不会开放发布。"
            ),
        }

    @staticmethod
    def _mapping_is_verified(vehicle: dict[str, Any], platform_code: str) -> bool:
        spec = REPUTATION_PLATFORMS[platform_code]
        mapping = vehicle.get("mappings", {}).get(platform_code, {})
        return (
            mapping.get("validation_status") == "verified"
            and mapping.get("validation_contract_version") == spec.validation_contract_version
            and mapping.get("validated_mapping_hash")
            == _mapping_hash(str(vehicle.get("id") or ""), mapping, platform_code)
        )

    def publish_scope(self, value: ScopePublishRequest) -> dict[str, Any]:
        preview = self.publish_preview()
        if value.revision != preview["revision"]:
            raise DomainError(
                "REPUTATION_SCOPE_CONFLICT",
                "范围草稿已经变化，请重新预览。",
                status_code=409,
            )
        if not preview["can_publish"]:
            raise DomainError("REPUTATION_SCOPE_NOT_VERIFIED", "全部映射验证通过后才能发布。")
        if preview["initial_publish"] and not value.initial_review_acknowledged:
            raise DomainError("REPUTATION_INITIAL_REVIEW_REQUIRED", "请先确认首发全量复核。")
        with self.sessions.begin() as db:
            draft = db.get(ReputationScopeDraft, "current")
            if not draft or draft.revision != value.revision:
                raise DomainError(
                    "REPUTATION_SCOPE_CONFLICT",
                    "范围草稿已经变化，请重新预览。",
                    status_code=409,
                )
            next_version = (db.scalar(select(func.max(ReputationScopeVersion.version))) or 0) + 1
            version = ReputationScopeVersion(
                version=next_version,
                snapshot=json.loads(json.dumps(draft.data, ensure_ascii=False)),
                source_revision=draft.revision,
            )
            db.add(version)
            db.flush()
            draft.published_version_id = version.id
        return self.get_scope()

    def _published_version_dict(
        self, db: Session, draft: ReputationScopeDraft
    ) -> dict[str, Any] | None:
        if not draft.published_version_id:
            return None
        version = db.get(ReputationScopeVersion, draft.published_version_id)
        return (
            {
                "id": version.id,
                "version": version.version,
                "published_at": version.published_at.isoformat(),
            }
            if version
            else None
        )

    def _create_evidence(
        self, run_dir: Path, result: ReputationResult, metrics: dict[str, Any]
    ) -> ReputationEvidence:
        evidence_id = uuid7()
        evidence_dir = run_dir / "evidence" / result.vehicle_id
        evidence_dir.mkdir(parents=True, exist_ok=True)
        metric_path = evidence_dir / "region.png"
        self._draw_fixture(metric_path, result, metrics, size=(980, 220), compact=True)
        digest = _sha256(metric_path)
        return ReputationEvidence(
            id=evidence_id,
            result_id=result.id,
            full_page_path=str(metric_path),
            metric_region_path=str(metric_path),
            full_page_sha256=digest,
            metric_region_sha256=digest,
            width=980,
            height=220,
        )

    @staticmethod
    def _draw_fixture(
        path: Path,
        result: ReputationResult,
        metrics: dict[str, Any],
        *,
        size: tuple[int, int],
        compact: bool,
    ) -> None:
        image = Image.new("RGB", size, "#F8FAFC")
        draw = ImageDraw.Draw(image)
        draw.rounded_rectangle(
            (28, 24, size[0] - 28, size[1] - 24), radius=22, fill="#FFFFFF", outline="#CBD5E1", width=2
        )
        draw.rectangle((28, 24, size[0] - 28, 82), fill="#1E3A8A")
        draw.text((52, 44), f"SYNTHETIC EVIDENCE / {result.vehicle_id}", fill="white")
        labels = (("SCORE", "score"), ("RANK", "rank"), ("VOLUME", "volume"))
        top = 112 if compact else 180
        box_height = 76 if compact else 150
        gap = 20
        width = (size[0] - 104 - gap * 2) // 3
        for index, (label, key) in enumerate(labels):
            left = 52 + index * (width + gap)
            draw.rounded_rectangle(
                (left, top, left + width, top + box_height),
                radius=16,
                fill="#EFF6FF",
                outline="#BFDBFE",
                width=2,
            )
            value = metrics[key].get("raw") or metrics[key].get("comparison_status") or "-"
            draw.text((left + 18, top + 18), label, fill="#475569")
            draw.text((left + 18, top + 44), str(value), fill="#0F172A")
        if not compact:
            draw.text((52, 112), "Full-page synthetic fixture. No platform request was issued.", fill="#64748B")
            draw.text((52, size[1] - 64), "ThreadSnap reputation inspection acceptance fixture", fill="#64748B")
        image.save(path, format="PNG", optimize=False)

    def _render_report(
        self,
        run: ReputationRun,
        results: list[ReputationResult],
        evidence_result_ids: set[str] | None = None,
    ) -> str:
        prefix = "【不完整汇报】" if run.status == "partial_success" else ""
        title = f"{prefix}{run.planned_date}口碑巡检指标变动如下："
        lines = [title]
        if run.run_type == "baseline_initialization":
            for platform_code in run.platform_codes:
                platform_results = [
                    item for item in results if item.platform_code == platform_code
                ]
                lines.extend(
                    [
                        "",
                        f"【{REPUTATION_PLATFORMS[platform_code].display_name}】",
                        "首次基线初始化，无前日变化可比较。",
                        "页面证据："
                        f"{sum(item.id in evidence_result_ids for item in platform_results) if evidence_result_ids is not None else sum(bool(item.evidence_required and item.status == 'success') for item in platform_results)}"
                        f"/{len(platform_results)}",
                    ]
                )
            if run.schedule_type == "month_end":
                lines.append("月末巡检：本批次同时作为首次基线并执行当前范围全量页面证据。")
            missing = [item for item in results if item.status != "success"]
            if missing:
                lines.extend(
                    ["", "异常与缺失："]
                    + [f"- {item.platform_name}/{item.vehicle_name}：{item.error_message or item.status}" for item in missing]
                )
            return "\n".join(lines) + "\n"
        for platform_code in run.platform_codes:
            lines.extend(["", f"【{REPUTATION_PLATFORMS[platform_code].display_name}】"])
            platform_changed = 0
            platform_anomalies: list[str] = []
            for result in [item for item in results if item.platform_code == platform_code]:
                changes: list[str] = []
                for key, name in (
                ("score", "口碑分"),
                ("rank", "排名"),
                ("volume", "口碑量"),
                ("review_article_count", "口碑评价篇数"),
                ("negative_rate", "差评率"),
            ):
                    metric = result.metrics.get(key)
                    if not metric:
                        continue
                    if metric.get("direction") in {"up", "down"}:
                        direction = "上升" if metric["direction"] == "up" else "下降"
                        changes.append(
                            f"{name}{metric['raw']}，较昨日{direction}{str(metric['delta']).lstrip('+-')}"
                        )
                if changes:
                    platform_changed += 1
                    lines.append(f"{platform_changed}. 【{result.vehicle_name}】" + "；".join(changes) + "。")
                if result.status != "success" or any(
                    item.get("comparison_status") not in {"comparable"}
                    for item in result.metrics.values()
                ):
                    states = sorted(
                        {
                            item.get("comparison_status", "unknown")
                            for item in result.metrics.values()
                            if item.get("comparison_status") != "comparable"
                        }
                    )
                    platform_anomalies.append(
                        f"- {result.vehicle_name}：{result.error_message or '、'.join(states)}"
                    )
            if not platform_changed and not platform_anomalies:
                lines.append("今日无口碑指标变化。")
            elif not platform_changed:
                lines.append("今日没有可确认的正常变化。")
            if platform_anomalies:
                lines.extend(["", "异常与缺失：", *platform_anomalies])
        if run.schedule_type == "month_end" or run.run_type == "month_end":
            lines.extend(["", "月末巡检：已执行当前范围全量页面证据。"])
        return "\n".join(lines) + "\n"

    def _create_xlsx(
        self,
        run: ReputationRun,
        results: list[ReputationResult],
        evidence_by_result: dict[str, ReputationEvidence],
        path: Path,
    ) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "口碑巡检"
        platform_codes = list(run.platform_codes or [PLATFORM_CODE])
        single_platform = len(platform_codes) == 1
        metric_headers = ("口碑分", "排名", "口碑量", "口碑评价篇数", "差评率")
        headers = ["日期", "角色", "车系", "车型"] + [
            label if single_platform else f"{REPUTATION_PLATFORMS[code].display_name}-{label}"
            for code in platform_codes
            for label in metric_headers
        ] + ["备注"]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        vehicles = list({result.vehicle_id: result for result in results}.values())
        by_target = {(result.vehicle_id, result.platform_code): result for result in results}
        preview_dir = path.parent / "xlsx-previews"
        preview_manifest: list[dict[str, Any]] = []
        for row_index, vehicle in enumerate(vehicles, start=2):
            values = [
                run.planned_date,
                "重点车型" if vehicle.role == "focus" else "竞品车型",
                vehicle.series_name,
                vehicle.vehicle_name,
            ]
            for code in platform_codes:
                result = by_target.get((vehicle.vehicle_id, code))
                values.extend(
                    [
                        result.metrics["score"].get("raw") or "—" if result else "—",
                        result.metrics["rank"].get("raw") or "—" if result else "—",
                        result.metrics["volume"].get("raw") or "—" if result else "—",
                        result.metrics.get("review_article_count", {}).get("raw") or "—" if result else "—",
                        result.metrics.get("negative_rate", {}).get("raw") or "—" if result else "—",
                    ]
                )
            missing: list[str] = []
            for code in platform_codes:
                result = by_target.get((vehicle.vehicle_id, code))
                evidence = evidence_by_result.get(result.id) if result else None
                if not evidence:
                    reason = result.error_message if result and result.error_message else "证据缺失"
                    missing.append(f"{REPUTATION_PLATFORMS[code].display_name}：{reason}")
            values.append("；".join(missing))
            sheet.append(values)
            for platform_index, code in enumerate(platform_codes):
                result = by_target.get((vehicle.vehicle_id, code))
                first_column = 5 + platform_index * 5
                for offset, metric_name in enumerate(
                    ("score", "rank", "volume", "review_article_count", "negative_rate")
                ):
                    tone = result.metrics.get(metric_name, {}).get("tone") if result else None
                    cell = sheet.cell(row_index, first_column + offset)
                    cell.fill = GREEN_FILL if tone == "positive" else RED_FILL if tone == "negative" else NEUTRAL_FILL
                    cell.alignment = Alignment(horizontal="center")
            preview_path, preview_record = self._xlsx_preview(
                vehicle.vehicle_id,
                platform_codes,
                by_target,
                evidence_by_result,
                preview_dir,
            )
            if preview_path:
                preview = WorksheetImage(preview_path)
                if single_platform:
                    preview.width, preview.height = 294, 66
                    sheet.row_dimensions[row_index].height = 52
                else:
                    preview.width, preview.height = 720, 135
                    sheet.row_dimensions[row_index].height = 104
                note_column = 5 + len(platform_codes) * 5
                sheet.add_image(preview, f"{get_column_letter(note_column)}{row_index}")
            if preview_record:
                preview_manifest.append(preview_record)
        widths = [13, 12, 18, 22] + [
            value for _ in platform_codes for value in (12, 12, 14, 16, 12)
        ] + [105 if not single_platform else 44]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.freeze_panes = "E2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(vehicles) + 1}"
        workbook.save(path)
        if preview_manifest:
            (preview_dir / "manifest.json").write_text(
                json.dumps(
                    {
                        "schema_version": "reputation-xlsx-preview-v1",
                        "run_id": run.id,
                        "items": preview_manifest,
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )

    @staticmethod
    def _xlsx_preview(
        vehicle_id: str,
        platform_codes: list[str],
        by_target: dict[tuple[str, str], ReputationResult],
        evidence_by_result: dict[str, ReputationEvidence],
        preview_dir: Path,
    ) -> tuple[str | None, dict[str, Any] | None]:
        """为多平台XLSX生成单张可追溯预览，独立证据字节保持不变。"""

        sources: list[dict[str, Any]] = []
        for code in platform_codes:
            result = by_target.get((vehicle_id, code))
            evidence = evidence_by_result.get(result.id) if result else None
            sources.append(
                {
                    "platform_code": code,
                    "platform_name": REPUTATION_PLATFORMS[code].display_name,
                    "result_id": result.id if result else None,
                    "evidence_id": evidence.id if evidence else None,
                    "source_sha256": evidence.metric_region_sha256 if evidence else None,
                    "source_path": evidence.metric_region_path if evidence else None,
                }
            )
        available = [source for source in sources if source["source_path"]]
        if len(platform_codes) == 1:
            return (str(available[0]["source_path"]), None) if available else (None, None)
        if not available:
            return None, None

        input_hash = _text_hash(
            [
                {
                    "platform_code": source["platform_code"],
                    "source_sha256": source["source_sha256"],
                }
                for source in sources
            ]
        )
        preview_dir.mkdir(parents=True, exist_ok=True)
        preview_path = preview_dir / f"{vehicle_id}-{input_hash[:16]}.png"
        if not preview_path.is_file():
            tile_width, content_height, label_height = 320, 180, 28
            canvas = Image.new(
                "RGB", (tile_width * len(platform_codes), content_height + label_height), "white"
            )
            draw = ImageDraw.Draw(canvas)
            try:
                font = ImageFont.truetype("C:/Windows/Fonts/msyh.ttc", 16)
            except OSError:
                font = ImageFont.load_default()
            for index, source in enumerate(sources):
                left = index * tile_width
                draw.rectangle(
                    (left, 0, left + tile_width - 1, label_height - 1),
                    fill="#E8EEF8",
                    outline="#CBD5E1",
                )
                draw.text((left + 10, 4), source["platform_name"], fill="#0F172A", font=font)
                draw.rectangle(
                    (left, label_height, left + tile_width - 1, label_height + content_height - 1),
                    outline="#CBD5E1",
                )
                if source["source_path"]:
                    with Image.open(source["source_path"]) as original:
                        tile = original.convert("RGB")
                        tile.thumbnail((tile_width - 12, content_height - 12), Image.Resampling.LANCZOS)
                    x = left + (tile_width - tile.width) // 2
                    y = label_height + (content_height - tile.height) // 2
                    canvas.paste(tile, (x, y))
                else:
                    draw.text(
                        (left + 94, label_height + 78),
                        "EVIDENCE MISSING",
                        fill="#B91C1C",
                    )
            canvas.save(preview_path, format="PNG", optimize=False)
        record = {
            "vehicle_id": vehicle_id,
            "input_hash": input_hash,
            "preview_path": str(preview_path),
            "preview_sha256": _sha256(preview_path),
            "sources": sources,
        }
        return str(preview_path), record

    @staticmethod
    def _run_dict(run: ReputationRun) -> dict[str, Any]:
        return {
            "id": run.id,
            "number": run.number,
            "source_type": run.source_type,
            "scenario_id": run.scenario_id,
            "run_type": run.run_type,
            "schedule_type": run.schedule_type,
            "planned_date": run.planned_date,
            "root_run_id": run.root_run_id,
            "parent_run_id": run.parent_run_id,
            "scope_version_id": run.scope_version_id,
            "planned_at": run.planned_at.isoformat() if run.planned_at else None,
            "report_planned_at": (
                run.report_planned_at.isoformat() if run.report_planned_at else None
            ),
            "report_generated_at": (
                run.report_generated_at.isoformat() if run.report_generated_at else None
            ),
            "delayed": run.delayed,
            "concurrency": run.concurrency,
            "baseline_date": run.baseline_date,
            "baseline_frozen_at": (
                run.baseline_frozen_at.isoformat() if run.baseline_frozen_at else None
            ),
            "baseline_source_run_id": run.baseline_source_run_id,
            "status": run.status,
            "platform_codes": run.platform_codes,
            "planned_count": run.planned_count,
            "completed_count": run.completed_count,
            "failed_count": run.failed_count,
            "required_evidence_count": run.required_evidence_count,
            "complete_evidence_count": run.complete_evidence_count,
            "report_status": run.report_status,
            "report_attempt_count": run.report_attempt_count,
            "report_text": run.report_text,
            "created_at": run.created_at.isoformat(),
            "started_at": run.started_at.isoformat() if run.started_at else None,
            "finished_at": run.finished_at.isoformat() if run.finished_at else None,
        }

    @staticmethod
    def _result_dict(
        result: ReputationResult,
        evidence: ReputationEvidence | None,
        prefix: str,
    ) -> dict[str, Any]:
        return {
            "id": result.id,
            "vehicle_id": result.vehicle_id,
            "series_name": result.series_name,
            "vehicle_name": result.vehicle_name,
            "role": result.role,
            "role_position": result.role_position,
            "vehicle_position": result.vehicle_position,
            "platform_code": result.platform_code,
            "platform_name": result.platform_name,
            "status": result.status,
            "metrics": result.metrics,
            "evidence_required": result.evidence_required,
            "attempt_count": result.attempt_count,
            "duration_ms": result.duration_ms,
            "error_code": result.error_code,
            "error_message": result.error_message,
            "collected_at": result.collected_at.isoformat(),
            "evidence": {
                "id": evidence.id,
                "full_page_url": f"{prefix}/reputation/evidence/{evidence.id}/full",
                "metric_region_url": f"{prefix}/reputation/evidence/{evidence.id}/metric",
                "full_page_sha256": evidence.full_page_sha256,
                "metric_region_sha256": evidence.metric_region_sha256,
            }
            if evidence
            else None,
        }
